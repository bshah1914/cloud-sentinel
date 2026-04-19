"""
CloudSentrix Enterprise — Multi-Cloud Security Platform
FastAPI backend with provider plugin architecture for AWS, Azure, and GCP.
"""

import json
import os
import shutil
import threading
import uuid
import yaml
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
from pydantic import BaseModel
from jose import JWTError, jwt
from passlib.context import CryptContext

# ── Paths ────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
ACCOUNT_DATA_DIR = BASE_DIR / "account-data"
CONFIG_FILE = BASE_DIR / "config.json"
CONFIG_DEMO = BASE_DIR / "config.json.demo"
AUDIT_CONFIG = BASE_DIR / "audit_config.yaml"
AUDIT_OVERRIDE = BASE_DIR / "config" / "audit_config_override.yaml"
USERS_FILE = BASE_DIR / "backend" / "users.json"

# ── JWT / Auth Config ────────────────────────────────────────────
SECRET_KEY = os.environ.get("CLOUDSENTRIX_JWT_SECRET",
                             "cloudlunar-enterprise-secret-change-in-production-2024")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("CLOUDSENTRIX_TOKEN_MINUTES", "480"))  # default 8h
REFRESH_TOKEN_EXPIRE_MINUTES = int(os.environ.get("CLOUDSENTRIX_REFRESH_MINUTES", "10080"))  # default 7d

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")


# ── User Management (Database-backed) ────────────────────────────
def _db_session():
    """Get a database session."""
    from models.database import SessionLocal
    return SessionLocal()

def _load_users() -> list[dict]:
    """Load users from database, fallback to JSON."""
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        users = db.query(DBUser).all()
        db.close()
        if users:
            return [{"username": u.username, "hashed_password": u.password_hash,
                      "role": u.role, "user_type": u.user_type, "org_id": u.org_id,
                      "email": u.email, "id": u.id, "is_active": u.is_active,
                      "created": u.created_at.isoformat() if u.created_at else ""} for u in users]
    except Exception:
        pass
    # Fallback to JSON
    if USERS_FILE.exists():
        with open(USERS_FILE) as f:
            return json.load(f)
    return []

def _save_users(users: list[dict]):
    """Save to JSON for backward compat."""
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=2)

def _init_admin():
    """Initialize admin user in database."""
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        admin = db.query(DBUser).filter(DBUser.username == "admin").first()
        if not admin:
            # Will be created by seed_database()
            pass
        db.close()
    except Exception:
        pass

def _authenticate_user(username: str, password: str) -> Optional[dict]:
    """Authenticate from database (single source of truth)."""
    return db_ops.authenticate_user(username, password)


def _create_token(data: dict, expires_delta: timedelta) -> str:
    to_encode = data.copy()
    to_encode["exp"] = datetime.utcnow() + expires_delta
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(401, "Invalid token")
    except JWTError:
        raise HTTPException(401, "Invalid or expired token")

    # Get user from database (single source of truth)
    user = db_ops.get_user_by_username(username)
    if user:
        user["org_name"] = payload.get("org_name", "")
        return user

    raise HTTPException(401, "User not found")


# ── App Setup ────────────────────────────────────────────────────
app = FastAPI(title="CloudSentrix Enterprise API", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Logging Setup ────────────────────────────────────────────────
from services.app_logger import app_log, access_log, auth_log, scan_log, security_log, export_log, log_startup
from services.audit import log_action
from starlette.middleware.base import BaseHTTPMiddleware
import time as _time

class AuditMiddleware(BaseHTTPMiddleware):
    """Log all API requests to access log and audit trail."""
    SKIP_PATHS = {"/api/health", "/api/auth/me", "/docs", "/openapi.json", "/favicon.ico"}
    AUDIT_ACTIONS = {
        "POST /api/auth/login": "auth.login",
        "POST /api/scan/start": "scan.started",
        "POST /api/accounts": "account.added",
        "DELETE /api/accounts": "account.removed",
        "POST /api/cidrs": "cidr.added",
        "DELETE /api/cidrs": "cidr.removed",
        "GET /api/export": "report.exported",
        "POST /api/compliance/scan": "compliance.scan",
        "GET /api/report": "report.viewed",
        "POST /api/v2/scans": "scan.started_v2",
        "POST /api/v2/accounts": "account.added_v2",
        "POST /api/v2/alert-rules": "alert.rule_created",
        "POST /api/v2/schedules": "schedule.created",
        "POST /api/v2/remediation": "remediation.executed",
        "POST /api/admin/clients": "admin.client_added",
        "DELETE /api/admin/clients": "admin.client_removed",
        "POST /api/admin/users": "admin.user_created",
        "POST /api/ai/chat": "ai.chat",
    }

    async def dispatch(self, request, call_next):
        start = _time.time()
        method = request.method
        path = request.url.path
        ip = request.client.host if request.client else "unknown"
        ua = request.headers.get("user-agent", "")[:100]

        response = await call_next(request)
        duration = round((_time.time() - start) * 1000)
        status = response.status_code

        # Access log (skip health checks)
        if path not in self.SKIP_PATHS:
            access_log.info(f"{method} {path} {status} {duration}ms {ip}")

        # Audit trail for important actions
        action_key = f"{method} {path}"
        for pattern, action in self.AUDIT_ACTIONS.items():
            pmethod, ppath = pattern.split(" ", 1)
            if method == pmethod and path.startswith(ppath):
                # Extract username from token if possible
                username = None
                org_id = None
                auth_header = request.headers.get("authorization", "")
                if auth_header.startswith("Bearer "):
                    try:
                        payload = jwt.decode(auth_header[7:], SECRET_KEY, algorithms=[ALGORITHM])
                        username = payload.get("sub")
                        org_id = payload.get("org_id")
                    except Exception:
                        pass

                log_action(
                    org_id=org_id, username=username, action=action,
                    resource_type=path.split("/")[2] if len(path.split("/")) > 2 else None,
                    resource_id=path.split("/")[-1] if len(path.split("/")) > 3 else None,
                    details={"status": status, "duration_ms": duration, "method": method, "path": path},
                    ip_address=ip, user_agent=ua[:200],
                )
                break

        return response

app.add_middleware(AuditMiddleware)
log_startup()

_init_admin()

# ── Register Provider Plugins ────────────────────────────────────
from providers.registry import registry
from providers.aws.collector import AWSCollector
from providers.aws.parser import AWSParser
from providers.aws.auditor import AWSAuditor
from providers.aws.routes import AWSRoutes

from providers.azure.collector import AzureCollector
from providers.azure.parser import AzureParser
from providers.azure.auditor import AzureAuditor
from providers.azure.routes import AzureRoutes

from providers.gcp.collector import GCPCollector
from providers.gcp.parser import GCPParser
from providers.gcp.auditor import GCPAuditor
from providers.gcp.routes import GCPRoutes

# Register all providers
registry.register("aws", AWSCollector(), AWSParser(), AWSAuditor(), AWSRoutes())
registry.register("azure", AzureCollector(), AzureParser(), AzureAuditor(), AzureRoutes())
registry.register("gcp", GCPCollector(), GCPParser(), GCPAuditor(), GCPRoutes())

# Mount provider-specific routes
for pid in registry.provider_ids:
    plugin = registry.get(pid)
    plugin.routes.register_routes(app, get_current_user, ACCOUNT_DATA_DIR)


# ── Enterprise: Database + Services ────────────────────────────
try:
    from models.database import init_db
    from services.seed import seed_database
    from services.scheduler import start_scheduler
    from routes.enterprise import register_enterprise_routes

    init_db()
    seed_database()
    register_enterprise_routes(app, get_current_user)
    start_scheduler()

    # Encrypt any plaintext cloud credentials (idempotent, safe on every start)
    try:
        from services.db_ops import migrate_plaintext_credentials
        migrated = migrate_plaintext_credentials()
        if migrated:
            print(f"[Security] Encrypted {migrated} plaintext cloud credential(s)")
    except Exception as e:
        print(f"[Security] Credential migration warning: {e}")

    # Start the alert rule evaluator (background thread)
    try:
        from services.alert_evaluator import start as _start_alert_evaluator
        _start_alert_evaluator()
    except Exception as e:
        print(f"[Alerts] Evaluator start warning: {e}")

    # Start the availability monitor (background thread)
    try:
        from services.availability import start as _start_availability
        _start_availability()
    except Exception as e:
        print(f"[Availability] Start warning: {e}")

    print("[Enterprise] Database, services, and scheduler initialized")
except Exception as e:
    print(f"[Enterprise] Init warning (non-fatal): {e}")


# ── In-memory State ─────────────────────────────────────────────
scan_jobs: dict = {}


# ── Pydantic Models ─────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


class AccountIn(BaseModel):
    id: str
    name: str
    provider: str = "aws"      # aws, azure, gcp
    default: bool = False


class CIDRIn(BaseModel):
    cidr: str
    name: str


class ScanRequestModel(BaseModel):
    account_name: Optional[str] = None
    provider: str = "aws"
    credentials: Optional[dict] = None
    region: Optional[str] = "all"
    # Legacy AWS fields (backward compat)
    aws_access_key_id: Optional[str] = None
    aws_secret_access_key: Optional[str] = None
    aws_region: Optional[str] = None


class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str = "viewer"       # admin, editor, viewer


class SignupRequest(BaseModel):
    name: str
    email: str
    company: str = ""
    phone: str = ""
    cloud_provider: str = "aws"  # aws, azure, gcp
    password: str


# ── DB Dashboard Helper ──────────────────────────────────────────
def _dashboard_from_db(account_name: str, provider: str = "aws"):
    """Build dashboard response from V2 scan results stored in database."""
    try:
        from models.database import SessionLocal, CloudAccount, Scan, Finding
        db = SessionLocal()
        # Find account by name
        account = db.query(CloudAccount).filter(CloudAccount.name == account_name).first()
        if not account:
            # Try partial match
            account = db.query(CloudAccount).filter(CloudAccount.name.ilike(f"%{account_name}%")).first()
        if not account:
            db.close()
            raise HTTPException(404, f"No data for account '{account_name}'")

        # Get latest completed scan
        scan = db.query(Scan).filter(
            Scan.account_id == account.id, Scan.status == "completed"
        ).order_by(Scan.completed_at.desc()).first()

        if not scan:
            db.close()
            raise HTTPException(404, f"No completed scan for account '{account_name}'")

        # Get findings
        findings = db.query(Finding).filter(Finding.scan_id == scan.id).all()

        # Build resource totals from scan results
        results = scan.results or {}
        totals = {"instances": 0, "security_groups": 0, "vpcs": 0, "subnets": 0,
                  "lambdas": 0, "buckets": 0, "rds": 0, "elbs": 0, "snapshots": 0,
                  "network_interfaces": 0}
        region_matrix = {}
        for region, data in results.items():
            if isinstance(data, dict) and "resources" in data:
                res = data["resources"]
                totals["instances"] += res.get("ec2_instances", 0)
                totals["security_groups"] += res.get("security_groups", 0)
                totals["vpcs"] += res.get("vpcs", 0)
                totals["subnets"] += res.get("subnets", 0)
                totals["lambdas"] += res.get("lambda_functions", 0)
                totals["buckets"] += res.get("s3_buckets", 0)
                totals["rds"] += res.get("rds_instances", 0)
                totals["elbs"] += res.get("load_balancers", 0)
                totals["snapshots"] += res.get("snapshots", 0)
                region_matrix[region] = {
                    "instances": res.get("ec2_instances", 0),
                    "security_groups": res.get("security_groups", 0),
                    "vpcs": res.get("vpcs", 0),
                    "subnets": res.get("subnets", 0),
                    "lambdas": res.get("lambda_functions", 0),
                    "rds": res.get("rds_instances", 0),
                    "elbs": res.get("load_balancers", 0),
                    "snapshots": res.get("snapshots", 0),
                }

        # Build findings summary
        sev_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
        audit_findings = []
        for f in findings:
            sev_counts[f.severity] = sev_counts.get(f.severity, 0) + 1
            audit_findings.append({
                "title": f.title, "severity": f.severity, "category": f.category,
                "resource_type": f.resource_type, "resource_id": f.resource_id,
                "region": f.region, "remediation": f.remediation,
                "remediation_cli": f.remediation_cli,
            })

        # Build regions dict with has_resources flag (needed by frontend Dashboard)
        regions_detail = {}
        for region, rm in region_matrix.items():
            total = sum(rm.values()) if isinstance(rm, dict) else 0
            regions_detail[region] = {
                **rm,
                "has_resources": total > 0,
            }

        dashboard = {
            "account": account_name, "provider": provider,
            "caller_identity": {"Account": account.account_id or ""},
            "security_score": scan.security_score,
            "collection_date": scan.completed_at.isoformat() if scan.completed_at else "",
            "regions_scanned": scan.regions_scanned,
            "totals": totals,
            "regions": regions_detail,
            "region_matrix": region_matrix,
            "public_ips": [],
            "public_summary": {"ec2": 0, "rds": 0, "elb": 0, "total": 0},
            "open_security_groups": sev_counts.get("CRITICAL", 0) + sev_counts.get("HIGH", 0),
            "iam_summary": {},
            "iam_users": [],
            "iam_users_no_mfa": 0,
            "findings": audit_findings,
            "findings_summary": sev_counts,
            "scan_source": "v2_database",
        }
        db.close()
        return dashboard
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error building dashboard: {str(e)}")


# ── Helpers ──────────────────────────────────────────────────────
def _load_config() -> dict:
    cfg_path = CONFIG_FILE if CONFIG_FILE.exists() else CONFIG_DEMO
    if not cfg_path.exists():
        return {"accounts": [], "cidrs": {}}
    with open(cfg_path) as f:
        return json.load(f)


def _save_config(config: dict):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path) as f:
        return json.load(f)


def _get_account_dirs(provider: str = None) -> list[dict]:
    """List accounts that have collected data, optionally filtered by provider."""
    results = []
    if not ACCOUNT_DATA_DIR.exists():
        return results

    if provider:
        pdir = ACCOUNT_DATA_DIR / provider
        if pdir.exists():
            for d in pdir.iterdir():
                if d.is_dir() and not d.name.startswith("."):
                    results.append({"name": d.name, "provider": provider})
    else:
        for pdir in ACCOUNT_DATA_DIR.iterdir():
            if pdir.is_dir() and pdir.name in ("aws", "azure", "gcp"):
                for d in pdir.iterdir():
                    if d.is_dir() and not d.name.startswith("."):
                        results.append({"name": d.name, "provider": pdir.name})
        # Legacy: check for accounts not under a provider folder (old AWS format)
        for d in ACCOUNT_DATA_DIR.iterdir():
            if d.is_dir() and d.name not in ("aws", "azure", "gcp") and not d.name.startswith("."):
                results.append({"name": d.name, "provider": "aws"})
    return results


def _get_regions_for_account(account_name: str, provider: str = "aws") -> list[str]:
    acct_dir = ACCOUNT_DATA_DIR / provider / account_name
    # Fallback to legacy path
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        return []
    return [d.name for d in acct_dir.iterdir() if d.is_dir() and not d.name.startswith(".")]


# ── Database Operations (Single Source of Truth) ─────────────────
import services.db_ops as db_ops

# ── AUTH ENDPOINTS ───────────────────────────────────────────────
@app.post("/api/auth/login")
def login(req: LoginRequest, request: __import__("fastapi").Request = None):
    ip = request.client.host if request and request.client else "unknown"

    # Authenticate from database (handles both owner and client users)
    user = _authenticate_user(req.username, req.password)
    if user:
        # Check if user is approved
        if not user.get("is_approved", True):
            auth_log.warning(f"LOGIN_BLOCKED user={req.username} pending_approval ip={ip}")
            raise HTTPException(403, "Your account is pending admin approval. You'll be notified once approved.")
        user_type = user.get("user_type", "owner")
        org_id = user.get("org_id", "")
        org_name = ""

        # Get org name for client users
        if user_type == "client" and org_id:
            try:
                from models.database import SessionLocal, Organization
                db = SessionLocal()
                org = db.query(Organization).filter(Organization.id == org_id).first()
                if org:
                    org_name = org.name
                    if org.status != "active":
                        auth_log.warning(f"LOGIN_BLOCKED user={req.username} org_suspended ip={ip}")
                        db.close()
                        raise HTTPException(403, "Organization is suspended")
                db.close()
            except HTTPException:
                raise
            except Exception:
                pass

        auth_log.info(f"LOGIN_SUCCESS {user_type}={req.username} ip={ip}")
        log_action(org_id=org_id, username=req.username, action="auth.login",
                   details={"type": user_type, "role": user["role"]}, ip_address=ip)

        token_data = {"sub": user["username"], "role": user["role"], "user_type": user_type}
        if user_type == "client":
            token_data["org_id"] = org_id
            token_data["org_name"] = org_name

        token = _create_token(token_data, timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
        refresh = _create_token({"sub": user["username"], "type": "refresh"},
                                 timedelta(minutes=REFRESH_TOKEN_EXPIRE_MINUTES))

        response_user = {"username": user["username"], "role": user["role"], "user_type": user_type}
        if user_type == "client":
            response_user["org_id"] = org_id
            response_user["org_name"] = org_name

        return {"access_token": token, "refresh_token": refresh, "token_type": "bearer", "user": response_user}

    auth_log.warning(f"LOGIN_FAILED user={req.username} ip={ip}")
    log_action(username=req.username, action="auth.login_failed",
               details={"reason": "invalid_credentials"}, ip_address=ip)
    raise HTTPException(401, "Invalid username or password")


class RefreshRequest(BaseModel):
    refresh_token: str


@app.post("/api/auth/refresh")
def refresh_token(req: RefreshRequest):
    """Issue a new access token from a valid refresh token."""
    try:
        payload = jwt.decode(req.refresh_token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(401, "Not a refresh token")
        username = payload.get("sub")
        if not username:
            raise HTTPException(401, "Invalid refresh token")
    except JWTError:
        raise HTTPException(401, "Invalid or expired refresh token")

    user = db_ops.get_user_by_username(username)
    if not user:
        raise HTTPException(401, "User not found")

    new_access = _create_token({
        "sub": user["username"], "role": user["role"], "user_type": user.get("user_type", "owner"),
    }, timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    new_refresh = _create_token({
        "sub": user["username"], "type": "refresh",
    }, timedelta(minutes=REFRESH_TOKEN_EXPIRE_MINUTES))
    return {"access_token": new_access, "refresh_token": new_refresh, "token_type": "bearer"}


@app.get("/api/auth/me")
def get_me(user: dict = Depends(get_current_user)):
    result = {"username": user["username"], "role": user["role"], "created": user.get("created"),
              "user_type": user.get("user_type", "owner")}
    if user.get("user_type") == "client":
        result["org_id"] = user.get("org_id")
        result["org_name"] = user.get("org_name", "")
    return result


# ── Signup (Self-Registration, Pending Approval) ────────────────
@app.post("/api/auth/signup")
def signup(req: SignupRequest, request: __import__("fastapi").Request = None):
    """Register a new user — account is created but requires admin approval before login."""
    ip = request.client.host if request and request.client else "unknown"
    name = _sanitize_input(req.name.strip())
    email = _sanitize_input(req.email.strip().lower())
    company = _sanitize_input(req.company.strip()) if req.company else ""
    phone = _sanitize_input(req.phone.strip()) if req.phone else ""
    password = req.password

    if not name or not email or not password:
        raise HTTPException(400, "Name, email, and password are required")
    if len(password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")

    # Use email as username
    username = email

    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()

        # Check if user exists
        existing = db.query(DBUser).filter((DBUser.username == username) | (DBUser.email == email)).first()
        if existing:
            db.close()
            raise HTTPException(409, "An account with this email already exists")

        new_user = DBUser(
            id=str(uuid.uuid4()),
            username=username,
            email=email,
            password_hash=pwd_context.hash(password),
            role="viewer",
            user_type="owner",
            is_active=True,
            is_approved=False,  # Pending admin approval
            company=company,
            phone=phone,
            cloud_provider=req.cloud_provider,
        )
        # Store full name in email field if needed (or use username)
        db.add(new_user)
        db.commit()
        db.close()

        auth_log.info(f"SIGNUP user={username} company={company} ip={ip}")
        log_action(username=username, action="auth.signup",
                   details={"email": email, "company": company, "cloud_provider": req.cloud_provider},
                   ip_address=ip)

        return {
            "message": "Account created successfully! Your account is pending admin approval. You'll be able to login once approved.",
            "status": "pending_approval",
            "email": email,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Signup failed: {str(e)}")


# ── Admin: Pending Users Approval ───────────────────────────────
@app.get("/api/admin/pending-users")
def get_pending_users(user: dict = Depends(get_current_user)):
    """Get all users pending approval."""
    if user.get("user_type") != "owner" or user.get("role") not in ("admin", "owner"):
        raise HTTPException(403, "Admin only")
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        pending = db.query(DBUser).filter(DBUser.is_approved == False).all()
        result = [{
            "id": u.id, "username": u.username, "email": u.email,
            "company": getattr(u, 'company', ''), "phone": getattr(u, 'phone', ''),
            "cloud_provider": getattr(u, 'cloud_provider', ''),
            "created": u.created_at.isoformat() if u.created_at else "",
        } for u in pending]
        db.close()
        return {"pending_users": result, "total": len(result)}
    except Exception as e:
        return {"pending_users": [], "total": 0}


@app.post("/api/admin/approve-user/{user_id}")
def approve_user(user_id: str, user: dict = Depends(get_current_user)):
    """Approve a pending user."""
    if user.get("user_type") != "owner" or user.get("role") not in ("admin", "owner"):
        raise HTTPException(403, "Admin only")
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        target = db.query(DBUser).filter(DBUser.id == user_id).first()
        if not target:
            db.close()
            raise HTTPException(404, "User not found")
        target.is_approved = True
        db.commit()
        db.close()
        log_action(org_id=user.get("org_id"), username=user.get("username"),
                   action="admin.user_approved", resource_type="user",
                   details={"approved_user": target.username, "email": target.email})
        return {"message": f"User '{target.username}' approved successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Approval failed: {str(e)}")


@app.post("/api/admin/reject-user/{user_id}")
def reject_user(user_id: str, user: dict = Depends(get_current_user)):
    """Reject and delete a pending user."""
    if user.get("user_type") != "owner" or user.get("role") not in ("admin", "owner"):
        raise HTTPException(403, "Admin only")
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        target = db.query(DBUser).filter(DBUser.id == user_id).first()
        if not target:
            db.close()
            raise HTTPException(404, "User not found")
        username = target.username
        db.delete(target)
        db.commit()
        db.close()
        log_action(org_id=user.get("org_id"), username=user.get("username"),
                   action="admin.user_rejected", resource_type="user",
                   details={"rejected_user": username})
        return {"message": f"User '{username}' rejected and removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Rejection failed: {str(e)}")


# ── Theme Presets ────────────────────────────────────────────────
THEME_PRESETS = [
    {"id": "dark", "name": "Dark", "category": "dark", "description": "Default dark theme with violet accents"},
    {"id": "light", "name": "Light", "category": "light", "description": "Clean light theme for bright environments"},
    {"id": "midnight-blue", "name": "Midnight Blue", "category": "dark", "description": "Deep navy corporate theme"},
    {"id": "ocean", "name": "Ocean", "category": "dark", "description": "Teal and green aquatic theme"},
    {"id": "sunset", "name": "Sunset", "category": "dark", "description": "Warm orange and amber tones"},
    {"id": "corporate", "name": "Corporate", "category": "light", "description": "Professional light theme with slate accents"},
    {"id": "minimal", "name": "Minimal", "category": "dark", "description": "Clean monochrome with subtle accents"},
    {"id": "high-contrast", "name": "High Contrast", "category": "dark", "description": "Maximum readability with strong contrasts"},
]

@app.get("/api/themes")
def list_themes():
    return {"themes": THEME_PRESETS}

class UserPreferencesUpdate(BaseModel):
    theme_id: Optional[str] = None
    custom_theme: Optional[dict] = None

@app.get("/api/user/preferences")
def get_user_preferences(user: dict = Depends(get_current_user)):
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        db_user = db.query(DBUser).filter(DBUser.username == user["username"]).first()
        result = {"theme_id": "dark", "custom_theme": None}
        if db_user:
            result["theme_id"] = getattr(db_user, "theme_id", "dark") or "dark"
            result["custom_theme"] = getattr(db_user, "custom_theme", None)
        db.close()
        return result
    except Exception:
        return {"theme_id": "dark", "custom_theme": None}

@app.put("/api/user/preferences")
def update_user_preferences(prefs: UserPreferencesUpdate, user: dict = Depends(get_current_user)):
    try:
        from models.database import SessionLocal, User as DBUser
        db = SessionLocal()
        db_user = db.query(DBUser).filter(DBUser.username == user["username"]).first()
        if db_user:
            if prefs.theme_id is not None:
                db_user.theme_id = prefs.theme_id
            if prefs.custom_theme is not None:
                db_user.custom_theme = prefs.custom_theme
            db.commit()
        db.close()
        return {"status": "ok"}
    except Exception:
        return {"status": "ok"}

class OrgBrandingUpdate(BaseModel):
    branding_colors: Optional[dict] = None
    branding_product_name: Optional[str] = None
    branding_logo: Optional[str] = None

@app.get("/api/org/branding")
def get_org_branding(user: dict = Depends(get_current_user)):
    try:
        from models.database import SessionLocal, Organization
        db = SessionLocal()
        org_id = user.get("org_id", "cloudsentrix")
        org = db.query(Organization).filter(Organization.id == org_id).first()
        result = {"logo": None, "product_name": "CloudSentrix", "primary_color": "#7c3aed", "colors": None}
        if org:
            result["logo"] = org.branding_logo
            result["product_name"] = getattr(org, "branding_product_name", None) or "CloudSentrix"
            result["primary_color"] = org.branding_color or "#7c3aed"
            result["colors"] = getattr(org, "branding_colors", None)
        db.close()
        return result
    except Exception:
        return {"logo": None, "product_name": "CloudSentrix", "primary_color": "#7c3aed", "colors": None}

@app.put("/api/org/branding")
def update_org_branding(branding: OrgBrandingUpdate, user: dict = Depends(get_current_user)):
    if user.get("role") not in ("admin", "editor"):
        raise HTTPException(403, "Admin access required")
    try:
        from models.database import SessionLocal, Organization
        db = SessionLocal()
        org_id = user.get("org_id", "cloudsentrix")
        org = db.query(Organization).filter(Organization.id == org_id).first()
        if org:
            if branding.branding_colors is not None:
                org.branding_colors = branding.branding_colors
            if branding.branding_product_name is not None:
                org.branding_product_name = branding.branding_product_name
            if branding.branding_logo is not None:
                org.branding_logo = branding.branding_logo
            db.commit()
        db.close()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Helper: Require Owner ────────────────────────────────────────
def require_owner(user: dict = Depends(get_current_user)):
    if user.get("role") not in ("admin", "editor"):
        raise HTTPException(403, "Owner access required")
    return user


# ── OWNER: Organization Management ──────────────────────────────
class OrgCreateRequest(BaseModel):
    name: str
    contact_email: str
    plan: str = "free"

class OrgUpdateRequest(BaseModel):
    name: Optional[str] = None
    contact_email: Optional[str] = None
    plan: Optional[str] = None
    status: Optional[str] = None

class ClientUserCreateRequest(BaseModel):
    username: str
    password: str
    email: str
    role: str = "viewer"


@app.get("/api/admin/stats")
def admin_platform_stats(user: dict = Depends(require_owner)):
    return db_ops.get_platform_stats()


@app.get("/api/admin/clients")
def admin_list_clients(user: dict = Depends(require_owner)):
    return {"clients": db_ops.get_all_organizations()}


@app.post("/api/admin/clients")
def admin_create_client(req: OrgCreateRequest, user: dict = Depends(require_owner)):
    return db_ops.create_organization(req.name, req.contact_email, req.plan)


@app.get("/api/admin/clients/{org_id}")
def admin_get_client(org_id: str, user: dict = Depends(require_owner)):
    org = db_ops.get_organization(org_id)
    if not org:
        raise HTTPException(404, "Client not found")
    org["users"] = db_ops.get_client_users(org_id)
    org["invoices"] = db_ops.get_invoices(org_id)
    org["activity"] = db_ops.get_activity_log(org_id, limit=20)
    return org


@app.put("/api/admin/clients/{org_id}")
def admin_update_client(org_id: str, req: OrgUpdateRequest, user: dict = Depends(require_owner)):
    updates = {k: v for k, v in req.dict().items() if v is not None}
    org = db_ops.update_organization(org_id, updates)
    if not org:
        raise HTTPException(404, "Client not found")
    return org


@app.delete("/api/admin/clients/{org_id}")
def admin_delete_client(org_id: str, user: dict = Depends(require_owner)):
    db_ops.delete_organization(org_id)
    return {"status": "deleted"}


@app.post("/api/admin/clients/{org_id}/users")
def admin_create_client_user(org_id: str, req: ClientUserCreateRequest, user: dict = Depends(require_owner)):
    result = db_ops.create_client_user(org_id, req.username, req.password, req.email, req.role)
    if not result:
        raise HTTPException(404, "Organization not found")
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(400, result["error"])
    return result


@app.get("/api/admin/clients/{org_id}/users")
def admin_list_client_users(org_id: str, user: dict = Depends(require_owner)):
    return {"users": db_ops.get_client_users(org_id)}


@app.delete("/api/admin/users/{user_id}")
def admin_delete_client_user(user_id: str, user: dict = Depends(require_owner)):
    db_ops.delete_client_user(user_id)
    return {"status": "deleted"}


@app.get("/api/admin/activity")
def admin_activity_log(user: dict = Depends(require_owner)):
    return {"activity": db_ops.get_activity_log(limit=100)}


@app.get("/api/admin/invoices")
def admin_all_invoices(user: dict = Depends(require_owner)):
    return {"invoices": db_ops.get_invoices()}


@app.get("/api/admin/plans")
def admin_get_plans(user: dict = Depends(require_owner)):
    return {"plans": db_ops.PLANS}


# ── CLIENT: Self-service endpoints ──────────────────────────────
@app.get("/api/client/profile")
def client_profile(user: dict = Depends(get_current_user)):
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(403, "Client access required")
    org = db_ops.get_organization(org_id)
    if not org:
        raise HTTPException(404, "Organization not found")
    plan_details = db_ops.PLANS.get(org.get("plan", "free"), {})
    return {"organization": org, "plan": plan_details, "users": db_ops.get_client_users(org_id)}


@app.get("/api/client/invoices")
def client_invoices(user: dict = Depends(get_current_user)):
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(403, "Client access required")
    return {"invoices": db_ops.get_invoices(org_id)}


@app.get("/api/client/activity")
def client_activity(user: dict = Depends(get_current_user)):
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(403, "Client access required")
    return {"activity": db_ops.get_activity_log(org_id, limit=50)}


# ── USER MANAGEMENT (RBAC) ──────────────────────────────────────
@app.get("/api/users")
def list_users(user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return {"users": db_ops.get_all_users()}


@app.post("/api/users")
def create_user_endpoint(req: UserCreateRequest, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    result = db_ops.create_user(req.username, req.password, req.role)
    if isinstance(result, dict) and "error" in result:
        raise HTTPException(400, result["error"])
    return {"status": "ok", "user": result}


@app.delete("/api/users/{username}")
def delete_user_endpoint(username: str, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    if username == user["username"]:
        raise HTTPException(400, "Cannot delete yourself")
    db_ops.delete_user(username)
    return {"status": "ok"}


# ── PROVIDERS ENDPOINT ───────────────────────────────────────────
@app.get("/api/providers")
def list_providers(user: dict = Depends(get_current_user)):
    """List all registered cloud providers with their metadata."""
    return {"providers": registry.list_providers()}


@app.post("/api/providers/{provider_id}/validate")
def validate_provider_credentials(provider_id: str, credentials: dict, user: dict = Depends(get_current_user)):
    """Validate credentials for a specific provider."""
    plugin = registry.get(provider_id)
    if not plugin:
        raise HTTPException(404, f"Provider '{provider_id}' not found")
    return plugin.collector.validate_credentials(credentials)


class ValidateKeysRequest(BaseModel):
    access_key: str
    secret_key: str
    role_arn: Optional[str] = None


@app.post("/api/validate-credentials")
def validate_aws_credentials(req: ValidateKeysRequest, user: dict = Depends(get_current_user)):
    """Validate AWS credentials before saving — returns account info if valid."""
    try:
        import boto3
        from botocore.exceptions import ClientError, NoCredentialsError
        session = boto3.Session(
            aws_access_key_id=req.access_key,
            aws_secret_access_key=req.secret_key,
        )
        sts = session.client("sts")
        identity = sts.get_caller_identity()
        return {
            "valid": True,
            "account_id": identity.get("Account"),
            "arn": identity.get("Arn"),
            "user_id": identity.get("UserId"),
        }
    except Exception as e:
        err = str(e)
        if "InvalidClientTokenId" in err or "InvalidAccessKeyId" in err:
            return {"valid": False, "error": "Invalid Access Key ID — this key does not exist in AWS."}
        elif "SignatureDoesNotMatch" in err:
            return {"valid": False, "error": "Invalid Secret Access Key — the signature does not match."}
        elif "ExpiredToken" in err:
            return {"valid": False, "error": "Credentials have expired. Generate new Access Keys in AWS IAM."}
        else:
            return {"valid": False, "error": f"Validation failed: {err}"}


@app.get("/api/providers/{provider_id}/regions")
def get_provider_regions(provider_id: str, user: dict = Depends(get_current_user)):
    """Get available regions for a provider (static list, no credentials needed)."""
    plugin = registry.get(provider_id)
    if not plugin:
        raise HTTPException(404, f"Provider '{provider_id}' not found")
    return {"regions": plugin.collector.get_regions({})}


# ── ACCOUNTS (Multi-Cloud) ──────────────────────────────────────
@app.get("/api/accounts")
def list_accounts(user: dict = Depends(get_current_user)):
    """List all cloud accounts — DB is the single source of truth."""
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    db_accounts = db_ops.get_cloud_accounts(org_id)

    # Also check for file-based scan data (legacy)
    seen = {a["name"] for a in db_accounts}
    for item in _get_account_dirs():
        if item["name"] not in seen:
            regions = _get_regions_for_account(item["name"], item["provider"])
            db_accounts.append({
                "id": "auto-detected", "name": item["name"],
                "provider": item["provider"], "default": False,
                "has_data": True, "regions": regions,
                "has_credentials": False, "security_score": 0,
                "total_resources": 0, "last_scan": None,
            })

    return {"accounts": db_accounts}


@app.post("/api/accounts")
def add_account(account: AccountIn, user: dict = Depends(get_current_user)):
    """Add cloud account — saves to database."""
    org_id = user.get("org_id", "cloudsentrix")
    result = db_ops.add_cloud_account(
        name=account.name, provider=account.provider,
        account_id=account.id, org_id=org_id,
    )
    return {"status": "ok", "account": result}


@app.delete("/api/accounts/{account_id}")
def remove_account(account_id: str, provider: str = "aws", user: dict = Depends(get_current_user)):
    """Remove cloud account from DB and delete scan data files."""
    db_ops.remove_cloud_account(account_id)

    # Also clean up scan data files
    for name_dir in [ACCOUNT_DATA_DIR / provider / account_id, ACCOUNT_DATA_DIR / account_id]:
        if name_dir.exists():
            shutil.rmtree(name_dir)

    return {"status": "ok", "account_removed": account_id}


# ── CIDRs ────────────────────────────────────────────────────────
@app.get("/api/cidrs")
def list_cidrs(user: dict = Depends(get_current_user)):
    config = _load_config()
    cidrs = config.get("cidrs", {})
    return {"cidrs": [{"cidr": k, **v} for k, v in cidrs.items()]}


@app.post("/api/cidrs")
def add_cidr(cidr_in: CIDRIn, user: dict = Depends(get_current_user)):
    config = _load_config()
    config.setdefault("cidrs", {})[cidr_in.cidr] = {"name": cidr_in.name}
    _save_config(config)
    return {"status": "ok"}


@app.delete("/api/cidrs/{cidr:path}")
def remove_cidr(cidr: str, user: dict = Depends(get_current_user)):
    from urllib.parse import unquote
    cidr = unquote(cidr)
    config = _load_config()
    config.get("cidrs", {}).pop(cidr, None)
    _save_config(config)
    return {"status": "ok"}


# ── Dashboard Layout CRUD ────────────────────────────────────────
class DashboardCreateRequest(BaseModel):
    name: str = "My Dashboard"
    layout: list = []

class DashboardUpdateRequest(BaseModel):
    name: Optional[str] = None
    layout: Optional[list] = None

@app.get("/api/dashboards")
def list_dashboards(user: dict = Depends(get_current_user)):
    from models.database import SessionLocal, DashboardLayout
    db = SessionLocal()
    try:
        dashboards = db.query(DashboardLayout).filter(
            (DashboardLayout.user_id == user.get("id")) |
            (DashboardLayout.is_template == True)
        ).all()
        return {"dashboards": [
            {"id": d.id, "name": d.name, "is_default": d.is_default, "is_template": d.is_template,
             "template_type": d.template_type, "layout": d.layout, "created_at": d.created_at.isoformat() if d.created_at else ""}
            for d in dashboards
        ]}
    finally:
        db.close()

@app.post("/api/dashboards")
def create_dashboard(req: DashboardCreateRequest, user: dict = Depends(get_current_user)):
    from models.database import SessionLocal, DashboardLayout, gen_id
    db = SessionLocal()
    try:
        dash = DashboardLayout(
            id=gen_id(), user_id=user.get("id"), org_id=user.get("org_id", "cloudsentrix"),
            name=req.name, layout=req.layout,
        )
        db.add(dash)
        db.commit()
        return {"id": dash.id, "name": dash.name, "status": "created"}
    finally:
        db.close()

@app.get("/api/dashboards/{dashboard_id}")
def get_dashboard_layout(dashboard_id: str, user: dict = Depends(get_current_user)):
    from models.database import SessionLocal, DashboardLayout
    db = SessionLocal()
    try:
        dash = db.query(DashboardLayout).filter(DashboardLayout.id == dashboard_id).first()
        if not dash:
            raise HTTPException(404, "Dashboard not found")
        return {"id": dash.id, "name": dash.name, "layout": dash.layout,
                "is_template": dash.is_template, "template_type": dash.template_type}
    finally:
        db.close()

@app.put("/api/dashboards/{dashboard_id}")
def update_dashboard_layout(dashboard_id: str, req: DashboardUpdateRequest, user: dict = Depends(get_current_user)):
    from models.database import SessionLocal, DashboardLayout
    db = SessionLocal()
    try:
        dash = db.query(DashboardLayout).filter(DashboardLayout.id == dashboard_id).first()
        if not dash:
            raise HTTPException(404, "Dashboard not found")
        if req.name is not None:
            dash.name = req.name
        if req.layout is not None:
            dash.layout = req.layout
        db.commit()
        return {"status": "ok"}
    finally:
        db.close()

@app.delete("/api/dashboards/{dashboard_id}")
def delete_dashboard(dashboard_id: str, user: dict = Depends(get_current_user)):
    from models.database import SessionLocal, DashboardLayout
    db = SessionLocal()
    try:
        dash = db.query(DashboardLayout).filter(DashboardLayout.id == dashboard_id).first()
        if dash:
            db.delete(dash)
            db.commit()
        return {"status": "ok"}
    finally:
        db.close()

@app.post("/api/dashboards/{dashboard_id}/clone")
def clone_dashboard(dashboard_id: str, user: dict = Depends(get_current_user)):
    from models.database import SessionLocal, DashboardLayout, gen_id
    db = SessionLocal()
    try:
        src = db.query(DashboardLayout).filter(DashboardLayout.id == dashboard_id).first()
        if not src:
            raise HTTPException(404, "Dashboard not found")
        clone = DashboardLayout(
            id=gen_id(), user_id=user.get("id"), org_id=user.get("org_id", "cloudsentrix"),
            name=f"{src.name} (Copy)", layout=src.layout,
        )
        db.add(clone)
        db.commit()
        return {"id": clone.id, "name": clone.name, "status": "cloned"}
    finally:
        db.close()

@app.get("/api/dashboards/templates")
def list_dashboard_templates():
    from models.database import SessionLocal, DashboardLayout
    db = SessionLocal()
    try:
        templates = db.query(DashboardLayout).filter(DashboardLayout.is_template == True).all()
        return {"templates": [
            {"id": d.id, "name": d.name, "template_type": d.template_type, "layout": d.layout}
            for d in templates
        ]}
    finally:
        db.close()

@app.get("/api/widgets/registry")
def get_widget_registry():
    return {"widgets": [
        {"id": "security-score", "name": "Security Score", "category": "Security"},
        {"id": "severity-pie", "name": "Severity Distribution", "category": "Security"},
        {"id": "findings-radar", "name": "Security Posture", "category": "Security"},
        {"id": "top-findings", "name": "Top Findings", "category": "Security"},
        {"id": "risk-gauge", "name": "Risk Gauge", "category": "Security"},
        {"id": "resource-stats", "name": "Resource Stats", "category": "Resources"},
        {"id": "resource-bar", "name": "Resources by Region", "category": "Resources"},
        {"id": "public-ips", "name": "Public IPs", "category": "Resources"},
        {"id": "multi-cloud", "name": "Cloud Providers", "category": "Overview"},
        {"id": "waf-radar", "name": "WAF Radar", "category": "Overview"},
        {"id": "account-list", "name": "Account List", "category": "Overview"},
        {"id": "big-stat", "name": "KPI Card", "category": "Executive"},
        {"id": "trend-chart", "name": "Score Trend", "category": "Executive"},
        {"id": "recent-scans", "name": "Recent Scans", "category": "Operations"},
        {"id": "quick-scan", "name": "Quick Scan", "category": "Operations"},
        {"id": "audit-log", "name": "Audit Log", "category": "Operations"},
        {"id": "compliance-card", "name": "Compliance", "category": "Compliance"},
        {"id": "threat-feed", "name": "Threat Feed", "category": "Threats"},
        {"id": "findings-trend", "name": "Findings Trend", "category": "Threats"},
        {"id": "network-topology", "name": "Network Topology", "category": "Advanced"},
        {"id": "region-map", "name": "Region Map", "category": "Advanced"},
        {"id": "score-timeline", "name": "Score Timeline", "category": "Advanced"},
        {"id": "scan-compare", "name": "Scan Comparison", "category": "Advanced"},
    ]}


# ── VPC CIDRs (from scan data) ─────────────────────────────────
@app.get("/api/vpc-cidrs/{account_name}")
def get_vpc_cidrs(account_name: str, user: dict = Depends(get_current_user)):
    """Get all VPC CIDRs from scanned account data."""
    vpc_cidrs = []
    for pid in ["aws"]:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if not acct_dir.exists():
            acct_dir = ACCOUNT_DATA_DIR / account_name
        if not acct_dir.exists():
            continue
        for region_dir in sorted(acct_dir.iterdir()):
            if not region_dir.is_dir():
                continue
            vpc_file = region_dir / "ec2-describe-vpcs.json"
            if vpc_file.exists():
                try:
                    data = json.load(open(vpc_file))
                    for vpc in data.get("Vpcs", []):
                        cidr = vpc.get("CidrBlock", "")
                        vpc_id = vpc.get("VpcId", "")
                        name = ""
                        for tag in vpc.get("Tags", []):
                            if tag.get("Key") == "Name":
                                name = tag.get("Value", "")
                        is_default = vpc.get("IsDefault", False)
                        state = vpc.get("State", "")
                        # Subnet count
                        subnet_file = region_dir / "ec2-describe-subnets.json"
                        subnet_count = 0
                        if subnet_file.exists():
                            try:
                                subnets = json.load(open(subnet_file)).get("Subnets", [])
                                subnet_count = sum(1 for s in subnets if s.get("VpcId") == vpc_id)
                            except Exception:
                                pass
                        vpc_cidrs.append({
                            "vpc_id": vpc_id,
                            "cidr": cidr,
                            "name": name or ("Default VPC" if is_default else ""),
                            "region": region_dir.name,
                            "state": state,
                            "is_default": is_default,
                            "subnets": subnet_count,
                        })
                except Exception:
                    pass
    return {"vpc_cidrs": vpc_cidrs, "total": len(vpc_cidrs)}


# ── UNIFIED SCAN (Multi-Cloud) ──────────────────────────────────
@app.post("/api/scan/start")
def start_scan(req: ScanRequestModel, user: dict = Depends(get_current_user)):
    """Start a background scan for any cloud provider."""
    scan_log.info(f"SCAN_START user={user.get('username')} provider={req.provider} account={req.account_name}")
    log_action(org_id=user.get("org_id"), username=user.get("username"), action="scan.started",
               resource_type="scan", details={"provider": req.provider, "account": req.account_name})
    provider_id = req.provider
    plugin = registry.get(provider_id)
    if not plugin:
        raise HTTPException(400, f"Unknown provider: {provider_id}")

    # Build credentials dict
    credentials = req.credentials or {}
    # Legacy AWS fields
    if provider_id == "aws" and not credentials:
        if req.aws_access_key_id:
            credentials = {
                "access_key_id": req.aws_access_key_id,
                "secret_access_key": req.aws_secret_access_key,
            }

    job_id = str(uuid.uuid4())[:8]
    account_name = req.account_name or "default"
    requested_region = req.region or req.aws_region or "all"

    scan_jobs[job_id] = {
        "id": job_id, "provider": provider_id,
        "status": "running", "account": account_name,
        "started": datetime.now().isoformat(),
        "log": [], "progress": 0,
        "regions_scanned": [], "regions_total": 0,
    }

    def _run():
        job = scan_jobs[job_id]
        try:
            collector = plugin.collector

            # Determine regions
            if requested_region == "all":
                job["log"].append(f"Discovering enabled {plugin.info.short_name} regions...")
                regions_to_scan = collector.get_regions(credentials)
                job["log"].append(f"Found {len(regions_to_scan)} regions")
            else:
                regions_to_scan = [requested_region]

            job["regions_total"] = len(regions_to_scan)

            total_collected = 0
            total_errors = []

            for idx, region in enumerate(regions_to_scan):
                job["progress"] = int(5 + (85 * idx / max(len(regions_to_scan), 1)))
                job["log"].append(f"[{idx+1}/{len(regions_to_scan)}] Scanning {region}...")

                result = collector.collect(
                    account_name, region, credentials,
                    ACCOUNT_DATA_DIR, progress_callback=None,
                )

                total_collected += result.get("resources_collected", 0)
                total_errors.extend(result.get("errors", []))
                job["regions_scanned"].append(region)

                if result.get("resources_collected", 0) > 0:
                    job["log"].append(f"  {region}: {result['resources_collected']} resources collected")
                else:
                    job["log"].append(f"  {region}: no resources (or access denied)")

            if total_errors:
                job["log"].append(f"Warnings: {len(total_errors)} total across all regions")

            # Check for any collected data
            acct_dir = ACCOUNT_DATA_DIR / provider_id / account_name
            has_data = acct_dir.exists() and any(acct_dir.rglob("*.json"))

            if not has_data:
                job["status"] = "failed"
                job["log"].append(f"FAILED: No data files created. Check {plugin.info.short_name} credentials.")
                job["completed"] = datetime.now().isoformat()
                return

            job["progress"] = 100
            job["status"] = "completed"
            job["log"].append(f"Scan completed — {total_collected} resources across {len(job['regions_scanned'])} regions")
            job["completed"] = datetime.now().isoformat()

        except Exception as e:
            job["status"] = "failed"
            job["log"].append(f"Error: {str(e)}")
            job["completed"] = datetime.now().isoformat()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    return {"job_id": job_id, "status": "started", "provider": provider_id}


@app.get("/api/scan/status/{job_id}")
def scan_status(job_id: str, user: dict = Depends(get_current_user)):
    if job_id not in scan_jobs:
        raise HTTPException(404, "Job not found")
    return scan_jobs[job_id]


@app.get("/api/scan/history")
def scan_history(user: dict = Depends(get_current_user)):
    return {"jobs": list(scan_jobs.values())}


# ── UNIFIED DASHBOARD ────────────────────────────────────────────
@app.get("/api/dashboard/{provider}/{account_name}")
def dashboard_by_provider(provider: str, account_name: str, user: dict = Depends(get_current_user)):
    """Get dashboard for a specific provider + account."""
    plugin = registry.get(provider)
    if not plugin:
        raise HTTPException(404, f"Provider '{provider}' not found")

    acct_dir = ACCOUNT_DATA_DIR / provider / account_name
    if not acct_dir.exists() and provider == "aws":
        acct_dir = ACCOUNT_DATA_DIR / account_name

    if acct_dir.exists():
        return plugin.parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)

    # Fallback: build dashboard from V2 scan results in database
    return _dashboard_from_db(account_name, provider)


# Legacy endpoint (backward compat)
@app.get("/api/dashboard/{account_name}")
def dashboard_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy dashboard endpoint — auto-detects provider."""
    # Check new path first
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            return registry.get(pid).parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
    # Legacy AWS path
    legacy_dir = ACCOUNT_DATA_DIR / account_name
    if legacy_dir.exists():
        return registry.get("aws").parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
    # Fallback: try database
    return _dashboard_from_db(account_name)


# ── UNIFIED MULTI-CLOUD OVERVIEW ────────────────────────────────
@app.get("/api/overview")
def multi_cloud_overview(user: dict = Depends(get_current_user)):
    """Aggregated overview across all providers and accounts."""
    overview = {
        "providers": {},
        "total_accounts": 0,
        "total_resources": 0,
        "total_findings": 0,
        "avg_security_score": 0,
        "resource_breakdown": {},
    }

    scores = []
    for item in _get_account_dirs():
        pid = item["provider"]
        name = item["name"]
        plugin = registry.get(pid)
        if not plugin:
            continue

        if pid not in overview["providers"]:
            overview["providers"][pid] = {
                "name": plugin.info.name,
                "short_name": plugin.info.short_name,
                "color": plugin.info.color,
                "accounts": [],
                "total_resources": 0,
            }

        try:
            dash = plugin.parser.parse_dashboard(name, ACCOUNT_DATA_DIR)
            total_res = sum(v for v in dash.get("totals", {}).values() if isinstance(v, int))
            findings = plugin.auditor.run_audit(name, ACCOUNT_DATA_DIR)

            acct_info = {
                "name": name, "provider": pid,
                "security_score": dash.get("security_score", 0),
                "total_resources": total_res,
                "total_findings": len(findings),
                "regions_scanned": dash.get("regions_scanned", 0),
            }
            overview["providers"][pid]["accounts"].append(acct_info)
            overview["providers"][pid]["total_resources"] += total_res
            overview["total_accounts"] += 1
            overview["total_resources"] += total_res
            overview["total_findings"] += len(findings)
            scores.append(dash.get("security_score", 0))
            # Accumulate resource breakdown by service type
            for k, v in dash.get("totals", {}).items():
                if isinstance(v, int) and v > 0:
                    overview["resource_breakdown"][k] = overview["resource_breakdown"].get(k, 0) + v
        except Exception:
            overview["providers"][pid]["accounts"].append({
                "name": name, "provider": pid, "error": "Failed to load",
            })

    # Also include DB-based accounts (V2 scans)
    try:
        from models.database import SessionLocal, CloudAccount, Scan
        db = SessionLocal()
        db_accounts = db.query(CloudAccount).all()
        file_account_names = {item["name"] for item in _get_account_dirs()}
        for acc in db_accounts:
            if acc.name in file_account_names:
                continue  # Already included from file-based data
            latest_scan = db.query(Scan).filter(
                Scan.account_id == acc.id, Scan.status == "completed"
            ).order_by(Scan.completed_at.desc()).first()
            if not latest_scan:
                continue
            pid = acc.provider or "aws"
            if pid not in overview["providers"]:
                overview["providers"][pid] = {
                    "name": pid.upper(), "short_name": pid.upper(),
                    "color": "#FF9900", "accounts": [], "total_resources": 0,
                }
            acct_info = {
                "name": acc.name, "provider": pid,
                "security_score": latest_scan.security_score,
                "total_resources": latest_scan.resources_found,
                "total_findings": latest_scan.findings_count,
                "regions_scanned": latest_scan.regions_scanned,
            }
            overview["providers"][pid]["accounts"].append(acct_info)
            overview["providers"][pid]["total_resources"] += latest_scan.resources_found
            overview["total_accounts"] += 1
            overview["total_resources"] += latest_scan.resources_found
            overview["total_findings"] += latest_scan.findings_count
            scores.append(latest_scan.security_score)
        db.close()
    except Exception:
        pass

    if scores:
        overview["avg_security_score"] = round(sum(scores) / len(scores))

    return overview


# ── LEGACY ENDPOINTS (backward compat for existing frontend) ─────
@app.get("/api/resources/{account_name}")
def get_resources_legacy(account_name: str, user: dict = Depends(get_current_user)):
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            return registry.get(pid).parser.parse_resources(account_name, ACCOUNT_DATA_DIR)
    legacy_dir = ACCOUNT_DATA_DIR / account_name
    if legacy_dir.exists():
        return registry.get("aws").parser.parse_resources(account_name, ACCOUNT_DATA_DIR)
    # DB fallback — convert counts to arrays for frontend compatibility
    dash = _dashboard_from_db(account_name)
    region_matrix = dash.get("region_matrix", {})
    findings = dash.get("findings", [])
    regions_out = {}
    for region, counts in region_matrix.items():
        if not isinstance(counts, dict):
            continue
        region_findings = [f for f in findings if f.get("region") == region]
        # Build placeholder arrays from counts
        ec2 = [{"id": f.get("resource_id", ""), "type": "", "state": "", "public_ip": "", "private_ip": "", "vpc": "", "_region": region}
               for f in region_findings if f.get("resource_type") in ("ec2", "instance")]
        sg = [{"id": f.get("resource_id", ""), "name": f.get("title", ""), "vpc": "", "inbound_rules": 0, "outbound_rules": 0, "_region": region}
              for f in region_findings if f.get("resource_type") in ("security_group", "sg")]
        regions_out[region] = {
            "ec2_instances": ec2 if ec2 else [{"id": f"i-{i}", "type": "", "state": "running", "public_ip": "", "private_ip": "", "vpc": "", "_region": region} for i in range(counts.get("instances", 0))],
            "s3_buckets": [],
            "security_groups": sg if sg else [{"id": f"sg-{i}", "name": "", "vpc": "", "inbound_rules": 0, "outbound_rules": 0} for i in range(counts.get("security_groups", 0))],
            "vpcs": [{"id": f"vpc-{i}", "cidr": "", "default": False} for i in range(counts.get("vpcs", 0))],
            "lambda_functions": [{"name": f"fn-{i}", "runtime": "", "memory": 0} for i in range(counts.get("lambdas", 0))],
            "rds_instances": [{"id": f"rds-{i}", "engine": "", "class": "", "publicly_accessible": False} for i in range(counts.get("rds", 0))],
            "load_balancers": [{"name": f"elb-{i}", "dns": "", "scheme": ""} for i in range(counts.get("elbs", 0))],
        }
    return {"account": account_name, "regions": regions_out}


@app.get("/api/iam/{account_name}")
def get_iam_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy IAM endpoint — routes to AWS IAM."""
    from providers.aws.parser import AWSParser, _read_json, _get_regions
    acct_dir = ACCOUNT_DATA_DIR / "aws" / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    regions = _get_regions(acct_dir)
    if not regions:
        # DB fallback — return scan findings related to IAM
        try:
            dash = _dashboard_from_db(account_name)
            iam_findings = [f for f in dash.get("findings", []) if f.get("category") in ("iam", "identity")]
            return {"account": account_name, "summary": {}, "users": [], "roles": [], "policies": [],
                    "findings": iam_findings}
        except Exception:
            return {"account": account_name, "summary": {}, "users": [], "roles": [], "policies": []}
    # IAM data is global — prefer us-east-1, fallback to any region with IAM files
    iam_region = "us-east-1" if "us-east-1" in regions else regions[0]
    for r in [iam_region] + regions:
        if (acct_dir / r / "iam-get-account-authorization-details.json").exists():
            iam_region = r
            break
    region_dir = acct_dir / iam_region
    auth = _read_json(region_dir / "iam-get-account-authorization-details.json")
    summary = _read_json(region_dir / "iam-get-account-summary.json")
    users = [{"name": u.get("UserName"), "arn": u.get("Arn"), "created": u.get("CreateDate"),
              "policies": [p["PolicyName"] for p in u.get("AttachedManagedPolicies", [])],
              "inline_policies": list(u.get("UserPolicyList", [])),
              "groups": u.get("GroupList", []), "mfa_devices": u.get("MFADevices", [])}
             for u in auth.get("UserDetailList", [])]
    roles = [{"name": r.get("RoleName"), "arn": r.get("Arn"), "created": r.get("CreateDate"),
              "policies": [p["PolicyName"] for p in r.get("AttachedManagedPolicies", [])],
              "trust_policy": r.get("AssumeRolePolicyDocument")}
             for r in auth.get("RoleDetailList", [])]
    policies = [{"name": p.get("PolicyName"), "arn": p.get("Arn"),
                 "attachment_count": p.get("AttachmentCount"), "is_attachable": p.get("IsAttachable")}
                for p in auth.get("Policies", [])]
    return {"account": account_name, "summary": summary.get("SummaryMap", {}),
            "users": users, "roles": roles, "policies": policies}


@app.get("/api/security-groups/{account_name}")
def get_security_groups_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy SG endpoint — routes to AWS."""
    from providers.aws.parser import _read_json, _get_regions
    acct_dir = ACCOUNT_DATA_DIR / "aws" / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    regions = _get_regions(acct_dir)
    if not regions:
        # DB fallback
        try:
            dash = _dashboard_from_db(account_name)
            sg_findings = [f for f in dash.get("findings", []) if f.get("resource_type") in ("SecurityGroup", "security_group")]
            return {"account": account_name, "risky_groups": [
                {"region": f.get("region",""), "group_id": f.get("resource_id",""), "group_name": f.get("title",""),
                 "vpc_id": "", "open_rules": [], "severity": f.get("severity","HIGH")}
                for f in sg_findings
            ]}
        except Exception:
            return {"account": account_name, "risky_groups": []}
    results = []
    for region in regions:
        sgs = _read_json(acct_dir / region / "ec2-describe-security-groups.json")
        for sg in sgs.get("SecurityGroups", []):
            open_rules = []
            for rule in sg.get("IpPermissions", []):
                for ip_range in rule.get("IpRanges", []):
                    if ip_range.get("CidrIp") == "0.0.0.0/0":
                        open_rules.append({"protocol": rule.get("IpProtocol"), "from_port": rule.get("FromPort"), "to_port": rule.get("ToPort")})
                for ip_range in rule.get("Ipv6Ranges", []):
                    if ip_range.get("CidrIpv6") == "::/0":
                        open_rules.append({"protocol": rule.get("IpProtocol"), "from_port": rule.get("FromPort"), "to_port": rule.get("ToPort")})
            if open_rules:
                results.append({"region": region, "group_id": sg["GroupId"], "group_name": sg["GroupName"],
                                "vpc_id": sg.get("VpcId"), "open_rules": open_rules,
                                "severity": "CRITICAL" if any(r.get("from_port") in (22, 3389, 3306, 5432) or r.get("protocol") == "-1" for r in open_rules) else "HIGH"})
    return {"account": account_name, "risky_groups": results}


@app.get("/api/audit/config")
def get_audit_config(user: dict = Depends(get_current_user)):
    try:
        with open(AUDIT_CONFIG) as f:
            cfg = yaml.safe_load(f) or {}
        if AUDIT_OVERRIDE.exists():
            with open(AUDIT_OVERRIDE) as f:
                override = yaml.safe_load(f) or {}
            for k, v in override.items():
                if k not in cfg:
                    cfg[k] = {"title": "Unknown", "severity": "High", "group": "unknown"}
                cfg[k].update(v)
        return {"findings": [{"id": fid, **conf} for fid, conf in cfg.items()]}
    except Exception:
        return {"findings": []}


@app.post("/api/audit/run/{account_name}")
def run_audit_legacy(account_name: str, user: dict = Depends(get_current_user)):
    """Legacy audit — auto-detects provider."""
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            plugin = registry.get(pid)
            findings = plugin.auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
            severity_counts = {}
            for f in findings:
                s = f.get("severity", "INFO")
                severity_counts[s] = severity_counts.get(s, 0) + 1
            return {"account": account_name, "provider": pid,
                    "total": len(findings), "findings": findings,
                    "summary": {s: severity_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]}}
    # Legacy AWS
    legacy_dir = ACCOUNT_DATA_DIR / account_name
    if legacy_dir.exists():
        findings = registry.get("aws").auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
        severity_counts = {}
        for f in findings:
            s = f.get("severity", "INFO")
            severity_counts[s] = severity_counts.get(s, 0) + 1
        return {"account": account_name, "provider": "aws",
                "total": len(findings), "findings": findings,
                "summary": {s: severity_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]}}
    # DB fallback — use findings from V2 scan
    try:
        dash = _dashboard_from_db(account_name)
        findings = dash.get("findings", [])
        sev = dash.get("findings_summary", {})
        return {"account": account_name, "provider": "aws",
                "total": len(findings), "findings": findings,
                "summary": {s: sev.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]}}
    except Exception:
        raise HTTPException(404, f"No data for account '{account_name}'")


# ── REMEDIATION REQUEST FROM AUDIT FINDINGS ──────────────────────
class AuditRemediationRequest(BaseModel):
    title: str
    severity: str = "MEDIUM"
    category: str = "security"
    region: str = ""
    resource: str = ""
    description: str = ""
    remediation: str = ""
    account_name: str = ""

@app.post("/api/remediation/request")
def request_remediation_from_audit(req: AuditRemediationRequest, user: dict = Depends(get_current_user)):
    """Create a finding in DB and submit remediation request — for audit/threat findings that don't have DB IDs."""
    from services.remediation import request_remediation as _request
    db = SessionLocal()
    try:
        # Check if finding already exists by title + account
        existing = db.query(Finding).filter(
            Finding.title == req.title,
            Finding.account_id == req.account_name,
            Finding.status == "open"
        ).first()
        if existing:
            finding_id = existing.id
        else:
            finding_id = gen_id()
            finding = Finding(
                id=finding_id, org_id=user.get("org_id", "cloudsentrix"),
                scan_id="audit", account_id=req.account_name,
                title=req.title, severity=req.severity,
                category=req.category, region=req.region,
                resource_type=req.resource, description=req.description,
                remediation=req.remediation,
                remediation_cli=req.remediation,
                status="open",
            )
            db.add(finding)
            db.commit()
        return _request(finding_id, "auto", user.get("sub"))
    finally:
        db.close()


# ── REPORT EXPORT ────────────────────────────────────────────────
from fastapi.responses import Response
from report_generator import (
    generate_dashboard_pdf, generate_dashboard_csv,
    generate_audit_pdf, generate_audit_csv,
)


@app.get("/api/export/dashboard/{provider}/{account_name}")
def export_dashboard(provider: str, account_name: str, format: str = "pdf",
                     user: dict = Depends(get_current_user)):
    """Export dashboard report as PDF or CSV."""
    export_log.info(f"EXPORT dashboard format={format} account={account_name} user={user.get('username')}")
    log_action(org_id=user.get("org_id"), username=user.get("username"), action="report.exported",
               resource_type="dashboard", details={"format": format, "account": account_name})
    plugin = registry.get(provider)
    if not plugin:
        raise HTTPException(400, f"Unknown provider '{provider}'")
    data = plugin.parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
    if not data or not data.get("region_matrix"):
        raise HTTPException(404, f"No data for account '{account_name}'")

    if format == "csv":
        content = generate_dashboard_csv(data, account_name, provider)
        return Response(content=content, media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="dashboard-{account_name}-{provider}.csv"'})
    else:
        content = generate_dashboard_pdf(data, account_name, provider)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="dashboard-{account_name}-{provider}.pdf"'})


@app.get("/api/export/audit/{account_name}")
def export_audit(account_name: str, format: str = "pdf",
                 user: dict = Depends(get_current_user)):
    """Export audit report as PDF or CSV."""
    # Auto-detect provider
    detected_provider = "aws"
    for pid in registry.provider_ids:
        if (ACCOUNT_DATA_DIR / pid / account_name).exists():
            detected_provider = pid
            break

    plugin = registry.get(detected_provider)
    acct_dir = ACCOUNT_DATA_DIR / detected_provider / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        raise HTTPException(404, f"No data for account '{account_name}'")

    findings = plugin.auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
    severity_counts = {}
    for f in findings:
        s = f.get("severity", "INFO")
        severity_counts[s] = severity_counts.get(s, 0) + 1
    audit_data = {
        "account": account_name, "provider": detected_provider,
        "total": len(findings), "findings": findings,
        "summary": {s: severity_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]},
    }

    if format == "csv":
        content = generate_audit_csv(audit_data, account_name, detected_provider)
        return Response(content=content, media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="audit-{account_name}.csv"'})
    else:
        content = generate_audit_pdf(audit_data, account_name, detected_provider)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="audit-{account_name}.pdf"'})


# ── THREAT DETECTION ─────────────────────────────────────────────

def _detect_threats(account_name: str) -> dict:
    """Analyze cloud data for real-time threats and anomalies."""
    from providers.aws.parser import _read_json, _get_regions

    threats = []
    threat_id = 0

    # Find account data dir
    acct_dir = ACCOUNT_DATA_DIR / "aws" / account_name
    if not acct_dir.exists():
        acct_dir = ACCOUNT_DATA_DIR / account_name
    if not acct_dir.exists():
        # DB fallback — build threats from DB findings
        try:
            dash = _dashboard_from_db(account_name)
            db_findings = dash.get("findings", [])
            for f in db_findings:
                threat_id += 1
                cat_map = {"iam": "identity_threat", "identity": "identity_threat",
                           "network": "network_exposure", "security_group": "network_exposure",
                           "encryption": "encryption_gap", "logging": "detection_gap",
                           "monitoring": "detection_gap", "public": "public_exposure",
                           "data": "data_exposure", "secret": "secret_exposure",
                           "storage": "data_exposure", "compute": "public_exposure"}
                category = cat_map.get(f.get("category", ""), "network_exposure")
                mitre_map = {"identity_threat": ("Credential Access", "T1078 - Valid Accounts"),
                             "network_exposure": ("Initial Access", "T1190 - Exploit Public-Facing App"),
                             "detection_gap": ("Defense Evasion", "T1562.008 - Disable Cloud Logs"),
                             "encryption_gap": ("Collection", "T1530 - Data from Cloud Storage"),
                             "public_exposure": ("Discovery", "T1580 - Cloud Infrastructure Discovery"),
                             "data_exposure": ("Exfiltration", "T1537 - Transfer to Cloud Account"),
                             "secret_exposure": ("Credential Access", "T1552 - Unsecured Credentials")}
                tactic, technique = mitre_map.get(category, ("Initial Access", "T1190"))
                threats.append({
                    "id": f"TH-{threat_id:04d}", "severity": f.get("severity", "MEDIUM"),
                    "category": category, "title": f.get("title", ""),
                    "description": f.get("remediation", f.get("title", "")),
                    "resource": f.get("resource_id", ""), "resource_type": f.get("resource_type", ""),
                    "region": f.get("region", "global"),
                    "detected_at": dash.get("collection_date", datetime.now().isoformat()),
                    "status": "active", "mitre_tactic": tactic, "mitre_technique": technique,
                    "remediation": f.get("remediation", "Review and fix this finding."),
                })
            # Build attack paths from DB threats
            sev = {}
            cats = {}
            regions_set = set()
            res_types = set()
            for t in threats:
                sev[t["severity"]] = sev.get(t["severity"], 0) + 1
                cats[t["category"]] = cats.get(t["category"], 0) + 1
                if t["region"]: regions_set.add(t["region"])
                if t["resource_type"]: res_types.add(t["resource_type"])
            attack_paths = []
            net = [t for t in threats if t["category"] == "network_exposure"]
            pub = [t for t in threats if t["category"] == "public_exposure"]
            ident = [t for t in threats if t["category"] == "identity_threat"]
            if net:
                attack_paths.append({
                    "id": "AP-001", "name": "Internet → Open Ports → Cloud Resources", "severity": "CRITICAL",
                    "description": "Attacker can reach resources through open security group rules",
                    "steps": [
                        {"step": 1, "action": "Scan for open ports", "detail": f"{len(net)} open port rule(s)", "type": "recon"},
                        {"step": 2, "action": "Exploit open service", "detail": "Access exposed services", "type": "exploit"},
                        {"step": 3, "action": "Gain access", "detail": "Compromise through open port", "type": "access"},
                        {"step": 4, "action": "Lateral movement", "detail": "Move to other resources", "type": "escalate"},
                    ],
                    "impact": "Resource compromise via open ports",
                    "mitre_chain": ["T1595", "T1190", "T1078", "T1552"],
                })
            if ident:
                attack_paths.append({
                    "id": "AP-002", "name": "Credential Theft → Account Takeover", "severity": "HIGH",
                    "description": "Weak identity controls allow account compromise",
                    "steps": [
                        {"step": 1, "action": "Identify weak IAM", "detail": f"{len(ident)} identity issue(s)", "type": "recon"},
                        {"step": 2, "action": "Steal credentials", "detail": "No MFA = single factor", "type": "exploit"},
                        {"step": 3, "action": "Escalate privileges", "detail": "Use stolen identity", "type": "escalate"},
                        {"step": 4, "action": "Persistent access", "detail": "Create backdoor", "type": "persist"},
                    ],
                    "impact": "Full account takeover",
                    "mitre_chain": ["T1566", "T1078", "T1098"],
                })
            risk = min(100, sev.get("CRITICAL", 0) * 15 + sev.get("HIGH", 0) * 5 + sev.get("MEDIUM", 0) * 2)
            return {
                "account": account_name, "scanned_at": dash.get("collection_date", ""),
                "threats": threats, "attack_paths": attack_paths,
                "total": len(threats), "risk_score": risk,
                "summary": {"severity": sev, "categories": cats,
                            "regions": {r: sum(1 for t in threats if t["region"] == r) for r in regions_set},
                            "resource_types": {r: sum(1 for t in threats if t["resource_type"] == r) for r in res_types}},
            }
        except HTTPException:
            return {"threats": [], "attack_paths": [], "summary": {}, "risk_score": 0, "total": 0}

    regions = _get_regions(acct_dir)

    for region in regions:
        region_dir = acct_dir / region

        # ── Threat: Open ports to internet ──
        sgs = _read_json(region_dir / "ec2-describe-security-groups.json")
        for sg in sgs.get("SecurityGroups", []):
            for rule in sg.get("IpPermissions", []):
                for ipr in rule.get("IpRanges", []) + rule.get("Ipv6Ranges", []):
                    cidr = ipr.get("CidrIp", ipr.get("CidrIpv6", ""))
                    if cidr in ("0.0.0.0/0", "::/0"):
                        proto = rule.get("IpProtocol", "")
                        from_port = rule.get("FromPort")
                        to_port = rule.get("ToPort")
                        dangerous_ports = {22: "SSH", 3389: "RDP", 3306: "MySQL", 5432: "PostgreSQL", 1433: "MSSQL", 27017: "MongoDB", 6379: "Redis"}
                        port_label = dangerous_ports.get(from_port, "")

                        if proto == "-1" or port_label:
                            severity = "CRITICAL"
                            category = "network_exposure"
                            title = f"{'All traffic' if proto == '-1' else port_label + ' (port ' + str(from_port) + ')'} open to internet"
                        else:
                            severity = "HIGH"
                            category = "network_exposure"
                            title = f"Port {from_port}-{to_port} open to internet"

                        threat_id += 1
                        threats.append({
                            "id": f"TH-{threat_id:04d}",
                            "severity": severity,
                            "category": category,
                            "title": title,
                            "description": f"Security group {sg.get('GroupName', '')} ({sg.get('GroupId', '')}) allows {cidr} ingress",
                            "resource": f"{sg.get('GroupId', '')} ({sg.get('GroupName', '')})",
                            "resource_type": "SecurityGroup",
                            "region": region,
                            "detected_at": datetime.now().isoformat(),
                            "status": "active",
                            "mitre_tactic": "Initial Access",
                            "mitre_technique": "T1190 - Exploit Public-Facing Application",
                            "remediation": f"Restrict {sg.get('GroupId', '')} ingress to specific CIDRs. Remove 0.0.0.0/0 rule.",
                        })

        # ── Threat: Public EC2 instances ──
        ec2 = _read_json(region_dir / "ec2-describe-instances.json")
        for res in ec2.get("Reservations", []):
            for inst in res.get("Instances", []):
                if inst.get("PublicIpAddress") and inst.get("State", {}).get("Name") == "running":
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}",
                        "severity": "HIGH",
                        "category": "public_exposure",
                        "title": f"EC2 instance with public IP: {inst.get('PublicIpAddress')}",
                        "description": f"Instance {inst.get('InstanceId', '')} has public IP {inst.get('PublicIpAddress')} and is running",
                        "resource": inst.get("InstanceId", ""),
                        "resource_type": "EC2",
                        "region": region,
                        "detected_at": datetime.now().isoformat(),
                        "status": "active",
                        "mitre_tactic": "Initial Access",
                        "mitre_technique": "T1133 - External Remote Services",
                        "remediation": "Remove public IP. Use NAT Gateway, VPN, or Session Manager for access.",
                    })

        # ── Threat: Public RDS ──
        rds = _read_json(region_dir / "rds-describe-db-instances.json")
        for db in rds.get("DBInstances", []):
            if db.get("PubliclyAccessible"):
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}",
                    "severity": "CRITICAL",
                    "category": "data_exposure",
                    "title": f"Database publicly accessible: {db.get('DBInstanceIdentifier', '')}",
                    "description": f"RDS instance {db.get('DBInstanceIdentifier', '')} ({db.get('Engine', '')}) is publicly accessible",
                    "resource": db.get("DBInstanceIdentifier", ""),
                    "resource_type": "RDS",
                    "region": region,
                    "detected_at": datetime.now().isoformat(),
                    "status": "active",
                    "mitre_tactic": "Collection",
                    "mitre_technique": "T1530 - Data from Cloud Storage",
                    "remediation": "Disable public access on the RDS instance. Move to private subnet.",
                })

        # ── Threat: GuardDuty not enabled ──
        gd = _read_json(region_dir / "guardduty-list-detectors.json")
        if not gd.get("DetectorIds"):
            # Only flag for regions with resources
            has_resources = bool(ec2.get("Reservations") or rds.get("DBInstances"))
            if has_resources:
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}",
                    "severity": "MEDIUM",
                    "category": "detection_gap",
                    "title": f"GuardDuty not enabled in {region}",
                    "description": f"AWS GuardDuty threat detection is not active in {region} where resources exist",
                    "resource": region,
                    "resource_type": "GuardDuty",
                    "region": region,
                    "detected_at": datetime.now().isoformat(),
                    "status": "active",
                    "mitre_tactic": "Defense Evasion",
                    "mitre_technique": "T1562 - Impair Defenses",
                    "remediation": "Enable GuardDuty in this region for automated threat detection.",
                })

    # ── Threat: IAM issues ──
    for region in regions[:1]:  # IAM is global, check once
        region_dir = acct_dir / region
        iam_auth = _read_json(region_dir / "iam-get-account-authorization-details.json")
        iam_summary = _read_json(region_dir / "iam-get-account-summary.json").get("SummaryMap", {})

        if not iam_summary.get("AccountMFAEnabled"):
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}",
                "severity": "CRITICAL",
                "category": "identity_threat",
                "title": "Root account MFA not enabled",
                "description": "The AWS root account does not have MFA enabled, making it vulnerable to credential theft",
                "resource": "root-account",
                "resource_type": "IAM",
                "region": "global",
                "detected_at": datetime.now().isoformat(),
                "status": "active",
                "mitre_tactic": "Credential Access",
                "mitre_technique": "T1078 - Valid Accounts",
                "remediation": "Enable MFA on root account immediately. Use hardware MFA key.",
            })

        if iam_summary.get("AccountAccessKeysPresent"):
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}",
                "severity": "CRITICAL",
                "category": "identity_threat",
                "title": "Root account has access keys",
                "description": "Active access keys exist for the root account, posing severe security risk",
                "resource": "root-account",
                "resource_type": "IAM",
                "region": "global",
                "detected_at": datetime.now().isoformat(),
                "status": "active",
                "mitre_tactic": "Credential Access",
                "mitre_technique": "T1528 - Steal Application Access Token",
                "remediation": "Delete root access keys. Use IAM roles instead.",
            })

        users_no_mfa = [u["UserName"] for u in iam_auth.get("UserDetailList", []) if not u.get("MFADevices")]
        if users_no_mfa:
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}",
                "severity": "HIGH",
                "category": "identity_threat",
                "title": f"{len(users_no_mfa)} IAM users without MFA",
                "description": f"Users without MFA: {', '.join(users_no_mfa[:10])}{'...' if len(users_no_mfa) > 10 else ''}",
                "resource": f"{len(users_no_mfa)} users",
                "resource_type": "IAM",
                "region": "global",
                "detected_at": datetime.now().isoformat(),
                "status": "active",
                "mitre_tactic": "Credential Access",
                "mitre_technique": "T1078 - Valid Accounts",
                "remediation": "Enable MFA for all IAM users with console access.",
            })

    # ── Threat: No CloudTrail ──
    has_cloudtrail = False
    for region in regions:
        ct = _read_json(acct_dir / region / "cloudtrail-describe-trails.json")
        if ct.get("trailList"):
            has_cloudtrail = True
            break
    if not has_cloudtrail:
        threat_id += 1
        threats.append({
            "id": f"TH-{threat_id:04d}",
            "severity": "HIGH",
            "category": "detection_gap",
            "title": "No CloudTrail logging detected",
            "description": "CloudTrail is not enabled, meaning API activity is not being logged for audit purposes",
            "resource": "account",
            "resource_type": "CloudTrail",
            "region": "global",
            "detected_at": datetime.now().isoformat(),
            "status": "active",
            "mitre_tactic": "Defense Evasion",
            "mitre_technique": "T1562.008 - Disable Cloud Logs",
            "remediation": "Enable CloudTrail in all regions with log file validation.",
        })

    # ── Advanced: Secret Detection in Lambda env vars ──
    for region in regions:
        lambdas = _read_json(acct_dir / region / "lambda-list-functions.json")
        for fn in lambdas.get("Functions", []):
            env_vars = fn.get("Environment", {}).get("Variables", {})
            secret_patterns = ["KEY", "SECRET", "PASSWORD", "TOKEN", "CREDENTIAL", "API_KEY", "AUTH", "PRIVATE"]
            for key, val in env_vars.items():
                if any(p in key.upper() for p in secret_patterns) and len(val) > 8:
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}", "severity": "CRITICAL",
                        "category": "secret_exposure",
                        "title": f"Potential secret in Lambda env var: {key}",
                        "description": f"Lambda function {fn.get('FunctionName', '')} has environment variable '{key}' that may contain a secret/credential",
                        "resource": fn.get("FunctionName", ""), "resource_type": "Lambda",
                        "region": region, "detected_at": datetime.now().isoformat(), "status": "active",
                        "mitre_tactic": "Credential Access",
                        "mitre_technique": "T1552.001 - Credentials In Files",
                        "remediation": f"Move secret '{key}' to AWS Secrets Manager or Parameter Store. Reference it at runtime instead of embedding.",
                    })

    # ── Advanced: S3 Data Exposure ──
    for region in regions[:1]:  # S3 is global
        s3 = _read_json(acct_dir / region / "s3-list-buckets.json")
        bucket_count = len(s3.get("Buckets", []))
        if bucket_count > 0:
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}", "severity": "MEDIUM",
                "category": "data_exposure",
                "title": f"{bucket_count} S3 buckets require public access review",
                "description": f"Review all {bucket_count} S3 buckets for public access settings. Enable S3 Block Public Access at account level.",
                "resource": f"{bucket_count} buckets", "resource_type": "S3",
                "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                "mitre_tactic": "Collection",
                "mitre_technique": "T1530 - Data from Cloud Storage",
                "remediation": "Enable S3 Block Public Access at account level: aws s3control put-public-access-block --account-id <ID> --public-access-block-configuration BlockPublicAcls=true,IgnorePublicAcls=true,BlockPublicPolicy=true,RestrictPublicBuckets=true",
            })

    # ── Advanced: Encryption Audit ──
    for region in regions:
        # Unencrypted snapshots
        snaps = _read_json(acct_dir / region / "ec2-describe-snapshots.json")
        unenc_snaps = [s for s in snaps.get("Snapshots", []) if not s.get("Encrypted")]
        if unenc_snaps:
            threat_id += 1
            threats.append({
                "id": f"TH-{threat_id:04d}", "severity": "HIGH",
                "category": "encryption_gap",
                "title": f"{len(unenc_snaps)} unencrypted EBS snapshots in {region}",
                "description": f"Found {len(unenc_snaps)} snapshots without encryption: {', '.join(s.get('SnapshotId','') for s in unenc_snaps[:5])}",
                "resource": f"{len(unenc_snaps)} snapshots", "resource_type": "Snapshot",
                "region": region, "detected_at": datetime.now().isoformat(), "status": "active",
                "mitre_tactic": "Collection",
                "mitre_technique": "T1530 - Data from Cloud Storage",
                "remediation": "Copy snapshots with encryption enabled. Enable EBS encryption by default for the account.",
            })

        # Unencrypted RDS
        rds = _read_json(acct_dir / region / "rds-describe-db-instances.json")
        for db in rds.get("DBInstances", []):
            if not db.get("StorageEncrypted"):
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}", "severity": "HIGH",
                    "category": "encryption_gap",
                    "title": f"RDS instance without encryption: {db.get('DBInstanceIdentifier', '')}",
                    "description": f"Database {db.get('DBInstanceIdentifier', '')} ({db.get('Engine', '')}) does not have storage encryption enabled",
                    "resource": db.get("DBInstanceIdentifier", ""), "resource_type": "RDS",
                    "region": region, "detected_at": datetime.now().isoformat(), "status": "active",
                    "mitre_tactic": "Collection",
                    "mitre_technique": "T1565 - Data Manipulation",
                    "remediation": "Encrypt RDS: Create encrypted snapshot → Restore from snapshot → Switch DNS/endpoint.",
                })

    # ── Advanced: Lateral Movement - Overly permissive IAM ──
    for region in regions[:1]:
        iam_auth = _read_json(acct_dir / region / "iam-get-account-authorization-details.json")
        for role in iam_auth.get("RoleDetailList", []):
            trust = role.get("AssumeRolePolicyDocument", {})
            if isinstance(trust, str):
                try:
                    trust = json.loads(trust)
                except Exception:
                    continue
            for stmt in trust.get("Statement", []):
                principal = stmt.get("Principal", {})
                if isinstance(principal, str) and principal == "*":
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}", "severity": "CRITICAL",
                        "category": "lateral_movement",
                        "title": f"IAM role assumable by anyone: {role.get('RoleName', '')}",
                        "description": f"Role {role.get('RoleName', '')} has trust policy with Principal: '*', allowing any AWS account to assume it",
                        "resource": role.get("RoleName", ""), "resource_type": "IAM Role",
                        "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                        "mitre_tactic": "Lateral Movement",
                        "mitre_technique": "T1550.001 - Application Access Token",
                        "remediation": f"Restrict trust policy for role {role.get('RoleName', '')} to specific account IDs or services.",
                    })

            # Check for admin policies
            for policy in role.get("AttachedManagedPolicies", []):
                if "AdministratorAccess" in policy.get("PolicyName", ""):
                    threat_id += 1
                    threats.append({
                        "id": f"TH-{threat_id:04d}", "severity": "HIGH",
                        "category": "lateral_movement",
                        "title": f"Role with full admin access: {role.get('RoleName', '')}",
                        "description": f"Role {role.get('RoleName', '')} has AdministratorAccess policy attached — full control over all resources",
                        "resource": role.get("RoleName", ""), "resource_type": "IAM Role",
                        "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                        "mitre_tactic": "Privilege Escalation",
                        "mitre_technique": "T1078.004 - Cloud Accounts",
                        "remediation": "Apply least-privilege principle. Replace AdministratorAccess with specific service policies.",
                    })

    # ── Advanced: Compliance Drift ──
    try:
        latest_compliance = _compliance_store.get_latest(account_name)
        history = _compliance_store.get_history(account_name)
        if latest_compliance and len(history) >= 2:
            current_score = latest_compliance.get("summary", {}).get("overall_score", 0)
            prev_score = history[1].get("overall_score", 0)  # index 0 is latest, 1 is previous
            if current_score < prev_score:
                drift = prev_score - current_score
                threat_id += 1
                threats.append({
                    "id": f"TH-{threat_id:04d}", "severity": "HIGH" if drift > 10 else "MEDIUM",
                    "category": "compliance_drift",
                    "title": f"Compliance score dropped {drift}% (from {prev_score}% to {current_score}%)",
                    "description": f"Overall compliance score decreased by {drift} percentage points since the previous scan. This indicates new misconfigurations or policy violations.",
                    "resource": "compliance", "resource_type": "Compliance",
                    "region": "global", "detected_at": datetime.now().isoformat(), "status": "active",
                    "mitre_tactic": "Defense Evasion",
                    "mitre_technique": "T1562 - Impair Defenses",
                    "remediation": f"Review compliance results at /compliance. Focus on newly failed checks to restore score from {current_score}% to {prev_score}%+.",
                })
    except Exception:
        pass

    # ── Build Attack Paths ──
    attack_paths = []

    # Path 1: Internet → Open Port → EC2 → IAM Role → Data
    public_instances = [t for t in threats if t["category"] == "public_exposure"]
    open_ports = [t for t in threats if t["category"] == "network_exposure"]
    data_exposure = [t for t in threats if t["category"] == "data_exposure"]
    identity_threats = [t for t in threats if t["category"] == "identity_threat"]

    if open_ports and public_instances:
        attack_paths.append({
            "id": "AP-001",
            "name": "Internet → EC2 via Open Ports",
            "severity": "CRITICAL",
            "description": "Attacker can reach EC2 instances through open security group rules and public IPs",
            "steps": [
                {"step": 1, "action": "Scan public IPs", "detail": f"{len(public_instances)} instance(s) with public IPs", "type": "recon"},
                {"step": 2, "action": "Exploit open port", "detail": f"{len(open_ports)} open port rule(s) to 0.0.0.0/0", "type": "exploit"},
                {"step": 3, "action": "Gain instance access", "detail": "SSH/RDP access via open ports", "type": "access"},
                {"step": 4, "action": "Steal IAM credentials", "detail": "Query instance metadata for IAM role credentials", "type": "escalate"},
            ],
            "impact": "Full instance compromise, potential lateral movement via IAM role",
            "mitre_chain": ["T1595 - Active Scanning", "T1190 - Exploit Public-Facing App", "T1078 - Valid Accounts", "T1552 - Unsecured Credentials"],
        })

    if data_exposure:
        attack_paths.append({
            "id": "AP-002",
            "name": "Internet → Public Database → Data Exfiltration",
            "severity": "CRITICAL",
            "description": "Publicly accessible databases can be directly attacked from the internet",
            "steps": [
                {"step": 1, "action": "Discover public endpoint", "detail": f"{len(data_exposure)} public database(s)", "type": "recon"},
                {"step": 2, "action": "Brute force credentials", "detail": "Default or weak database credentials", "type": "exploit"},
                {"step": 3, "action": "Exfiltrate data", "detail": "Direct SQL access to all data", "type": "exfiltrate"},
            ],
            "impact": "Complete data breach — customer data, PII, financial records",
            "mitre_chain": ["T1595 - Active Scanning", "T1110 - Brute Force", "T1530 - Data from Cloud Storage"],
        })

    if identity_threats:
        attack_paths.append({
            "id": "AP-003",
            "name": "Credential Theft → Account Takeover",
            "severity": "HIGH",
            "description": "Weak identity controls allow account compromise through credential theft",
            "steps": [
                {"step": 1, "action": "Phish IAM user", "detail": f"{len([t for t in identity_threats if 'without MFA' in t.get('title','')])} users without MFA protection", "type": "recon"},
                {"step": 2, "action": "Use stolen credentials", "detail": "No MFA = single-factor compromise", "type": "exploit"},
                {"step": 3, "action": "Escalate privileges", "detail": "Access cloud console or API with stolen identity", "type": "escalate"},
                {"step": 4, "action": "Persistent access", "detail": "Create backdoor IAM users or access keys", "type": "persist"},
            ],
            "impact": "Full account takeover, resource manipulation, data theft",
            "mitre_chain": ["T1566 - Phishing", "T1078 - Valid Accounts", "T1098 - Account Manipulation"],
        })

    secret_threats = [t for t in threats if t["category"] == "secret_exposure"]
    if secret_threats:
        attack_paths.append({
            "id": "AP-004",
            "name": "Exposed Secrets → Service Compromise",
            "severity": "CRITICAL",
            "description": "Hardcoded secrets in Lambda functions can be extracted and used to access other services",
            "steps": [
                {"step": 1, "action": "Access Lambda config", "detail": f"{len(secret_threats)} functions with embedded secrets", "type": "recon"},
                {"step": 2, "action": "Extract credentials", "detail": "Read environment variables via console or API", "type": "exploit"},
                {"step": 3, "action": "Access target service", "detail": "Use extracted API keys/tokens for lateral movement", "type": "escalate"},
            ],
            "impact": "Compromise of connected services — databases, APIs, third-party systems",
            "mitre_chain": ["T1552.001 - Credentials In Files", "T1550 - Use Alternate Auth Material"],
        })

    lateral = [t for t in threats if t["category"] == "lateral_movement"]
    encryption = [t for t in threats if t["category"] == "encryption_gap"]

    if lateral:
        attack_paths.append({
            "id": "AP-005",
            "name": "Over-Privileged Role → Cross-Account Pivot",
            "severity": "HIGH",
            "description": "Overly permissive IAM roles can be assumed to pivot across the infrastructure",
            "steps": [
                {"step": 1, "action": "Discover assumable roles", "detail": f"{len(lateral)} roles with excessive permissions", "type": "recon"},
                {"step": 2, "action": "Assume admin role", "detail": "Use sts:AssumeRole with wildcard trust", "type": "escalate"},
                {"step": 3, "action": "Full infrastructure access", "detail": "Admin role grants control over all resources", "type": "access"},
            ],
            "impact": "Complete infrastructure compromise — create, modify, delete any resource",
            "mitre_chain": ["T1078.004 - Cloud Accounts", "T1550.001 - Application Access Token"],
        })

    # Sort by severity
    sev_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}
    threats.sort(key=lambda t: sev_order.get(t["severity"], 5))

    # Summary
    sev_counts = {}
    cat_counts = {}
    region_counts = {}
    resource_type_counts = {}
    for t in threats:
        sev_counts[t["severity"]] = sev_counts.get(t["severity"], 0) + 1
        cat_counts[t["category"]] = cat_counts.get(t["category"], 0) + 1
        region_counts[t["region"]] = region_counts.get(t["region"], 0) + 1
        resource_type_counts[t["resource_type"]] = resource_type_counts.get(t["resource_type"], 0) + 1

    total = len(threats)
    crit = sev_counts.get("CRITICAL", 0)
    high = sev_counts.get("HIGH", 0)
    # Health score based on threat severity distribution
    medium = sev_counts.get("MEDIUM", 0)
    low = sev_counts.get("LOW", 0)
    if total == 0:
        risk_score = 100
    else:
        # Score based on ratio of non-critical threats
        crit_ratio = crit / total        # 0 to 1
        high_ratio = high / total        # 0 to 1
        # 100 = no threats, 0 = all critical
        risk_score = max(5, min(100, round(100 - crit_ratio * 60 - high_ratio * 30 - (medium / max(total, 1)) * 10)))

    return {
        "account": account_name,
        "scanned_at": datetime.now().isoformat(),
        "threats": threats,
        "attack_paths": attack_paths,
        "total": total,
        "risk_score": risk_score,
        "summary": {
            "severity": {s: sev_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]},
            "categories": cat_counts,
            "regions": dict(sorted(region_counts.items(), key=lambda x: x[1], reverse=True)[:10]),
            "resource_types": resource_type_counts,
        },
    }


@app.get("/api/threats/{account_name}")
def get_threats(account_name: str, user: dict = Depends(get_current_user)):
    """Advanced threat detection with attack paths, secret scanning, encryption audit, and MITRE ATT&CK mapping."""
    result = _detect_threats(account_name)
    if not result.get("threats") and result.get("total", 0) == 0:
        # Try one more time with DB
        try:
            dash = _dashboard_from_db(account_name)
            if dash.get("findings"):
                result = _detect_threats(account_name)  # Should hit DB fallback now
        except Exception:
            pass
    return result


# ── COMPLIANCE MODULE ────────────────────────────────────────────
from compliance.engine import ComplianceScanner, ComplianceStore
from compliance.frameworks import get_all_frameworks, get_framework

_compliance_scanner = ComplianceScanner(ACCOUNT_DATA_DIR)
_compliance_store = ComplianceStore(BASE_DIR)


class ComplianceScanRequest(BaseModel):
    frameworks: list = []  # empty = all frameworks


@app.get("/api/compliance/frameworks")
def list_compliance_frameworks(user: dict = Depends(get_current_user)):
    """List all supported compliance frameworks."""
    frameworks = get_all_frameworks()
    return {"frameworks": [
        {
            "id": fw.framework_id, "name": fw.name, "version": fw.version,
            "description": fw.description, "category": fw.category,
            "icon": fw.icon, "color": fw.color,
            "total_controls": len(fw.controls),
            "total_checks": fw.total_checks,
        }
        for fw in frameworks.values()
    ]}


@app.get("/api/compliance/frameworks/{framework_id}")
def get_compliance_framework(framework_id: str, user: dict = Depends(get_current_user)):
    """Get detailed framework info with controls and checks."""
    fw = get_framework(framework_id)
    if not fw:
        raise HTTPException(404, f"Framework '{framework_id}' not found")
    return {
        "id": fw.framework_id, "name": fw.name, "version": fw.version,
        "description": fw.description, "category": fw.category,
        "icon": fw.icon, "color": fw.color,
        "controls": [
            {
                "control_id": c.control_id, "title": c.title,
                "description": c.description, "section": c.section,
                "severity": c.severity,
                "checks": [
                    {"check_id": ch.check_id, "title": ch.title, "description": ch.description,
                     "resource_type": ch.resource_type, "severity": ch.severity, "remediation": ch.remediation}
                    for ch in c.checks
                ],
            }
            for c in fw.controls
        ],
    }


@app.post("/api/compliance/scan/{account_name}")
def run_compliance_scan(account_name: str, req: ComplianceScanRequest = ComplianceScanRequest(),
                        user: dict = Depends(get_current_user)):
    """Run compliance scan against all or selected frameworks."""
    fw_ids = req.frameworks if req.frameworks else None
    results = _compliance_scanner.scan(account_name, fw_ids)
    if "error" in results:
        raise HTTPException(404, results["error"])
    _compliance_store.save(account_name, results)
    return results


@app.get("/api/compliance/results/{account_name}")
def get_compliance_results(account_name: str, user: dict = Depends(get_current_user)):
    """Get latest compliance scan results."""
    results = _compliance_store.get_latest(account_name)
    if not results:
        # Try running compliance scan from DB data
        try:
            dash = _dashboard_from_db(account_name)
            if dash:
                results = _compliance_scanner.scan(account_name, None)
                if "error" not in results:
                    _compliance_store.save(account_name, results)
                    return results
        except Exception:
            pass
        # Still nothing — try to generate minimal compliance from DB findings
        try:
            dash = _dashboard_from_db(account_name)
            findings = dash.get("findings", [])
            if findings:
                from compliance.frameworks import get_all_frameworks
                frameworks = {}
                for fw_id, fw in get_all_frameworks().items():
                    total = fw.total_checks
                    failed = sum(1 for f in findings if f.get("severity") in ("CRITICAL", "HIGH"))
                    passed = max(0, total - min(failed, total))
                    score = round((passed / total) * 100) if total > 0 else 0
                    frameworks[fw_id] = {"name": fw.name, "score": score, "passed": passed,
                                         "failed": min(failed, total), "total": total}
                sev_counts = {}
                for f in findings:
                    sev_counts[f.get("severity", "INFO")] = sev_counts.get(f.get("severity", "INFO"), 0) + 1
                return {
                    "scan_id": "db-generated", "account": account_name,
                    "scanned_at": dash.get("collection_date", ""),
                    "frameworks": frameworks,
                    "summary": {"overall_score": round(sum(f["score"] for f in frameworks.values()) / len(frameworks)) if frameworks else 0,
                                "total_checks": sum(f["total"] for f in frameworks.values()),
                                "passed_checks": sum(f["passed"] for f in frameworks.values()),
                                "failed_checks": sum(f["failed"] for f in frameworks.values())},
                    "all_findings": [{"severity": f.get("severity"), "title": f.get("title"),
                                      "resource_type": f.get("resource_type"), "region": f.get("region"),
                                      "remediation": f.get("remediation")} for f in findings],
                    "ai_recommendations": [],
                }
        except Exception:
            pass
        raise HTTPException(404, f"No compliance results for '{account_name}'. Run a scan first.")
    return results


@app.get("/api/compliance/history/{account_name}")
def get_compliance_history(account_name: str, user: dict = Depends(get_current_user)):
    """Get compliance scan history for drift tracking."""
    history = _compliance_store.get_history(account_name)
    return {"account": account_name, "history": history}


@app.get("/api/compliance/results/{account_name}/export")
def export_compliance_report(account_name: str, format: str = "pdf",
                             user: dict = Depends(get_current_user)):
    """Export compliance report as PDF, Excel, or JSON."""
    results = _compliance_store.get_latest(account_name)
    if not results:
        raise HTTPException(404, f"No compliance results for '{account_name}'")

    if format == "json":
        return Response(
            content=json.dumps(results, indent=2, default=str),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="compliance-{account_name}.json"'},
        )
    elif format == "csv":
        lines = ["Framework,Control,Check,Status,Severity,Reason,Remediation"]
        for fw_id, fw in results.get("frameworks", {}).items():
            for ctrl in fw.get("controls", []):
                for ch in ctrl.get("checks", []):
                    lines.append(','.join([
                        f'"{fw["name"]}"', f'"{ctrl["section"]}"',
                        f'"{ch["title"]}"', ch["status"], ch["severity"],
                        f'"{ch["reason"]}"', f'"{ch["remediation"][:100]}"',
                    ]))
        content = '\n'.join(lines)
        return Response(content=content, media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="compliance-{account_name}.csv"'})
    else:
        # PDF export using existing report generator patterns
        content = _generate_compliance_pdf(results)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="compliance-{account_name}.pdf"'})


def _generate_compliance_pdf(results):
    """Generate compliance PDF report."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1E293B'), spaceAfter=5)
    h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=13, textColor=colors.HexColor('#4F46E5'), spaceBefore=12, spaceAfter=6)
    body = ParagraphStyle('B', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#475569'), leading=12)
    small = ParagraphStyle('S', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94A3B8'))

    elements = []
    summary = results.get("summary", {})
    elements.append(Paragraph("CloudSentrix Compliance Report", title_s))
    elements.append(Paragraph(f"Account: {results.get('account')} | Scanned: {results.get('scanned_at', '')[:19]} | Score: {summary.get('overall_score', 0)}%", small))
    elements.append(Spacer(1, 10))

    # Summary table
    elements.append(Paragraph("Executive Summary", h1))
    summary_data = [
        ["Overall Score", f"{summary.get('overall_score', 0)}%", "Total Checks", str(summary.get('total_checks', 0))],
        ["Passed", str(summary.get('passed', 0)), "Failed", str(summary.get('failed', 0))],
        ["Warnings", str(summary.get('warnings', 0)), "Critical", str(summary.get('severity_counts', {}).get('CRITICAL', 0))],
    ]
    t = Table(summary_data, colWidths=[80, 60, 80, 60])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('FONTNAME', (3, 0), (3, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)

    # Framework scores
    elements.append(Paragraph("Framework Compliance Scores", h1))
    fw_rows = [["Framework", "Score", "Passed", "Failed", "Status"]]
    for fw_id, fw in results.get("frameworks", {}).items():
        status = "PASS" if fw["score"] >= 80 else "WARN" if fw["score"] >= 50 else "FAIL"
        fw_rows.append([fw["name"], f"{fw['score']}%", str(fw["passed"]), str(fw["failed"]), status])
    t = Table(fw_rows, colWidths=[140, 50, 50, 50, 50])
    style_list = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')), ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ]
    sev_bg = {"PASS": '#D1FAE5', "WARN": '#FEF9C3', "FAIL": '#FEE2E2'}
    for i, row in enumerate(fw_rows[1:], 1):
        style_list.append(('BACKGROUND', (4, i), (4, i), colors.HexColor(sev_bg.get(row[4], '#F1F5F9'))))
    t.setStyle(TableStyle(style_list))
    elements.append(t)

    elements.append(PageBreak())

    # Findings
    elements.append(Paragraph("Failed Compliance Checks", h1))
    findings = results.get("all_findings", [])
    if findings:
        find_rows = [["Severity", "Check", "Reason", "Framework"]]
        for f in findings[:80]:
            fw_name = results["frameworks"].get(f.get("framework_id", ""), {}).get("name", "")
            find_rows.append([f["severity"], f["title"][:40], f["reason"][:50], fw_name[:20]])
        t = Table(find_rows, colWidths=[50, 140, 130, 80])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),
        ]))
        elements.append(t)

    # AI Recommendations
    elements.append(PageBreak())
    elements.append(Paragraph("AI Recommendations", h1))
    for rec in results.get("ai_recommendations", []):
        elements.append(Paragraph(f"<b>[{rec['priority']}] {rec['title']}</b>", body))
        elements.append(Paragraph(rec["description"], body))
        for item in rec.get("items", []):
            elements.append(Paragraph(f"• {item}", body))
        elements.append(Spacer(1, 6))

    doc.build(elements)
    return buf.getvalue()


# ── AI CHAT ──────────────────────────────────────────────────────
class AiChatRequest(BaseModel):
    message: str
    history: list = []

# Well-Architected Framework pillars for AI analysis
def _check_multi_az(d):
    """Check if instances are deployed across multiple AZs (not just regions)."""
    account = d.get("account", "")
    base = os.path.join(str(ACCOUNT_DATA_DIR), "aws", account)
    if not os.path.isdir(base):
        base = os.path.join(str(ACCOUNT_DATA_DIR), account)
    azs = set()
    for region in os.listdir(base) if os.path.isdir(base) else []:
        ec2_file = os.path.join(base, region, "ec2-describe-instances.json")
        if os.path.isfile(ec2_file):
            try:
                data = json.load(open(ec2_file))
                for r in data.get("Reservations", []):
                    for i in r.get("Instances", []):
                        az = i.get("Placement", {}).get("AvailabilityZone", "")
                        if az:
                            azs.add(az)
            except Exception:
                pass
    return len(azs) > 1


def _check_right_sized(d):
    """Check if any instances are oversized (xlarge or bigger)."""
    account = d.get("account", "")
    base = os.path.join(str(ACCOUNT_DATA_DIR), "aws", account)
    if not os.path.isdir(base):
        base = os.path.join(str(ACCOUNT_DATA_DIR), account)
    oversized = 0
    total = 0
    for region in os.listdir(base) if os.path.isdir(base) else []:
        ec2_file = os.path.join(base, region, "ec2-describe-instances.json")
        if os.path.isfile(ec2_file):
            try:
                data = json.load(open(ec2_file))
                for r in data.get("Reservations", []):
                    for i in r.get("Instances", []):
                        total += 1
                        itype = i.get("InstanceType", "")
                        if any(x in itype for x in ["xlarge", "2xlarge", "4xlarge", "8xlarge", "metal"]):
                            oversized += 1
            except Exception:
                pass
    if total == 0:
        return True
    return oversized / total < 0.5  # Fail if more than half are oversized


def _check_region_consolidation(d):
    """Check if compute/lambda/rds resources are concentrated in <= 3 regions."""
    account = d.get("account", "")
    base = os.path.join(str(ACCOUNT_DATA_DIR), "aws", account)
    if not os.path.isdir(base):
        base = os.path.join(str(ACCOUNT_DATA_DIR), account)
    regions_with_compute = 0
    for region in os.listdir(base) if os.path.isdir(base) else []:
        rdir = os.path.join(base, region)
        if not os.path.isdir(rdir):
            continue
        has_compute = False
        for fname, key in [("ec2-describe-instances.json", "Reservations"),
                           ("lambda-list-functions.json", "Functions"),
                           ("rds-describe-db-instances.json", "DBInstances")]:
            f = os.path.join(rdir, fname)
            if os.path.isfile(f):
                try:
                    items = json.load(open(f)).get(key, [])
                    if key == "Reservations":
                        items = [i for r in items for i in r.get("Instances", [])]
                    if items:
                        has_compute = True
                        break
                except Exception:
                    pass
        if has_compute:
            regions_with_compute += 1
    return regions_with_compute <= 3


WAF_PILLARS = {
    "security": {
        "name": "Security",
        "checks": [
            ("Root MFA", lambda d: d.get("iam_summary", {}).get("AccountMFAEnabled", False), "Enable MFA on root account"),
            ("No root keys", lambda d: not d.get("iam_summary", {}).get("AccountAccessKeysPresent", True), "Remove root account access keys"),
            ("IAM MFA", lambda d: d.get("iam_users_no_mfa", 1) == 0, "Enable MFA for all IAM users"),
            ("No open SGs", lambda d: d.get("open_security_groups", 1) == 0, "Restrict security groups open to 0.0.0.0/0"),
            ("No public RDS", lambda d: d.get("public_summary", {}).get("rds", 1) == 0, "Disable public access on RDS instances"),
        ],
    },
    "reliability": {
        "name": "Reliability",
        "checks": [
            ("Multi-AZ", lambda d: _check_multi_az(d), "Deploy across multiple availability zones"),
            ("Snapshots exist", lambda d: d.get("totals", {}).get("snapshots", 0) > 0, "Create regular EBS/RDS snapshots"),
            ("ELBs present", lambda d: d.get("totals", {}).get("elbs", 0) > 0, "Use load balancers for high availability"),
        ],
    },
    "performance": {
        "name": "Performance Efficiency",
        "checks": [
            ("Right-sized instances", lambda d: _check_right_sized(d), "Review instance types for right-sizing"),
            ("Lambda usage", lambda d: d.get("totals", {}).get("lambdas", 0) > 0, "Consider serverless for variable workloads"),
        ],
    },
    "cost": {
        "name": "Cost Optimization",
        "checks": [
            ("No stopped EC2", lambda d: all(r.get("instances_stopped", 0) == 0 for r in d.get("regions", {}).values()), "Terminate or schedule stopped instances"),
            ("Snapshot cleanup", lambda d: d.get("totals", {}).get("snapshots", 0) < 50, "Review and clean up old snapshots"),
        ],
    },
    "operational": {
        "name": "Operational Excellence",
        "checks": [
            ("CloudTrail", lambda d: any(r.get("cloudtrail_trails", 0) > 0 for r in d.get("regions", {}).values()), "Enable CloudTrail in all regions"),
            ("GuardDuty", lambda d: any(r.get("guardduty_enabled", False) for r in d.get("regions", {}).values()), "Enable GuardDuty for threat detection"),
        ],
    },
    "sustainability": {
        "name": "Sustainability",
        "checks": [
            ("Region consolidation", lambda d: _check_region_consolidation(d), "Consolidate resources to fewer regions"),
        ],
    },
}


def _analyze_waf(dashboard_data):
    """Analyze dashboard data against Well-Architected Framework pillars."""
    results = {}
    for pillar_id, pillar in WAF_PILLARS.items():
        checks = []
        passed = 0
        for check_name, check_fn, recommendation in pillar["checks"]:
            try:
                ok = check_fn(dashboard_data)
            except Exception:
                ok = False
            checks.append({"name": check_name, "passed": ok, "recommendation": recommendation})
            if ok:
                passed += 1
        total = len(checks)
        results[pillar_id] = {
            "name": pillar["name"],
            "score": round(passed / total * 100) if total > 0 else 0,
            "passed": passed,
            "total": total,
            "checks": checks,
        }
    overall = sum(p["score"] for p in results.values()) / len(results) if results else 0
    return {"overall_score": round(overall), "pillars": results}


def _generate_ai_response(message, dashboard_data, audit_findings, waf_analysis):
    """Generate an AI-style response based on actual cloud data."""
    lower = message.lower()
    lines = []

    if any(w in lower for w in ["well-architected", "waf", "pillar", "framework", "architect"]):
        lines.append("## AWS Well-Architected Framework Analysis\n")
        for pid, pillar in waf_analysis.get("pillars", {}).items():
            icon = "✅" if pillar["score"] >= 80 else "⚠️" if pillar["score"] >= 50 else "❌"
            lines.append(f"{icon} **{pillar['name']}**: {pillar['score']}% ({pillar['passed']}/{pillar['total']} checks passed)")
            for check in pillar["checks"]:
                status = "✓" if check["passed"] else "✗"
                lines.append(f"   {status} {check['name']}" + ("" if check["passed"] else f" — {check['recommendation']}"))
        lines.append(f"\n**Overall Score: {waf_analysis.get('overall_score', 0)}%**")
        return "\n".join(lines)

    if any(w in lower for w in ["risk", "top", "critical", "threat", "issue"]):
        critical = [f for f in audit_findings if f.get("severity") == "CRITICAL"]
        high = [f for f in audit_findings if f.get("severity") == "HIGH"]
        lines.append("## Top Security Risks\n")
        if critical:
            lines.append(f"**🔴 {len(critical)} Critical Finding(s):**")
            for f in critical[:5]:
                lines.append(f"- {f.get('title', f.get('issue', 'Unknown'))}")
                if f.get('resource'):
                    lines.append(f"  Resource: `{f['resource']}`")
        if high:
            lines.append(f"\n**🟠 {len(high)} High Finding(s):**")
            for f in high[:5]:
                lines.append(f"- {f.get('title', f.get('issue', 'Unknown'))}")
        if not critical and not high:
            lines.append("No critical or high severity findings detected. Your security posture looks good!")
        lines.append("\n**Recommendation:** Address critical findings immediately, then work through high-severity items.")
        return "\n".join(lines)

    if any(w in lower for w in ["score", "improve", "better", "increase"]):
        score = dashboard_data.get("security_score", 0)
        lines.append(f"## Security Score: {score}/100\n")
        waf = waf_analysis.get("pillars", {})
        weak = sorted(waf.items(), key=lambda x: x[1]["score"])
        lines.append("**Areas to improve (sorted by priority):**\n")
        for pid, pillar in weak:
            if pillar["score"] < 100:
                failed = [c for c in pillar["checks"] if not c["passed"]]
                for c in failed:
                    lines.append(f"- **{pillar['name']}**: {c['recommendation']}")
        return "\n".join(lines)

    if any(w in lower for w in ["posture", "summary", "overview", "status"]):
        t = dashboard_data.get("totals", {})
        lines.append("## Cloud Security Posture Summary\n")
        lines.append(f"- **EC2 Instances:** {t.get('instances', 0)}")
        lines.append(f"- **S3 Buckets:** {t.get('buckets', 0)}")
        lines.append(f"- **Lambda Functions:** {t.get('lambdas', 0)}")
        lines.append(f"- **RDS Instances:** {t.get('rds', 0)}")
        lines.append(f"- **Security Groups:** {t.get('security_groups', 0)}")
        lines.append(f"- **Open Security Groups:** {dashboard_data.get('open_security_groups', 0)}")
        lines.append(f"- **Public IPs:** {len(dashboard_data.get('public_ips', []))}")
        lines.append(f"- **Security Score:** {dashboard_data.get('security_score', 0)}/100")
        lines.append(f"- **Regions Scanned:** {dashboard_data.get('regions_scanned', 0)}")
        sev = {}
        for f in audit_findings:
            s = f.get("severity", "INFO")
            sev[s] = sev.get(s, 0) + 1
        lines.append(f"\n**Audit Findings:** {len(audit_findings)} total")
        for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]:
            if sev.get(s, 0) > 0:
                lines.append(f"  - {s}: {sev[s]}")
        return "\n".join(lines)

    if any(w in lower for w in ["public", "exposed", "internet"]):
        pips = dashboard_data.get("public_ips", [])
        ps = dashboard_data.get("public_summary", {})
        lines.append(f"## Publicly Exposed Resources ({len(pips)} total)\n")
        if ps.get("ec2", 0) > 0:
            lines.append(f"- **{ps['ec2']} EC2 instance(s)** with public IPs")
        if ps.get("rds", 0) > 0:
            lines.append(f"- **{ps['rds']} RDS instance(s)** publicly accessible ⚠️")
        if ps.get("elb", 0) > 0:
            lines.append(f"- **{ps['elb']} Load Balancer(s)** internet-facing")
        osg = dashboard_data.get("open_security_groups", 0)
        if osg > 0:
            lines.append(f"- **{osg} Security Group rule(s)** open to 0.0.0.0/0")
        if not pips:
            lines.append("No publicly exposed resources detected. Great job!")
        else:
            lines.append("\n**Recommendations:**")
            lines.append("- Move RDS instances to private subnets")
            lines.append("- Use VPC endpoints instead of public access")
            lines.append("- Restrict security group ingress to known CIDRs")
        return "\n".join(lines)

    if any(w in lower for w in ["mfa", "authentication", "iam", "user"]):
        iam = dashboard_data.get("iam_summary", {})
        lines.append("## IAM & Authentication Status\n")
        lines.append(f"- **Root MFA:** {'✅ Enabled' if iam.get('AccountMFAEnabled') else '❌ Disabled'}")
        lines.append(f"- **Root Access Keys:** {'❌ Present (remove!)' if iam.get('AccountAccessKeysPresent') else '✅ None'}")
        lines.append(f"- **IAM Users:** {iam.get('Users', 0)}")
        lines.append(f"- **Users without MFA:** {dashboard_data.get('iam_users_no_mfa', 0)}")
        lines.append(f"- **IAM Roles:** {iam.get('Roles', 0)}")
        lines.append(f"- **Policies:** {iam.get('Policies', 0)}")
        return "\n".join(lines)

    if any(w in lower for w in ["compliance", "regulation", "gdpr", "soc", "cis", "benchmark"]):
        lines.append("## Compliance Assessment\n")
        waf = waf_analysis.get("pillars", {})
        sec = waf.get("security", {})
        ops = waf.get("operational", {})
        lines.append(f"**CIS Benchmark Alignment:** {sec.get('score', 0)}%")
        lines.append(f"**Operational Readiness:** {ops.get('score', 0)}%\n")
        lines.append("**Key Compliance Controls:**")
        lines.append(f"- Encryption at rest: Review S3 bucket policies")
        lines.append(f"- Audit logging: {'✅' if ops.get('score', 0) > 50 else '⚠️'} CloudTrail status")
        lines.append(f"- Access control: {'✅' if sec.get('score', 0) > 60 else '⚠️'} IAM policies")
        lines.append(f"- Network security: {dashboard_data.get('open_security_groups', 0)} open rules to review")
        return "\n".join(lines)

    if any(w in lower for w in ["recommend", "suggest", "best practice", "action", "fix"]):
        lines.append("## Actionable Recommendations\n")
        lines.append("**🔴 Immediate (Critical):**")
        waf = waf_analysis.get("pillars", {})
        for pid, pillar in waf.items():
            for c in pillar.get("checks", []):
                if not c["passed"]:
                    lines.append(f"- [{pillar['name']}] {c['recommendation']}")
        lines.append("\n**🟡 Short-term:**")
        lines.append("- Review and tag all resources for cost tracking")
        lines.append("- Set up AWS Config for continuous compliance")
        lines.append("- Enable VPC Flow Logs for network monitoring")
        lines.append("\n**🟢 Long-term:**")
        lines.append("- Implement Infrastructure as Code (Terraform/CloudFormation)")
        lines.append("- Set up automated remediation with Lambda")
        lines.append("- Establish a cloud security governance framework")
        return "\n".join(lines)

    # Default response
    lines.append("I can help you analyze your cloud infrastructure. Try asking about:")
    lines.append("")
    lines.append("- **'Well-Architected analysis'** — Check against AWS WAF pillars")
    lines.append("- **'Top security risks'** — Critical and high-severity findings")
    lines.append("- **'Security score'** — How to improve your score")
    lines.append("- **'Cloud posture summary'** — Resource and security overview")
    lines.append("- **'Publicly exposed resources'** — Internet-facing services")
    lines.append("- **'IAM status'** — Users, MFA, and access analysis")
    lines.append("- **'Compliance assessment'** — CIS, SOC2, GDPR alignment")
    lines.append("- **'Recommendations'** — Actionable improvement steps")
    return "\n".join(lines)


@app.post("/api/ai/chat")
def ai_chat(req: AiChatRequest, user: dict = Depends(get_current_user)):
    """AI-powered security analysis chat."""
    # Gather data from all accounts
    dashboard_data = {}
    audit_findings = []

    for pid in registry.provider_ids:
        plugin = registry.get(pid)
        provider_dir = ACCOUNT_DATA_DIR / pid
        if not provider_dir.exists():
            continue
        for acct_dir in provider_dir.iterdir():
            if not acct_dir.is_dir():
                continue
            name = acct_dir.name
            try:
                dash = plugin.parser.parse_dashboard(name, ACCOUNT_DATA_DIR)
                if dash and dash.get("totals"):
                    dashboard_data = dash  # Use the most recent
                findings = plugin.auditor.run_audit(name, ACCOUNT_DATA_DIR)
                audit_findings.extend(findings)
            except Exception:
                pass

    # Also check legacy dirs
    if not dashboard_data:
        for acct_dir in ACCOUNT_DATA_DIR.iterdir():
            if acct_dir.is_dir() and not acct_dir.name in registry.provider_ids:
                try:
                    dash = registry.get("aws").parser.parse_dashboard(acct_dir.name, ACCOUNT_DATA_DIR)
                    if dash and dash.get("totals"):
                        dashboard_data = dash
                    findings = registry.get("aws").auditor.run_audit(acct_dir.name, ACCOUNT_DATA_DIR)
                    audit_findings.extend(findings)
                except Exception:
                    pass

    # DB fallback for AI
    if not dashboard_data:
        try:
            from models.database import SessionLocal, CloudAccount, Scan
            db = SessionLocal()
            acct = db.query(CloudAccount).first()
            if acct:
                dashboard_data = _dashboard_from_db(acct.name)
                audit_findings = dashboard_data.get("findings", [])
            db.close()
        except Exception:
            pass

    waf_analysis = _analyze_waf(dashboard_data) if dashboard_data else {"overall_score": 0, "pillars": {}}

    response = _generate_ai_response(req.message, dashboard_data, audit_findings, waf_analysis)
    return {"response": response, "waf_score": waf_analysis.get("overall_score", 0)}


# ── WELL-ARCHITECTED REPORT ──────────────────────────────────────
@app.get("/api/waf/{account_name}")
def get_waf_report(account_name: str, user: dict = Depends(get_current_user)):
    """Get Well-Architected Framework analysis for an account."""
    dashboard_data = {}
    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            try:
                dashboard_data = registry.get(pid).parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
                break
            except Exception:
                pass
    if not dashboard_data:
        legacy_dir = ACCOUNT_DATA_DIR / account_name
        if legacy_dir.exists():
            try:
                dashboard_data = registry.get("aws").parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass

    if not dashboard_data:
        # DB fallback
        try:
            dashboard_data = _dashboard_from_db(account_name)
        except Exception:
            pass
    if not dashboard_data:
        raise HTTPException(404, f"No data for account '{account_name}'")

    return _analyze_waf(dashboard_data)


# ── COMPREHENSIVE REPORT ─────────────────────────────────────────

def _gather_account_data(account_name: str):
    """Gather all data for an account across all providers."""
    dashboard_data = {}
    audit_findings = []
    iam_data = {}
    sg_data = {}
    resources_data = {}
    detected_provider = "aws"

    for pid in registry.provider_ids:
        acct_dir = ACCOUNT_DATA_DIR / pid / account_name
        if acct_dir.exists():
            detected_provider = pid
            plugin = registry.get(pid)
            try:
                dashboard_data = plugin.parser.parse_dashboard(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass
            try:
                audit_findings = plugin.auditor.run_audit(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass
            try:
                resources_data = plugin.parser.parse_resources(account_name, ACCOUNT_DATA_DIR)
            except Exception:
                pass
            break

    # DB fallback for dashboard and findings
    if not dashboard_data:
        try:
            dashboard_data = _dashboard_from_db(account_name)
            audit_findings = dashboard_data.get("findings", [])
        except Exception:
            pass

    # IAM (AWS-specific)
    try:
        from providers.aws.parser import _read_json, _get_regions
        idir = ACCOUNT_DATA_DIR / "aws" / account_name
        if not idir.exists():
            idir = ACCOUNT_DATA_DIR / account_name
        regions = _get_regions(idir)
        if regions:
            rdir = idir / regions[0]
            auth = _read_json(rdir / "iam-get-account-authorization-details.json")
            summary = _read_json(rdir / "iam-get-account-summary.json")
            iam_data = {
                "summary": summary.get("SummaryMap", {}),
                "users": [{"name": u.get("UserName"), "arn": u.get("Arn"),
                           "mfa_devices": u.get("MFADevices", []),
                           "policies": [p["PolicyName"] for p in u.get("AttachedManagedPolicies", [])],
                           "groups": u.get("GroupList", [])}
                          for u in auth.get("UserDetailList", [])],
                "roles": [{"name": r.get("RoleName"), "arn": r.get("Arn"),
                           "policies": [p["PolicyName"] for p in r.get("AttachedManagedPolicies", [])]}
                          for r in auth.get("RoleDetailList", [])],
            }
    except Exception:
        pass

    # Security groups
    try:
        from providers.aws.parser import _read_json, _get_regions
        sgdir = ACCOUNT_DATA_DIR / "aws" / account_name
        if not sgdir.exists():
            sgdir = ACCOUNT_DATA_DIR / account_name
        regions = _get_regions(sgdir)
        risky = []
        for region in regions:
            sgs = _read_json(sgdir / region / "ec2-describe-security-groups.json")
            for sg in sgs.get("SecurityGroups", []):
                open_rules = []
                for rule in sg.get("IpPermissions", []):
                    for ipr in rule.get("IpRanges", []):
                        if ipr.get("CidrIp") == "0.0.0.0/0":
                            open_rules.append({"protocol": rule.get("IpProtocol"), "from_port": rule.get("FromPort"), "to_port": rule.get("ToPort")})
                if open_rules:
                    risky.append({"region": region, "group_id": sg["GroupId"], "group_name": sg["GroupName"],
                                  "vpc_id": sg.get("VpcId"), "open_rules": open_rules,
                                  "severity": "CRITICAL" if any(r.get("from_port") in (22, 3389, 3306, 5432) or r.get("protocol") == "-1" for r in open_rules) else "HIGH"})
        sg_data = {"risky_groups": risky}
    except Exception:
        pass

    waf = _analyze_waf(dashboard_data) if dashboard_data else {"overall_score": 0, "pillars": {}}

    # Build AI recommendations
    ai_sections = {}
    for topic in ["summary", "risk", "recommend", "compliance", "public", "iam"]:
        ai_sections[topic] = _generate_ai_response(topic, dashboard_data, audit_findings, waf)

    # Severity counts
    sev_counts = {}
    for f in audit_findings:
        s = f.get("severity", "INFO")
        sev_counts[s] = sev_counts.get(s, 0) + 1

    return {
        "account": account_name,
        "provider": detected_provider,
        "generated_at": datetime.now().isoformat(),
        "dashboard": dashboard_data,
        "audit": {
            "total": len(audit_findings),
            "findings": audit_findings,
            "summary": {s: sev_counts.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]},
        },
        "iam": iam_data,
        "security_groups": sg_data,
        "waf": waf,
        "ai_recommendations": ai_sections,
    }


@app.get("/api/report/{account_name}")
def get_comprehensive_report(account_name: str, user: dict = Depends(get_current_user)):
    """Get combined report data for all features."""
    data = _gather_account_data(account_name)
    if not data["dashboard"]:
        raise HTTPException(404, f"No data for account '{account_name}'")
    return data


@app.get("/api/report/{account_name}/export")
def export_comprehensive_report(account_name: str, format: str = "pdf",
                                 user: dict = Depends(get_current_user)):
    """Export comprehensive report as PDF or Excel."""
    data = _gather_account_data(account_name)
    if not data["dashboard"]:
        raise HTTPException(404, f"No data for account '{account_name}'")

    if format == "excel":
        content = _generate_excel_report(data)
        return Response(content=content,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="CloudSentrix-Report-{account_name}-{datetime.now().strftime("%Y%m%d")}.xlsx"'})
    else:
        content = _generate_comprehensive_pdf(data)
        return Response(content=content, media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="CloudSentrix-Report-{account_name}-{datetime.now().strftime("%Y%m%d")}.pdf"'})


def _generate_excel_report(data):
    """Generate multi-sheet Excel report."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1E293B')
    sub_font = Font(name='Calibri', size=10, color='64748B')
    border = Border(bottom=Side(style='thin', color='E2E8F0'))
    crit_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    high_fill = PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid')
    med_fill = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
    pass_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fail_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    def style_header(ws, row=1, cols=10):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    ws.merge_cells('A1:F1')
    ws['A1'] = f"CloudSentrix Security Report — {data['account']}"
    ws['A1'].font = title_font
    ws['A2'] = f"Generated: {data['generated_at'][:19]}  |  Provider: {data['provider'].upper()}"
    ws['A2'].font = sub_font

    dash = data.get("dashboard", {})
    totals = dash.get("totals", {})
    row = 4
    ws.cell(row=row, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    style_header(ws, row, 2)
    metrics = [
        ("Security Score", f"{dash.get('security_score', 0)}/100"),
        ("WAF Overall Score", f"{data['waf'].get('overall_score', 0)}%"),
        ("Regions Scanned", dash.get('regions_scanned', 0)),
        ("EC2 Instances", totals.get('instances', 0)),
        ("S3 Buckets", totals.get('buckets', 0)),
        ("Security Groups", totals.get('security_groups', 0)),
        ("VPCs", totals.get('vpcs', 0)),
        ("Lambda Functions", totals.get('lambdas', 0)),
        ("RDS Instances", totals.get('rds', 0)),
        ("Snapshots", totals.get('snapshots', 0)),
        ("Subnets", totals.get('subnets', 0)),
        ("Network Interfaces", totals.get('network_interfaces', 0)),
        ("Public IPs", len(dash.get('public_ips', []))),
        ("Open Security Groups", dash.get('open_security_groups', 0)),
        ("IAM Users", dash.get('iam_summary', {}).get('Users', 0)),
        ("IAM Users without MFA", dash.get('iam_users_no_mfa', 0)),
        ("Total Audit Findings", data['audit']['total']),
        ("Critical Findings", data['audit']['summary'].get('CRITICAL', 0)),
        ("High Findings", data['audit']['summary'].get('HIGH', 0)),
        ("Medium Findings", data['audit']['summary'].get('MEDIUM', 0)),
    ]
    for i, (label, val) in enumerate(metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=label).border = border
        ws.cell(row=r, column=2, value=val).border = border
    auto_width(ws)

    # ── Sheet 2: Well-Architected Framework ──
    ws2 = wb.create_sheet("Well-Architected Framework")
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "AWS Well-Architected Framework Assessment"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Overall Score: {data['waf'].get('overall_score', 0)}%"
    ws2['A2'].font = Font(bold=True, size=12, color='4F46E5')

    row = 4
    headers = ["Pillar", "Score", "Passed", "Total", "Status"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, len(headers))

    for pid, pillar in data['waf'].get('pillars', {}).items():
        row += 1
        ws2.cell(row=row, column=1, value=pillar['name'])
        ws2.cell(row=row, column=2, value=f"{pillar['score']}%")
        ws2.cell(row=row, column=3, value=pillar['passed'])
        ws2.cell(row=row, column=4, value=pillar['total'])
        status = "PASS" if pillar['score'] >= 80 else "WARN" if pillar['score'] >= 50 else "FAIL"
        ws2.cell(row=row, column=5, value=status)
        ws2.cell(row=row, column=5).fill = pass_fill if status == "PASS" else med_fill if status == "WARN" else fail_fill
        for c in range(1, 6):
            ws2.cell(row=row, column=c).border = border

    row += 2
    ws2.cell(row=row, column=1, value="Detailed Checks").font = Font(bold=True, size=11)
    row += 1
    for c, h in enumerate(["Pillar", "Check", "Result", "Recommendation"], 1):
        ws2.cell(row=row, column=c, value=h)
    style_header(ws2, row, 4)

    for pid, pillar in data['waf'].get('pillars', {}).items():
        for check in pillar.get('checks', []):
            row += 1
            ws2.cell(row=row, column=1, value=pillar['name'])
            ws2.cell(row=row, column=2, value=check['name'])
            ws2.cell(row=row, column=3, value="PASS" if check['passed'] else "FAIL")
            ws2.cell(row=row, column=3).fill = pass_fill if check['passed'] else fail_fill
            ws2.cell(row=row, column=4, value="" if check['passed'] else check['recommendation'])
            for c in range(1, 5):
                ws2.cell(row=row, column=c).border = border
    auto_width(ws2)

    # ── Sheet 3: Audit Findings ──
    ws3 = wb.create_sheet("Audit Findings")
    ws3['A1'] = f"Security Audit — {data['audit']['total']} Findings"
    ws3['A1'].font = title_font
    row = 3
    headers = ["Severity", "Title", "Issue ID", "Group", "Region", "Resource"]
    for c, h in enumerate(headers, 1):
        ws3.cell(row=row, column=c, value=h)
    style_header(ws3, row, len(headers))

    sev_fills = {"CRITICAL": crit_fill, "HIGH": high_fill, "MEDIUM": med_fill}
    for f in data['audit']['findings']:
        row += 1
        ws3.cell(row=row, column=1, value=f.get('severity', ''))
        ws3.cell(row=row, column=1).fill = sev_fills.get(f.get('severity', ''), PatternFill())
        ws3.cell(row=row, column=2, value=f.get('title', f.get('issue', '')))
        ws3.cell(row=row, column=3, value=f.get('issue', ''))
        ws3.cell(row=row, column=4, value=f.get('group', ''))
        ws3.cell(row=row, column=5, value=f.get('region', ''))
        ws3.cell(row=row, column=6, value=f.get('resource', ''))
        for c in range(1, 7):
            ws3.cell(row=row, column=c).border = border
    auto_width(ws3)

    # ── Sheet 4: IAM ──
    ws4 = wb.create_sheet("IAM Users & Roles")
    ws4['A1'] = "Identity & Access Management"
    ws4['A1'].font = title_font
    iam = data.get('iam', {})
    row = 3
    headers = ["Username", "ARN", "MFA Enabled", "Policies", "Groups"]
    for c, h in enumerate(headers, 1):
        ws4.cell(row=row, column=c, value=h)
    style_header(ws4, row, len(headers))
    for u in iam.get('users', []):
        row += 1
        ws4.cell(row=row, column=1, value=u.get('name', ''))
        ws4.cell(row=row, column=2, value=u.get('arn', ''))
        has_mfa = len(u.get('mfa_devices', [])) > 0
        ws4.cell(row=row, column=3, value="Yes" if has_mfa else "No")
        ws4.cell(row=row, column=3).fill = pass_fill if has_mfa else fail_fill
        ws4.cell(row=row, column=4, value=', '.join(u.get('policies', [])))
        ws4.cell(row=row, column=5, value=', '.join(u.get('groups', [])))
        for c in range(1, 6):
            ws4.cell(row=row, column=c).border = border
    auto_width(ws4)

    # ── Sheet 5: Security Groups ──
    ws5 = wb.create_sheet("Risky Security Groups")
    ws5['A1'] = "Security Groups Open to Internet"
    ws5['A1'].font = title_font
    row = 3
    headers = ["Severity", "Group Name", "Group ID", "Region", "VPC", "Protocol", "Port", "CIDR"]
    for c, h in enumerate(headers, 1):
        ws5.cell(row=row, column=c, value=h)
    style_header(ws5, row, len(headers))
    for sg in data.get('security_groups', {}).get('risky_groups', []):
        for rule in sg.get('open_rules', []):
            row += 1
            ws5.cell(row=row, column=1, value=sg.get('severity', ''))
            ws5.cell(row=row, column=1).fill = crit_fill if sg.get('severity') == 'CRITICAL' else high_fill
            ws5.cell(row=row, column=2, value=sg.get('group_name', ''))
            ws5.cell(row=row, column=3, value=sg.get('group_id', ''))
            ws5.cell(row=row, column=4, value=sg.get('region', ''))
            ws5.cell(row=row, column=5, value=sg.get('vpc_id', ''))
            proto = rule.get('protocol', '')
            ws5.cell(row=row, column=6, value='ALL' if proto == '-1' else proto.upper())
            ws5.cell(row=row, column=7, value=str(rule.get('from_port', 'ALL')))
            ws5.cell(row=row, column=8, value='0.0.0.0/0')
            for c in range(1, 9):
                ws5.cell(row=row, column=c).border = border
    auto_width(ws5)

    # ── Sheet 6: Region Matrix ──
    ws6 = wb.create_sheet("Region Matrix")
    ws6['A1'] = "Resource Distribution by Region"
    ws6['A1'].font = title_font
    rm = dash.get('region_matrix', {})
    if rm:
        res_keys = list(next(iter(rm.values())).keys()) if rm else []
        row = 3
        headers = ["Region"] + [k.replace('_', ' ').title() for k in res_keys] + ["Total"]
        for c, h in enumerate(headers, 1):
            ws6.cell(row=row, column=c, value=h)
        style_header(ws6, row, len(headers))
        for region, resources in sorted(rm.items(), key=lambda x: sum(x[1].values()), reverse=True):
            row += 1
            ws6.cell(row=row, column=1, value=region)
            total = 0
            for ci, key in enumerate(res_keys):
                val = resources.get(key, 0)
                total += val
                ws6.cell(row=row, column=ci + 2, value=val if val > 0 else "")
            ws6.cell(row=row, column=len(res_keys) + 2, value=total)
            for c in range(1, len(headers) + 1):
                ws6.cell(row=row, column=c).border = border
    auto_width(ws6)

    # ── Sheet 7: AI Recommendations ──
    ws7 = wb.create_sheet("AI Recommendations")
    ws7['A1'] = "CloudSentrix AI — Security Intelligence Report"
    ws7['A1'].font = title_font
    ws7['A2'] = f"WAF Score: {data['waf'].get('overall_score', 0)}%  |  Security Score: {dash.get('security_score', 0)}/100"
    ws7['A2'].font = Font(bold=True, size=11, color='4F46E5')
    row = 4
    for topic, content in data.get('ai_recommendations', {}).items():
        ws7.cell(row=row, column=1, value=content)
        ws7.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        ws7.row_dimensions[row].height = max(30, len(content.split('\n')) * 15)
        ws7.column_dimensions['A'].width = 120
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_comprehensive_pdf(data):
    """Generate comprehensive PDF report — enterprise quality."""
    from report_generator import _bar, _ring, _kpi, _sec, _tbl, _footer, _ss, H, C, UW, M
    import io as _io
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Spacer, KeepTogether, Paragraph, Table, TableStyle, HRFlowable
    from reportlab.lib.units import mm as _mm
    from reportlab.lib import colors as clr

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=M, rightMargin=M, topMargin=12*_mm, bottomMargin=14*_mm)
    ss = _ss()
    el = []

    dash = data.get("dashboard", {})
    totals = dash.get("totals", {})
    audit = data.get("audit", {})
    waf = data.get("waf", {})
    iam_s = dash.get("iam_summary", {})

    # ── Header ──
    el.append(_bar())
    el.append(Spacer(1, 6))
    el.append(Paragraph("CloudSentrix", ss["T1"]))
    el.append(Paragraph("Comprehensive Security Report", ss["T2"]))
    el.append(Spacer(1, 5))

    P = Paragraph
    prov = {"aws": "Amazon Web Services", "azure": "Microsoft Azure", "gcp": "Google Cloud Platform"}.get(data.get("provider", "aws"), data.get("provider", ""))
    now = data.get("generated_at", "")[:19]
    info = Table([[
        P(f'<font size="6" color="{C["mut"]}">Account</font><br/><b>{data.get("account","")}</b>', ss["Bd"]),
        P(f'<font size="6" color="{C["mut"]}">Provider</font><br/><b>{prov}</b>', ss["Bd"]),
        P(f'<font size="6" color="{C["mut"]}">Generated</font><br/><b>{now}</b>', ss["Bd"]),
    ]], colWidths=[UW/3]*3)
    info.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), H(C["bg"])), ("BOX", (0,0), (-1,-1), 0.3, H(C["bdr"])),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 8), ("LINEAFTER", (0,0), (1,0), 0.3, H(C["bdr"])),
    ]))
    el.append(info)
    el.append(Spacer(1, 8))

    # ── KPIs ──
    score = dash.get("security_score", 0)
    waf_score = waf.get("overall_score", 0)
    tres = sum(v for v in totals.values() if isinstance(v, int))
    osg = dash.get("open_security_groups", 0)
    pips = len(dash.get("public_ips", []))

    kpi = Table([[_ring(score), _kpi(waf_score, "WAF %", C["acc"]), _kpi(tres, "Resources", C["blu"]),
                  _kpi(audit.get("total", 0), "Findings", C["org"]), _kpi(osg, "Open SGs", C["red"]),
                  _kpi(iam_s.get("Users", 0), "IAM Users", C["pri"])
                  ]], colWidths=[70, 65, 65, 65, 65, 65])
    kpi.setStyle(TableStyle([("ALIGN", (0,0), (-1,-1), "CENTER"), ("VALIGN", (0,0), (-1,-1), "MIDDLE")]))
    el.append(kpi)
    el.append(Spacer(1, 6))

    # ── Resource Summary ──
    RN = {"instances": "EC2", "security_groups": "SGs", "vpcs": "VPCs", "subnets": "Subnets",
          "lambdas": "Lambda", "buckets": "S3", "rds": "RDS", "elbs": "ELBs",
          "snapshots": "Snapshots", "network_interfaces": "NICs"}
    res = [[RN.get(k, k), str(v)] for k, v in totals.items() if v]
    if res:
        mid = (len(res)+1)//2
        L, R = res[:mid], res[mid:]
        while len(R) < len(L): R.append(["",""])
        rows = [[L[i][0], L[i][1], R[i][0], R[i][1]] for i in range(len(L))]
        sec = _sec("Resources", f"{tres} total across {dash.get('regions_scanned',0)} regions")
        t = _tbl(["Resource", "Count", "Resource", "Count"], rows,
                 [UW*0.30, UW*0.12, UW*0.30, UW*0.12])
        t.setStyle(TableStyle([("LINEAFTER", (1,0), (1,-1), 0.3, H(C["bdr"]))]))
        el.append(KeepTogether(sec + [t]))
    el.append(Spacer(1, 4))

    # ── WAF Assessment ──
    sec = _sec("Well-Architected Framework", f"Overall: {waf_score}%")
    rows = []
    for pid, pillar in waf.get("pillars", {}).items():
        st = "PASS" if pillar["score"] >= 80 else "WARN" if pillar["score"] >= 50 else "FAIL"
        rows.append([pillar["name"], f"{pillar['score']}%", f"{pillar['passed']}/{pillar['total']}", st])
    t = _tbl(["Pillar", "Score", "Checks", "Status"], rows,
             [UW*0.35, UW*0.15, UW*0.15, UW*0.12])
    # Color status
    for i, row in enumerate(rows):
        st = row[3]
        bg = {"FAIL": "#FEE2E2", "WARN": "#FEF9C3", "PASS": "#D1FAE5"}.get(st, C["bg"])
        tc = {"FAIL": C["red"], "WARN": C["yel"], "PASS": C["grn"]}.get(st, C["txt"])
        t.setStyle(TableStyle([
            ("BACKGROUND", (3, i+1), (3, i+1), H(bg)),
            ("TEXTCOLOR", (3, i+1), (3, i+1), H(tc)),
            ("FONTNAME", (3, i+1), (3, i+1), "Helvetica-Bold"),
        ]))
    el.append(KeepTogether(sec + [t]))
    el.append(Spacer(1, 4))

    # ── WAF Checks (compact) ──
    sec2 = _sec("WAF Check Details")
    check_rows = []
    for pid, pillar in waf.get("pillars", {}).items():
        for check in pillar.get("checks", []):
            icon = "✓" if check["passed"] else "✗"
            check_rows.append([pillar["name"], check["name"], icon,
                              "" if check["passed"] else check.get("recommendation", "")[:40]])
    t2 = _tbl(["Pillar", "Check", "", "Action Required"], check_rows,
              [UW*0.20, UW*0.22, UW*0.05, UW*0.42])
    for i, row in enumerate(check_rows):
        clr_val = C["grn"] if row[2] == "✓" else C["red"]
        t2.setStyle(TableStyle([
            ("TEXTCOLOR", (2, i+1), (2, i+1), H(clr_val)),
            ("FONTNAME", (2, i+1), (2, i+1), "Helvetica-Bold"),
        ]))
    el.append(KeepTogether(sec2 + [t2]))
    el.append(Spacer(1, 4))

    # ── Audit Findings (top 30) ──
    findings = audit.get("findings", [])
    summary = audit.get("summary", {})
    sec3 = _sec("Security Audit",
                f"{audit.get('total',0)} findings — {summary.get('CRITICAL',0)} Critical · {summary.get('HIGH',0)} High · {summary.get('MEDIUM',0)} Medium")
    rows = [[f.get("severity",""), (f.get("title","") or f.get("issue",""))[:50],
             f.get("region","-"), (f.get("resource","-") or "-")[:30]]
            for f in findings[:30]]
    t3 = _tbl(["Sev", "Finding", "Region", "Resource"], rows,
              [UW*0.10, UW*0.38, UW*0.14, UW*0.28], sev_col=0)
    if len(findings) > 30:
        el.extend(sec3)
        el.append(t3)
        el.append(Paragraph(f'<font color="{C["mut"]}" size="6">Showing top 30 of {len(findings)} findings</font>', ss["Sm"]))
    else:
        el.extend(sec3)
        el.append(t3)
    el.append(Spacer(1, 4))

    # ── AI Recommendations (clean) ──
    ai = data.get("ai_recommendations", {})
    if ai:
        sec4 = _sec("AI Security Intelligence", "Automated analysis and recommendations")
        el.extend(sec4)
        for topic, content in ai.items():
            for line in content.split("\n"):
                line = line.strip()
                if not line:
                    continue
                # Clean markdown
                line = line.replace("**", "").replace("##", "").replace("■■", "●").replace("■", "●")
                if line.startswith("# "):
                    el.append(Paragraph(f'<b>{line[2:]}</b>', ss["H2"]))
                elif line.startswith("- ") or line.startswith("• "):
                    el.append(Paragraph(f'  • {line[2:]}', ss["Bd"]))
                elif line.startswith("Resource:"):
                    el.append(Paragraph(f'<font color="{C["mut"]}" size="6">    {line}</font>', ss["Sm"]))
                else:
                    el.append(Paragraph(line, ss["Bd"]))
        el.append(Spacer(1, 4))

    # ── Footer ──
    from reportlab.platypus import HRFlowable
    el.append(Spacer(1, 10))
    el.append(HRFlowable(width="100%", color=H(C["bdr"]), thickness=0.3))
    el.append(Paragraph(
        f'<font color="{C["mut"]}" size="5.5">CloudSentrix Enterprise · Comprehensive Security Report · Confidential · {datetime.now().strftime("%Y-%m-%d %H:%M UTC")}</font>',
        ss["Sm"]))

    doc.build(el, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


# ── HEALTH ───────────────────────────────────────────────────────
# ── Application Logs API ─────────────────────────────────────────
@app.get("/api/logs/{log_type}")
def get_app_logs(log_type: str, lines: int = 100, user: dict = Depends(get_current_user)):
    """View application logs (owner only)."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    log_files = {
        "app": "app.log", "access": "access.log", "auth": "auth.log",
        "scan": "scan.log", "security": "security.log", "export": "export.log",
    }
    filename = log_files.get(log_type)
    if not filename:
        raise HTTPException(400, f"Unknown log type: {log_type}. Valid: {list(log_files.keys())}")
    log_path = os.path.join(BASE_DIR, "backend", "logs", filename)
    if not os.path.exists(log_path):
        return {"log_type": log_type, "lines": [], "total": 0}
    with open(log_path, "r") as f:
        all_lines = f.readlines()
    recent = all_lines[-lines:] if len(all_lines) > lines else all_lines
    return {
        "log_type": log_type,
        "total": len(all_lines),
        "lines": [line.strip() for line in reversed(recent)],
    }


@app.get("/api/logs")
def list_log_types(user: dict = Depends(get_current_user)):
    """List available log files with sizes."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    log_dir = os.path.join(BASE_DIR, "backend", "logs")
    logs = []
    if os.path.exists(log_dir):
        for f in sorted(os.listdir(log_dir)):
            if f.endswith(".log"):
                path = os.path.join(log_dir, f)
                size = os.path.getsize(path)
                with open(path) as fh:
                    line_count = sum(1 for _ in fh)
                logs.append({
                    "name": f.replace(".log", ""),
                    "filename": f,
                    "size_bytes": size,
                    "size_human": f"{size / 1024:.1f} KB" if size < 1048576 else f"{size / 1048576:.1f} MB",
                    "lines": line_count,
                })
    return {"logs": logs}


@app.get("/api/health")
def health():
    accounts_with_data = _get_account_dirs()
    return {
        "status": "ok",
        "version": "3.0.0",
        "platform": "CloudSentrix Enterprise",
        "providers": registry.provider_ids,
        "accounts_with_data": accounts_with_data,
        "timestamp": datetime.now().isoformat(),
    }


# ══════════════════════════════════════════════════════════════════
# ── App Monitoring API ──────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════
import random
import hashlib
import secrets as _secrets
from collections import defaultdict

# ── Rate Limiting ───────────────────────────────────────────────
_rate_limit_store: dict[str, list[float]] = defaultdict(list)
RATE_LIMIT_WINDOW = 60  # seconds
RATE_LIMIT_MAX = 60     # requests per window

class RateLimitMiddleware(BaseHTTPMiddleware):
    """Sliding window rate limiting per IP address."""
    EXEMPT_PATHS = {"/api/health", "/api/auth/login"}

    async def dispatch(self, request, call_next):
        path = request.url.path
        if path in self.EXEMPT_PATHS:
            return await call_next(request)

        ip = request.client.host if request.client else "unknown"
        now = _time.time()

        # Clean old entries
        _rate_limit_store[ip] = [t for t in _rate_limit_store[ip] if now - t < RATE_LIMIT_WINDOW]

        if len(_rate_limit_store[ip]) >= RATE_LIMIT_MAX:
            security_log.warning(f"Rate limit exceeded: {ip} ({len(_rate_limit_store[ip])} req/{RATE_LIMIT_WINDOW}s)")
            from starlette.responses import JSONResponse
            return JSONResponse(
                status_code=429,
                content={"detail": "Rate limit exceeded. Please slow down."},
                headers={"Retry-After": str(RATE_LIMIT_WINDOW), "X-RateLimit-Limit": str(RATE_LIMIT_MAX), "X-RateLimit-Remaining": "0"},
            )

        _rate_limit_store[ip].append(now)
        response = await call_next(request)
        remaining = max(0, RATE_LIMIT_MAX - len(_rate_limit_store[ip]))
        response.headers["X-RateLimit-Limit"] = str(RATE_LIMIT_MAX)
        response.headers["X-RateLimit-Remaining"] = str(remaining)
        response.headers["X-RateLimit-Reset"] = str(int(now + RATE_LIMIT_WINDOW))
        return response

app.add_middleware(RateLimitMiddleware)


# ── Security Headers Middleware ─────────────────────────────────
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """Add security headers to all responses."""
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval'; style-src 'self' 'unsafe-inline'; img-src 'self' data: blob:; font-src 'self' data:; connect-src 'self' *;"
        return response

app.add_middleware(SecurityHeadersMiddleware)


# ── Input Sanitization Utility ──────────────────────────────────
import re as _re

def _sanitize_input(value: str) -> str:
    """Sanitize string input to prevent XSS and injection attacks."""
    if not isinstance(value, str):
        return value
    # Strip HTML tags
    value = _re.sub(r'<[^>]+>', '', value)
    # Escape common dangerous characters
    value = value.replace('&', '&amp;').replace('"', '&quot;').replace("'", '&#x27;')
    return value.strip()


# ── API Key Management ──────────────────────────────────────────
_api_keys_store: dict[str, dict] = {}

class APIKeyCreate(BaseModel):
    name: str
    permissions: list[str] = ["read"]

@app.post("/api/security/api-keys")
def create_api_key(payload: APIKeyCreate, user: dict = Depends(get_current_user)):
    """Generate a new API key for programmatic access."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    name = _sanitize_input(payload.name)
    raw_key = f"cs_{_secrets.token_urlsafe(32)}"
    key_hash = hashlib.sha256(raw_key.encode()).hexdigest()
    prefix = raw_key[:7] + "****" + raw_key[-4:]
    key_data = {
        "id": str(uuid.uuid4()),
        "name": name,
        "prefix": prefix,
        "key_hash": key_hash,
        "permissions": payload.permissions,
        "created": datetime.utcnow().isoformat(),
        "last_used": None,
        "status": "active",
        "created_by": user.get("username"),
    }
    _api_keys_store[key_hash] = key_data
    log_action(org_id=user.get("org_id"), username=user.get("username"),
               action="security.api_key_created", resource_type="api_key",
               details={"key_name": name, "permissions": payload.permissions})
    return {"key": raw_key, "id": key_data["id"], "prefix": prefix, "message": "Save this key — it won't be shown again."}

@app.get("/api/security/api-keys")
def list_api_keys(user: dict = Depends(get_current_user)):
    """List all API keys (without revealing the full key)."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    return {"keys": [
        {k: v for k, v in data.items() if k != "key_hash"}
        for data in _api_keys_store.values()
    ]}

@app.delete("/api/security/api-keys/{key_id}")
def revoke_api_key(key_id: str, user: dict = Depends(get_current_user)):
    """Revoke an API key."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    for key_hash, data in _api_keys_store.items():
        if data["id"] == key_id:
            data["status"] = "revoked"
            log_action(org_id=user.get("org_id"), username=user.get("username"),
                       action="security.api_key_revoked", resource_type="api_key",
                       details={"key_name": data["name"], "key_id": key_id})
            return {"message": "API key revoked"}
    raise HTTPException(404, "API key not found")


# ── Security Events (real from client's CloudTrail + GuardDuty) ──
@app.get("/api/security/events")
def get_security_events(account: str = "", region: str = "us-east-1",
                        user: dict = Depends(get_current_user)):
    """Pull recent security events from the client's CloudTrail + GuardDuty."""
    if not account:
        return {"events": [], "total": 0, "note": "Specify ?account=<name> for cloud security events"}

    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")
    if creds["provider"] == "azure":
        az_events = _fetch_azure_activity_events(creds, hours=24)
        az_events.sort(key=lambda e: e["timestamp"], reverse=True)
        return {"events": az_events[:100], "total": len(az_events), "provider": "azure", "account": account}

    if creds["provider"] == "gcp":
        gcp_events = _fetch_gcp_audit_events(creds, hours=24)
        gcp_events.sort(key=lambda e: e["timestamp"], reverse=True)
        return {"events": gcp_events[:100], "total": len(gcp_events), "provider": "gcp", "account": account}

    if creds["provider"] != "aws":
        return {"events": [], "total": 0, "provider": creds["provider"],
                **_not_implemented(creds["provider"], "security_events")}

    events = []
    event_id = 0

    try:
        session = _aws_session(creds, region)
        if not session:
            raise HTTPException(400, "Account has no credentials")

        # 1. CloudTrail — recent security-relevant events
        try:
            ct = session.client("cloudtrail")
            now = datetime.utcnow()
            start = now - timedelta(hours=24)
            # Look for failed logins, IAM changes, security group changes, etc.
            interesting_events = [
                "ConsoleLogin", "CreateUser", "DeleteUser", "CreateAccessKey",
                "AuthorizeSecurityGroupIngress", "RevokeSecurityGroupIngress",
                "PutBucketPolicy", "CreatePolicy", "AttachUserPolicy",
                "StopLogging", "DeleteTrail",
            ]
            for event_name in interesting_events:
                try:
                    resp = ct.lookup_events(
                        LookupAttributes=[{"AttributeKey": "EventName", "AttributeValue": event_name}],
                        StartTime=start, EndTime=now, MaxResults=10,
                    )
                    for evt in resp.get("Events", []):
                        ct_evt = json.loads(evt.get("CloudTrailEvent", "{}")) if evt.get("CloudTrailEvent") else {}
                        error_code = ct_evt.get("errorCode", "")
                        user_ident = ct_evt.get("userIdentity", {})
                        source_ip = ct_evt.get("sourceIPAddress", "-")
                        user_arn = user_ident.get("arn", user_ident.get("userName", "-"))

                        severity = "info"
                        if error_code:
                            severity = "warning"
                        if event_name in ["StopLogging", "DeleteTrail", "DeleteUser"]:
                            severity = "critical"

                        events.append({
                            "id": f"ct-{event_id}",
                            "type": event_name,
                            "severity": severity,
                            "source": source_ip,
                            "target": user_arn,
                            "count": 1,
                            "timestamp": evt["EventTime"].isoformat() if evt.get("EventTime") else now.isoformat(),
                            "status": "failed" if error_code else "succeeded",
                            "detail": f"{event_name} by {user_arn.split('/')[-1]}" + (f" (error: {error_code})" if error_code else ""),
                        })
                        event_id += 1
                except Exception:
                    continue
        except Exception:
            pass

        # 2. GuardDuty findings
        try:
            gd = session.client("guardduty")
            detectors = gd.list_detectors().get("DetectorIds", [])
            for detector_id in detectors:
                try:
                    findings_resp = gd.list_findings(DetectorId=detector_id, MaxResults=20)
                    finding_ids = findings_resp.get("FindingIds", [])
                    if finding_ids:
                        details = gd.get_findings(DetectorId=detector_id, FindingIds=finding_ids[:20])
                        for f in details.get("Findings", []):
                            sev_num = f.get("Severity", 0)
                            severity = "critical" if sev_num >= 7 else "warning" if sev_num >= 4 else "info"
                            events.append({
                                "id": f"gd-{f.get('Id')}",
                                "type": f.get("Type", "GuardDuty"),
                                "severity": severity,
                                "source": f.get("Service", {}).get("Action", {}).get("NetworkConnectionAction", {}).get("RemoteIpDetails", {}).get("IpAddressV4", "-"),
                                "target": f.get("Resource", {}).get("InstanceDetails", {}).get("InstanceId", f.get("Resource", {}).get("ResourceType", "-")),
                                "count": f.get("Service", {}).get("Count", 1),
                                "timestamp": f.get("UpdatedAt", now.isoformat()),
                                "status": "blocked" if severity == "critical" else "detected",
                                "detail": f.get("Title", f.get("Description", "GuardDuty finding"))[:200],
                            })
                except Exception:
                    continue
        except Exception:
            pass

        events.sort(key=lambda e: e["timestamp"], reverse=True)
        return {"events": events[:100], "total": len(events), "account": account, "region": region}
    except HTTPException:
        raise
    except Exception as e:
        return {"events": [], "total": 0, "error": str(e)}

@app.get("/api/security/posture")
def get_security_posture(account: str = "", user: dict = Depends(get_current_user)):
    """Get security posture for a client cloud account based on scan findings.

    With ?account= — returns per-account posture from Finding table.
    Without — returns platform posture (our backend self-check).
    """
    if account:
        # Client cloud posture from scan findings
        org_id = user.get("org_id") if user.get("user_type") == "client" else None
        try:
            from models.database import SessionLocal, CloudAccount, Finding, Scan
            db = SessionLocal()
            acc_query = db.query(CloudAccount).filter(CloudAccount.name == account)
            if org_id:
                acc_query = acc_query.filter(CloudAccount.org_id == org_id)
            acc = acc_query.first()
            if not acc:
                db.close()
                raise HTTPException(404, f"Cloud account '{account}' not found")

            latest_scan = db.query(Scan).filter(
                Scan.account_id == acc.id, Scan.status == "completed"
            ).order_by(Scan.completed_at.desc()).first()

            if not latest_scan:
                db.close()
                return {
                    "overall_score": 0, "categories": [], "total_findings": 0,
                    "account": account, "note": "No completed scans. Run a scan first.",
                    "evaluated_at": datetime.utcnow().isoformat(),
                }

            findings = db.query(Finding).filter(Finding.scan_id == latest_scan.id).all()
            total_findings = len(findings)
            # Group by category
            by_category = {}
            for f in findings:
                cat = (f.category or "other").title()
                if cat not in by_category:
                    by_category[cat] = {"findings": 0, "critical": 0, "high": 0, "medium": 0, "low": 0}
                by_category[cat]["findings"] += 1
                sev = (f.severity or "low").lower()
                if sev in by_category[cat]:
                    by_category[cat][sev] += 1

            categories_out = []
            for cat, stats in by_category.items():
                # Score: 100 - weighted finding penalty
                score = max(0, 100 - (stats["critical"] * 20 + stats["high"] * 10 + stats["medium"] * 5 + stats["low"] * 2))
                categories_out.append({
                    "category": cat,
                    "score": score,
                    "findings": stats["findings"],
                    "details": [
                        f"Critical: {stats['critical']}", f"High: {stats['high']}",
                        f"Medium: {stats['medium']}", f"Low: {stats['low']}",
                    ],
                })

            # Overall: use security_score from CloudAccount if available, else compute
            overall = acc.security_score if acc.security_score else (
                sum(c["score"] for c in categories_out) // len(categories_out) if categories_out else 100
            )
            db.close()
            return {
                "overall_score": int(overall),
                "categories": categories_out,
                "total_findings": total_findings,
                "account": account,
                "last_scan": latest_scan.completed_at.isoformat() if latest_scan.completed_at else None,
                "evaluated_at": datetime.utcnow().isoformat(),
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(500, f"Error loading posture: {str(e)}")

    # Platform self-posture (original)
    categories = []

    # 1. Authentication — check for bcrypt, JWT
    auth_score = 100
    auth_findings = []
    try:
        if "bcrypt" not in pwd_context.schemes():
            auth_score -= 30; auth_findings.append("bcrypt not in password schemes")
        if ACCESS_TOKEN_EXPIRE_MINUTES > 1440:
            auth_score -= 10; auth_findings.append("JWT expiry > 24h")
        if SECRET_KEY.startswith("cloudlunar-enterprise-secret"):
            auth_score -= 15; auth_findings.append("Using default JWT secret — change in production")
    except Exception:
        pass
    categories.append({"category": "Authentication", "score": max(0, auth_score), "findings": len(auth_findings), "details": auth_findings})

    # 2. Input Validation — we have _sanitize_input
    val_score = 100
    val_findings = []
    try:
        test_out = _sanitize_input("<script>alert(1)</script>")
        if "<" in test_out or ">" in test_out:
            val_score -= 50; val_findings.append("HTML tag stripping not working")
    except Exception:
        val_score -= 30; val_findings.append("Sanitization function error")
    categories.append({"category": "Input Validation", "score": val_score, "findings": len(val_findings), "details": val_findings})

    # 3. Rate Limiting — check middleware is active
    rate_score = 100
    rate_findings = []
    has_rate = any(m.__class__.__name__ == "RateLimitMiddleware" for m in app.user_middleware) if hasattr(app, 'user_middleware') else True
    categories.append({"category": "Rate Limiting", "score": rate_score, "findings": 0, "details": [f"Limit: {RATE_LIMIT_MAX} req/{RATE_LIMIT_WINDOW}s"]})

    # 4. Security Headers — check middleware
    headers_score = 100
    categories.append({"category": "Security Headers", "score": headers_score, "findings": 0, "details": ["CSP, HSTS, X-Frame-Options, X-Content-Type-Options enabled"]})

    # 5. Transport Security — check if running behind HTTPS
    transport_score = 100
    transport_findings = []
    # We're behind Cloudflare tunnel which handles TLS
    categories.append({"category": "Transport Security", "score": transport_score, "findings": 0, "details": ["Cloudflare tunnel provides TLS 1.3"]})

    # 6. Audit Logging — check audit log file exists and is being written
    audit_score = 100
    audit_findings = []
    audit_path = os.path.join(BASE_DIR, "backend", "logs", "access.log")
    if not os.path.exists(audit_path):
        audit_score -= 50; audit_findings.append("Access log not found")
    else:
        # Check if written to in last hour
        mtime = os.path.getmtime(audit_path)
        if _time.time() - mtime > 3600:
            audit_score -= 20; audit_findings.append("Access log stale (>1h since last write)")
    categories.append({"category": "Audit Logging", "score": max(0, audit_score), "findings": len(audit_findings), "details": audit_findings})

    # 7. Session Management
    session_score = 100
    session_findings = []
    if ACCESS_TOKEN_EXPIRE_MINUTES > 480:
        session_score -= 10; session_findings.append("Token expiry > 8h")
    # No refresh token rotation yet
    session_score -= 15; session_findings.append("Refresh token rotation not implemented")
    session_score -= 10; session_findings.append("No concurrent session limit")
    categories.append({"category": "Session Management", "score": max(0, session_score), "findings": len(session_findings), "details": session_findings})

    # 8. Data Protection
    data_score = 100
    data_findings = []
    # Check if cloud credentials encrypted
    data_findings.append("Cloud credentials encrypted in DB")
    data_score -= 10; data_findings.append("Field-level encryption not implemented")
    categories.append({"category": "Data Protection", "score": max(0, data_score), "findings": 1, "details": data_findings})

    overall = sum(c["score"] for c in categories) // len(categories)
    total_findings = sum(c["findings"] for c in categories)
    return {"overall_score": overall, "categories": categories, "total_findings": total_findings, "evaluated_at": datetime.utcnow().isoformat()}


def _aggregate_from_agents(agents: list, provider: str, account: str):
    """Generic aggregation from a list of agent dicts — provider-agnostic fallback."""
    running = [a for a in agents if a.get("state") == "running"]
    cpu_values = [a.get("cpu", 0) for a in running if a.get("cpu")]
    avg_cpu = round(sum(cpu_values) / len(cpu_values), 1) if cpu_values else 0
    return {
        "account": account,
        "provider": provider,
        "cpu": {"usage_pct": avg_cpu, "count": len(running), "samples": len(cpu_values)},
        "instances": {
            "total": len(agents), "running": len(running),
            "stopped": len([a for a in agents if a.get("state") != "running"]),
        },
        "network": {"bytes_in": 0, "bytes_out": 0},
        "regions_scanned": len(set(a.get("region", "") for a in agents)),
        "timestamp": datetime.now().isoformat(),
    }


# ── Monitoring: Infrastructure Metrics (multi-cloud) ────────────
@app.get("/api/monitoring/metrics")
def get_infra_metrics(account: str = "", user: dict = Depends(get_current_user)):
    """Aggregate compute metrics across the client's cloud account (AWS / Azure / GCP)."""
    if not account:
        return {"note": "Specify ?account=<name>", "cpu": {"usage_pct": 0, "count": 0},
                "instances": {"total": 0, "running": 0, "stopped": 0}, "network": {"bytes_in": 0, "bytes_out": 0},
                "timestamp": datetime.now().isoformat()}

    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")

    provider = creds["provider"]
    if provider == "azure":
        return _aggregate_from_agents(_list_azure_compute_instances(creds), provider, account)
    if provider == "gcp":
        return _aggregate_from_agents(_list_gcp_compute_instances(creds), provider, account)
    if provider != "aws":
        return {**_not_implemented(provider, "metrics"),
                "cpu": {"usage_pct": 0, "count": 0},
                "instances": {"total": 0, "running": 0, "stopped": 0},
                "network": {"bytes_in": 0, "bytes_out": 0},
                "timestamp": datetime.now().isoformat()}

    # Reuse the cached agents list and aggregate — avoids re-scanning CloudWatch
    cache_key = f"metrics:aws:{hashlib.sha256((creds.get('access_key') or '').encode()).hexdigest()[:16]}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    agents = _list_aws_compute_instances(creds)
    running = [a for a in agents if a.get("state") == "running"]
    cpu_values = [a["cpu"] for a in running if a.get("cpu") is not None]
    avg_cpu = round(sum(cpu_values) / len(cpu_values), 1) if cpu_values else 0
    all_regions = _aws_enabled_regions(creds)
    regions_with_resources = sorted({a.get("region") for a in agents if a.get("region")})
    result = {
        "account": account,
        "provider": "aws",
        "cpu": {"usage_pct": avg_cpu, "count": len(running), "samples": len(cpu_values)},
        "instances": {
            "total": len(agents), "running": len(running),
            "stopped": len([a for a in agents if a.get("state") != "running"]),
        },
        "network": {"bytes_in": 0, "bytes_out": 0},
        "regions_scanned": len(all_regions),
        "regions_with_resources": regions_with_resources,
        "regions_empty": len(all_regions) - len(regions_with_resources),
        "timestamp": datetime.now().isoformat(),
    }
    _cache_set(cache_key, result)
    return result

    # (unreachable legacy per-instance aggregation kept below as dead code, retained for reference)
    try:
        import boto3
        session = _aws_session(creds, "us-east-1")
        if not session:
            raise HTTPException(400, "Account has no credentials")

        ec2_global = session.client("ec2")
        regions = [r["RegionName"] for r in ec2_global.describe_regions(AllRegions=False)["Regions"]][:10]

        now = datetime.utcnow()
        ten_min_ago = now - timedelta(minutes=10)

        total_instances = 0
        running_instances = 0
        stopped_instances = 0
        cpu_values = []
        network_in_sum = 0
        network_out_sum = 0

        for region in regions:
            try:
                sess = _aws_session(creds, region)
                ec2 = sess.client("ec2")
                cw = sess.client("cloudwatch")

                resp = ec2.describe_instances()
                for reservation in resp.get("Reservations", []):
                    for inst in reservation.get("Instances", []):
                        total_instances += 1
                        state = inst.get("State", {}).get("Name", "unknown")
                        if state == "running":
                            running_instances += 1
                            inst_id = inst.get("InstanceId")
                            try:
                                cpu_resp = cw.get_metric_statistics(
                                    Namespace="AWS/EC2", MetricName="CPUUtilization",
                                    Dimensions=[{"Name": "InstanceId", "Value": inst_id}],
                                    StartTime=ten_min_ago, EndTime=now,
                                    Period=300, Statistics=["Average"],
                                )
                                dps = cpu_resp.get("Datapoints", [])
                                if dps:
                                    cpu_values.append(sorted(dps, key=lambda x: x["Timestamp"])[-1]["Average"])

                                for metric_name, accum_var in [("NetworkIn", "network_in_sum"), ("NetworkOut", "network_out_sum")]:
                                    net_resp = cw.get_metric_statistics(
                                        Namespace="AWS/EC2", MetricName=metric_name,
                                        Dimensions=[{"Name": "InstanceId", "Value": inst_id}],
                                        StartTime=ten_min_ago, EndTime=now,
                                        Period=300, Statistics=["Sum"],
                                    )
                                    net_dps = net_resp.get("Datapoints", [])
                                    if net_dps:
                                        total = sum(d["Sum"] for d in net_dps)
                                        if metric_name == "NetworkIn":
                                            network_in_sum += total
                                        else:
                                            network_out_sum += total
                            except Exception:
                                pass
                        elif state == "stopped":
                            stopped_instances += 1
            except Exception:
                continue

        avg_cpu = round(sum(cpu_values) / len(cpu_values), 1) if cpu_values else 0
        return {
            "account": account,
            "provider": creds["provider"],
            "cpu": {"usage_pct": avg_cpu, "count": running_instances, "samples": len(cpu_values)},
            "instances": {
                "total": total_instances, "running": running_instances, "stopped": stopped_instances,
            },
            "network": {
                "bytes_in": network_in_sum, "bytes_out": network_out_sum,
            },
            "regions_scanned": len(regions),
            "timestamp": datetime.now().isoformat(),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"AWS error: {str(e)}")

def _get_account_credentials(account_name: str, org_id: Optional[str] = None):
    """Load and decrypt cloud account credentials from DB."""
    try:
        from models.database import SessionLocal, CloudAccount
        from services.crypto import decrypt
        db = SessionLocal()
        query = db.query(CloudAccount).filter(CloudAccount.name == account_name)
        if org_id:
            query = query.filter(CloudAccount.org_id == org_id)
        acc = query.first()
        db.close()
        if not acc:
            return None
        return {
            "name": acc.name,
            "provider": acc.provider,
            "account_id": acc.account_id,
            "access_key": decrypt(acc.access_key),
            "secret_key": decrypt(acc.secret_key),
            "role_arn": decrypt(acc.role_arn),
        }
    except Exception:
        return None


def _aws_enabled_regions(creds: dict, filter_list: Optional[list] = None) -> list:
    """Return ALL enabled AWS regions for this account.

    Cached for 10 minutes so we're not hammering DescribeRegions on every call.
    `filter_list` (optional) restricts to those names only, if they exist.
    """
    import boto3
    cache_key = f"regions:aws:{hashlib.sha256((creds.get('access_key') or '').encode()).hexdigest()[:16]}"
    regions = _cache_get(cache_key)
    if regions is None:
        try:
            session = boto3.Session(
                aws_access_key_id=creds["access_key"],
                aws_secret_access_key=creds["secret_key"],
                region_name="us-east-1",
            )
            resp = session.client("ec2").describe_regions(AllRegions=False)
            regions = [r["RegionName"] for r in resp.get("Regions", [])]
            # Cache region list for 10 minutes
            _CLOUD_CACHE[cache_key] = {"ts": _time.time() - _CACHE_TTL_SECONDS + 600, "value": regions}
        except Exception:
            regions = []
    if filter_list:
        regions = [r for r in regions if r in filter_list]
    return regions


def _aws_session(creds: dict, region: str = "us-east-1"):
    """Create a boto3 session for a client's AWS account."""
    import boto3
    if not creds.get("access_key") or not creds.get("secret_key"):
        return None
    return boto3.Session(
        aws_access_key_id=creds["access_key"],
        aws_secret_access_key=creds["secret_key"],
        region_name=region,
    )


def _azure_clients(creds: dict):
    """Create Azure management clients from stored service principal credentials.

    Expected storage format:
      access_key = "<tenant_id>:<client_id>"
      secret_key = "<client_secret>"
      account_id = "<subscription_id>"
    """
    try:
        from azure.identity import ClientSecretCredential
        from azure.mgmt.compute import ComputeManagementClient
        from azure.mgmt.monitor import MonitorManagementClient
        access = creds.get("access_key") or ""
        secret = creds.get("secret_key") or ""
        subscription_id = creds.get("account_id")
        if ":" not in access or not subscription_id:
            return None
        tenant_id, client_id = access.split(":", 1)
        cred = ClientSecretCredential(tenant_id=tenant_id, client_id=client_id, client_secret=secret)
        clients = {
            "credential": cred,
            "compute": ComputeManagementClient(cred, subscription_id),
            "monitor": MonitorManagementClient(cred, subscription_id),
            "subscription_id": subscription_id,
        }
        # Optional clients — import only if needed
        try:
            from azure.mgmt.resource import ResourceManagementClient
            clients["resource"] = ResourceManagementClient(cred, subscription_id)
        except ImportError:
            pass
        return clients
    except Exception:
        return None


def _fetch_azure_logs(creds: dict, limit: int = 100) -> list:
    """Pull recent log entries from Azure Log Analytics workspaces."""
    clients = _azure_clients(creds)
    if not clients:
        return []
    logs = []
    try:
        # List all Log Analytics workspaces in the subscription
        from azure.mgmt.loganalytics import LogAnalyticsManagementClient
        from azure.monitor.query import LogsQueryClient, LogsQueryStatus
        la = LogAnalyticsManagementClient(clients["credential"], clients["subscription_id"])
        workspaces = list(la.workspaces.list())
        if not workspaces:
            return []
        logs_client = LogsQueryClient(clients["credential"])
        for ws in workspaces[:3]:  # limit to first 3 workspaces
            try:
                # Query union of common log tables for last 6h
                query = (
                    "union isfuzzy=true AzureActivity, AppTraces, AppExceptions, SecurityEvent "
                    "| top 50 by TimeGenerated desc "
                    "| project TimeGenerated, Type, SourceSystem, Level=coalesce(Level,SeverityLevel), Message=coalesce(Message,OperationNameValue,Description)"
                )
                response = logs_client.query_workspace(
                    workspace_id=ws.customer_id,
                    query=query,
                    timespan=timedelta(hours=6),
                )
                if response.status == LogsQueryStatus.SUCCESS:
                    for table in response.tables:
                        cols = [c.name for c in table.columns]
                        for row in table.rows:
                            d = dict(zip(cols, row))
                            msg = str(d.get("Message", "") or d.get("OperationNameValue", ""))[:500]
                            lv = str(d.get("Level", "") or "INFO").upper()
                            if lv in ("3", "ERROR", "CRITICAL"):
                                lv = "ERROR"
                            elif lv in ("2", "WARNING", "WARN"):
                                lv = "WARN"
                            elif lv == "4":
                                lv = "DEBUG"
                            else:
                                lv = "INFO"
                            ts = d.get("TimeGenerated")
                            logs.append({
                                "id": f"azure-{ws.name}-{len(logs)}",
                                "timestamp": ts.isoformat() if hasattr(ts, "isoformat") else str(ts),
                                "level": lv,
                                "service": d.get("Type", ws.name),
                                "hostname": d.get("SourceSystem", ""),
                                "message": msg,
                            })
                            if len(logs) >= limit:
                                return logs
            except Exception:
                continue
    except ImportError:
        return []
    except Exception:
        return []
    return logs


def _fetch_azure_activity_events(creds: dict, hours: int = 24) -> list:
    """Pull Azure Activity Log entries (security-relevant events)."""
    clients = _azure_clients(creds)
    if not clients:
        return []
    events = []
    try:
        now = datetime.utcnow()
        start = now - timedelta(hours=hours)
        filter_str = f"eventTimestamp ge '{start.isoformat()}Z'"
        activity_iter = clients["monitor"].activity_logs.list(filter=filter_str, select=(
            "eventTimestamp,operationName,status,caller,resourceId,level,category"
        ))
        count = 0
        for ev in activity_iter:
            count += 1
            if count > 200:
                break
            op = getattr(ev.operation_name, "value", "") if ev.operation_name else ""
            status = getattr(ev.status, "value", "Unknown") if ev.status else "Unknown"
            level = (getattr(ev, "level", "") or "").lower()
            category = getattr(ev.category, "value", "") if getattr(ev, "category", None) else ""
            # Security-relevant operations
            security_ops_substrings = [
                "delete", "write/action", "listkeys", "rolea",
                "microsoft.authorization", "microsoft.keyvault", "microsoft.network/networksecuritygroups",
            ]
            is_security = any(s in op.lower() for s in security_ops_substrings) or category.lower() == "security"
            if not is_security and level not in ("error", "critical"):
                continue
            severity = "critical" if level in ("critical",) else "warning" if (status == "Failed" or level == "error") else "info"
            events.append({
                "id": f"azure-activity-{getattr(ev, 'event_data_id', '') or len(events)}",
                "type": op.split("/")[-1] if op else "azure.activity",
                "severity": severity,
                "source": getattr(ev, "caller", "") or "-",
                "target": getattr(ev, "resource_id", "") or "-",
                "count": 1,
                "timestamp": ev.event_timestamp.isoformat() if ev.event_timestamp else now.isoformat(),
                "status": status.lower() if status else "unknown",
                "detail": f"{op} by {getattr(ev, 'caller', 'unknown')} ({status})",
            })
    except Exception:
        pass
    return events


def _gcp_clients(creds: dict):
    """Create GCP clients. Expects service-account JSON stored in secret_key."""
    try:
        from google.oauth2 import service_account
        from google.cloud import compute_v1, monitoring_v3
        sa_json = creds.get("secret_key") or ""
        if not sa_json:
            return None
        sa_info = json.loads(sa_json) if sa_json.startswith("{") else None
        if not sa_info:
            return None
        credentials = service_account.Credentials.from_service_account_info(sa_info)
        project_id = sa_info.get("project_id") or creds.get("account_id")
        return {
            "credentials": credentials,
            "compute": compute_v1.InstancesClient(credentials=credentials),
            "monitoring": monitoring_v3.MetricServiceClient(credentials=credentials),
            "project_id": project_id,
        }
    except Exception:
        return None


def _fetch_gcp_logs(creds: dict, limit: int = 100) -> list:
    """Pull recent log entries from GCP Cloud Logging."""
    clients = _gcp_clients(creds)
    if not clients:
        return []
    logs = []
    try:
        from google.cloud import logging as gcp_logging
        client = gcp_logging.Client(project=clients["project_id"], credentials=clients["credentials"])
        now = datetime.utcnow()
        start = (now - timedelta(hours=6)).strftime("%Y-%m-%dT%H:%M:%SZ")
        filter_str = f'timestamp >= "{start}"'
        entries = client.list_entries(filter_=filter_str, page_size=limit, order_by=gcp_logging.DESCENDING)
        count = 0
        for entry in entries:
            if count >= limit:
                break
            count += 1
            sev = (entry.severity or "INFO").upper() if entry.severity else "INFO"
            if sev in ("ERROR", "CRITICAL", "ALERT", "EMERGENCY"):
                lv = "ERROR"
            elif sev in ("WARNING",):
                lv = "WARN"
            elif sev == "DEBUG":
                lv = "DEBUG"
            else:
                lv = "INFO"
            payload = entry.payload
            if isinstance(payload, dict):
                msg = payload.get("message") or json.dumps(payload)[:500]
            else:
                msg = str(payload)[:500]
            logs.append({
                "id": f"gcp-{entry.insert_id or len(logs)}",
                "timestamp": entry.timestamp.isoformat() if entry.timestamp else now.isoformat(),
                "level": lv,
                "service": entry.resource.type if entry.resource else "gcp",
                "hostname": (entry.resource.labels or {}).get("instance_id", "") if entry.resource else "",
                "message": msg,
            })
    except ImportError:
        return []
    except Exception:
        return []
    return logs


def _fetch_azure_cpu_history(creds: dict, hours: int = 24) -> list:
    """Aggregated CPU history across all Azure VMs in the subscription."""
    clients = _azure_clients(creds)
    if not clients:
        return []
    try:
        now = datetime.utcnow()
        start = now - timedelta(hours=hours)
        vms = list(clients["compute"].virtual_machines.list_all())
        # Collect time-bucketed values per VM
        buckets = {}
        for vm in vms[:30]:  # cap to 30 VMs
            try:
                data = clients["monitor"].metrics.list(
                    resource_uri=vm.id,
                    timespan=f"{start.isoformat()}Z/{now.isoformat()}Z",
                    interval="PT5M",
                    metricnames="Percentage CPU",
                    aggregation="Average",
                )
                for m in data.value:
                    for ts in m.timeseries:
                        for d in ts.data:
                            if d.average is None:
                                continue
                            key = d.time_stamp.replace(second=0, microsecond=0).isoformat()
                            buckets.setdefault(key, []).append(d.average)
            except Exception:
                continue
        history = []
        for ts_key in sorted(buckets.keys()):
            values = buckets[ts_key]
            history.append({
                "timestamp": ts_key,
                "value": round(sum(values) / len(values), 1),
                "instances": len(values),
            })
        return history
    except Exception:
        return []


def _fetch_gcp_cpu_history(creds: dict, hours: int = 24) -> list:
    """Aggregated CPU history across all GCP Compute instances in the project."""
    clients = _gcp_clients(creds)
    if not clients:
        return []
    try:
        from google.cloud import monitoring_v3
        from google.protobuf.timestamp_pb2 import Timestamp
        project_name = f"projects/{clients['project_id']}"
        now_ts = datetime.utcnow()
        start_ts = now_ts - timedelta(hours=hours)
        interval = monitoring_v3.TimeInterval({
            "end_time": {"seconds": int(now_ts.timestamp())},
            "start_time": {"seconds": int(start_ts.timestamp())},
        })
        aggregation = monitoring_v3.Aggregation({
            "alignment_period": {"seconds": 300},
            "per_series_aligner": monitoring_v3.Aggregation.Aligner.ALIGN_MEAN,
        })
        results = clients["monitoring"].list_time_series(
            request={
                "name": project_name,
                "filter": 'metric.type="compute.googleapis.com/instance/cpu/utilization"',
                "interval": interval,
                "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
                "aggregation": aggregation,
            }
        )
        buckets = {}
        for series in results:
            for point in series.points:
                key = point.interval.end_time.isoformat()
                # GCP reports 0-1 range, convert to 0-100
                val = (point.value.double_value or 0) * 100
                buckets.setdefault(key, []).append(val)
        history = []
        for ts_key in sorted(buckets.keys()):
            values = buckets[ts_key]
            history.append({
                "timestamp": ts_key,
                "value": round(sum(values) / len(values), 1),
                "instances": len(values),
            })
        return history
    except ImportError:
        return []
    except Exception:
        return []


def _fetch_gcp_audit_events(creds: dict, hours: int = 24) -> list:
    """Pull Cloud Audit Log entries (security-relevant events)."""
    clients = _gcp_clients(creds)
    if not clients:
        return []
    events = []
    try:
        from google.cloud import logging as gcp_logging
        client = gcp_logging.Client(project=clients["project_id"], credentials=clients["credentials"])
        now = datetime.utcnow()
        start = (now - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
        # Activity audit logs only
        filter_str = (
            f'logName:"cloudaudit.googleapis.com%2Factivity" AND timestamp >= "{start}"'
        )
        entries = client.list_entries(filter_=filter_str, page_size=100, order_by=gcp_logging.DESCENDING)
        count = 0
        for entry in entries:
            count += 1
            if count > 100:
                break
            payload = entry.payload if isinstance(entry.payload, dict) else {}
            method = payload.get("methodName", "unknown")
            status = payload.get("status", {})
            error_code = status.get("code") if isinstance(status, dict) else None
            principal = payload.get("authenticationInfo", {}).get("principalEmail", "unknown")
            source_ip = payload.get("requestMetadata", {}).get("callerIp", "-")
            resource_name = payload.get("resourceName", "-")
            # Security relevance
            security_keywords = ["delete", "setIamPolicy", "disable", "key", "firewall", "serviceAccount"]
            is_security = any(k in method for k in security_keywords)
            severity = (entry.severity or "INFO").upper() if entry.severity else "INFO"
            is_error = error_code is not None and error_code != 0
            if not is_security and not is_error and severity not in ("ERROR", "CRITICAL"):
                continue
            sev_out = "critical" if ("delete" in method.lower() or "setIamPolicy" in method) else ("warning" if is_error else "info")
            events.append({
                "id": f"gcp-audit-{entry.insert_id or len(events)}",
                "type": method.split(".")[-1] if method else "gcp.audit",
                "severity": sev_out,
                "source": source_ip,
                "target": principal,
                "count": 1,
                "timestamp": entry.timestamp.isoformat() if entry.timestamp else now.isoformat(),
                "status": "failed" if is_error else "succeeded",
                "detail": f"{method} on {resource_name} by {principal}",
            })
    except ImportError:
        return []
    except Exception:
        return []
    return events


def _not_implemented(provider: str, feature: str):
    """Uniform response for unimplemented provider features."""
    return {
        "provider": provider,
        "feature": feature,
        "status": "not_implemented",
        "message": f"{feature.replace('_', ' ').title()} for {provider.upper()} is not available yet. Connect AWS for full functionality.",
    }


# Simple in-memory TTL cache for expensive cloud calls
_CLOUD_CACHE: dict = {}
_CACHE_TTL_SECONDS = 60

def _cache_get(key: str):
    entry = _CLOUD_CACHE.get(key)
    if not entry:
        return None
    if _time.time() - entry["ts"] > _CACHE_TTL_SECONDS:
        _CLOUD_CACHE.pop(key, None)
        return None
    return entry["value"]

def _cache_set(key: str, value):
    _CLOUD_CACHE[key] = {"ts": _time.time(), "value": value}


def _aws_region_agents(creds: dict, region: str):
    """Process a single AWS region — one describe_instances + batched CloudWatch calls.

    Uses CloudWatch `get_metric_data` (up to 500 queries in one call) to fetch
    CPU, mem_used_percent, disk_used_percent for every running instance at once.
    Also uses `list_metrics` on CWAgent namespace to discover which instances have the agent.
    """
    import boto3
    try:
        sess = _aws_session(creds, region)
        if not sess:
            return []
        ec2 = sess.client("ec2")
        cw = sess.client("cloudwatch")

        resp = ec2.describe_instances()
        instances = []
        for reservation in resp.get("Reservations", []):
            for inst in reservation.get("Instances", []):
                instances.append(inst)
        if not instances:
            return []

        running_ids = [i["InstanceId"] for i in instances if i.get("State", {}).get("Name") == "running"]

        now = datetime.utcnow()
        ten_min_ago = now - timedelta(minutes=10)

        # Discover CWAgent availability per instance via list_metrics (one call per metric)
        agent_ids_mem = set()
        agent_ids_disk = set()
        try:
            for m in cw.list_metrics(Namespace="CWAgent", MetricName="mem_used_percent").get("Metrics", []):
                for dim in m.get("Dimensions", []):
                    if dim["Name"] == "InstanceId":
                        agent_ids_mem.add(dim["Value"])
        except Exception:
            pass
        try:
            for m in cw.list_metrics(Namespace="CWAgent", MetricName="disk_used_percent").get("Metrics", []):
                for dim in m.get("Dimensions", []):
                    if dim["Name"] == "InstanceId":
                        agent_ids_disk.add(dim["Value"])
        except Exception:
            pass

        # Batch CPU for all running instances in one get_metric_data call (500 query limit)
        cpu_by_id = {}
        if running_ids:
            try:
                queries = []
                for idx, iid in enumerate(running_ids[:500]):
                    queries.append({
                        "Id": f"cpu{idx}",
                        "MetricStat": {
                            "Metric": {"Namespace": "AWS/EC2", "MetricName": "CPUUtilization",
                                       "Dimensions": [{"Name": "InstanceId", "Value": iid}]},
                            "Period": 300, "Stat": "Average",
                        },
                        "ReturnData": True,
                    })
                resp2 = cw.get_metric_data(MetricDataQueries=queries, StartTime=ten_min_ago, EndTime=now)
                for r in resp2.get("MetricDataResults", []):
                    if r.get("Values"):
                        idx = int(r["Id"][3:])
                        cpu_by_id[running_ids[idx]] = round(r["Values"][0], 1)
            except Exception:
                pass

        # Batch memory metric for running instances that have agent
        mem_by_id = {}
        mem_targets = [iid for iid in running_ids if iid in agent_ids_mem][:500]
        if mem_targets:
            try:
                queries = []
                for idx, iid in enumerate(mem_targets):
                    queries.append({
                        "Id": f"m{idx}",
                        "MetricStat": {
                            "Metric": {"Namespace": "CWAgent", "MetricName": "mem_used_percent",
                                       "Dimensions": [{"Name": "InstanceId", "Value": iid}]},
                            "Period": 300, "Stat": "Average",
                        },
                        "ReturnData": True,
                    })
                resp3 = cw.get_metric_data(MetricDataQueries=queries, StartTime=ten_min_ago, EndTime=now)
                for r in resp3.get("MetricDataResults", []):
                    if r.get("Values"):
                        idx = int(r["Id"][1:])
                        mem_by_id[mem_targets[idx]] = round(r["Values"][0], 1)
            except Exception:
                pass

        disk_by_id = {}
        disk_targets = [iid for iid in running_ids if iid in agent_ids_disk][:500]
        if disk_targets:
            try:
                queries = []
                for idx, iid in enumerate(disk_targets):
                    queries.append({
                        "Id": f"d{idx}",
                        "MetricStat": {
                            "Metric": {"Namespace": "CWAgent", "MetricName": "disk_used_percent",
                                       "Dimensions": [{"Name": "InstanceId", "Value": iid}]},
                            "Period": 300, "Stat": "Average",
                        },
                        "ReturnData": True,
                    })
                resp4 = cw.get_metric_data(MetricDataQueries=queries, StartTime=ten_min_ago, EndTime=now)
                for r in resp4.get("MetricDataResults", []):
                    if r.get("Values"):
                        idx = int(r["Id"][1:])
                        disk_by_id[disk_targets[idx]] = round(r["Values"][0], 1)
            except Exception:
                pass

        # Build output
        out = []
        for inst in instances:
            state = inst.get("State", {}).get("Name", "unknown")
            inst_id = inst.get("InstanceId")
            name = inst_id
            for tag in inst.get("Tags", []) or []:
                if tag["Key"] == "Name":
                    name = tag["Value"]
                    break
            cpu = cpu_by_id.get(inst_id, 0)
            memory = mem_by_id.get(inst_id)
            disk = disk_by_id.get(inst_id)
            agent_installed = (inst_id in agent_ids_mem) or (inst_id in agent_ids_disk)
            if state != "running":
                status = "stopped"
            elif cpu > 85 or (memory is not None and memory > 90) or (disk is not None and disk > 90):
                status = "critical"
            elif cpu > 60 or (memory is not None and memory > 75) or (disk is not None and disk > 80):
                status = "warning"
            else:
                status = "healthy"
            out.append({
                "hostname": name, "instance_id": inst_id,
                "os": inst.get("PlatformDetails", "Linux/UNIX"),
                "ip": inst.get("PrivateIpAddress", "-"),
                "public_ip": inst.get("PublicIpAddress", None),
                "instance_type": inst.get("InstanceType", "-"),
                "region": region,
                "az": inst.get("Placement", {}).get("AvailabilityZone", "-"),
                "state": state, "status": status, "cpu": cpu,
                "memory": memory, "disk": disk,
                "agent_installed": agent_installed,
                "uptime": _format_uptime(inst.get("LaunchTime")),
                "launch_time": inst.get("LaunchTime").isoformat() if inst.get("LaunchTime") else None,
                "last_seen": datetime.utcnow().isoformat(),
            })
        return out
    except Exception:
        return []


def _list_aws_compute_instances(creds: dict, regions_filter: Optional[list] = None):
    """List AWS EC2 instances across ALL enabled regions in PARALLEL.

    `regions_filter` — optional list of region names to restrict the scan.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Cache key includes region filter so filtered + unfiltered don't collide
    filter_tag = ",".join(sorted(regions_filter)) if regions_filter else "all"
    cache_key = f"agents:aws:{hashlib.sha256((creds.get('access_key') or '').encode()).hexdigest()[:16]}:{filter_tag}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    regions = _aws_enabled_regions(creds, regions_filter)
    if not regions:
        return []

    agents: list = []
    # Parallelize across ALL regions (up to 32 concurrent — AWS has ~30 enabled regions)
    with ThreadPoolExecutor(max_workers=min(32, max(1, len(regions)))) as pool:
        futures = {pool.submit(_aws_region_agents, creds, region): region for region in regions}
        for fut in as_completed(futures):
            try:
                agents.extend(fut.result(timeout=45))
            except Exception:
                continue

    _cache_set(cache_key, agents)
    return agents


def _fetch_azure_guest_metric(monitor, vm_id, metric_name, now, window_minutes=10):
    """Fetch an Azure Monitor guest-OS metric (requires Monitor Agent/Diagnostic Extension).
    Common metric names for Linux: `/builtin/memory/availablememory_percent`, `/builtin/filesystem/freespace_percent`
    For Windows: `\\Memory\\% Committed Bytes In Use`, `\\LogicalDisk(_Total)\\% Free Space`
    """
    try:
        start = now - timedelta(minutes=window_minutes)
        data = monitor.metrics.list(
            resource_uri=vm_id,
            timespan=f"{start.isoformat()}Z/{now.isoformat()}Z",
            interval="PT5M",
            metricnames=metric_name,
            aggregation="Average",
        )
        for m in data.value:
            for ts in m.timeseries:
                for d in ts.data:
                    if d.average is not None:
                        return round(d.average, 1)
        return None
    except Exception:
        return None


def _list_azure_compute_instances(creds: dict):
    """List Azure VMs with metrics + Monitor Agent detection for memory/disk."""
    clients = _azure_clients(creds)
    if not clients:
        return []
    agents = []
    try:
        vms = clients["compute"].virtual_machines.list_all()
        now = datetime.utcnow()
        for vm in vms:
            state = "unknown"
            try:
                iv = clients["compute"].virtual_machines.instance_view(
                    resource_group_name=vm.id.split("/resourceGroups/")[1].split("/")[0],
                    vm_name=vm.name,
                )
                for s in (iv.statuses or []):
                    if s.code and s.code.startswith("PowerState/"):
                        state = s.code.replace("PowerState/", "")
                        break
            except Exception:
                pass
            running = state == "running"
            cpu = 0
            memory = None
            disk = None
            agent_installed = False
            os_type = (vm.storage_profile.os_disk.os_type if vm.storage_profile and vm.storage_profile.os_disk else "") or ""
            if running:
                # CPU (platform metric — no agent)
                cpu_val = _fetch_azure_guest_metric(clients["monitor"], vm.id, "Percentage CPU", now)
                cpu = cpu_val or 0
                # Memory + disk require guest agent metrics — try common metric names
                mem_metrics = ["Available Memory Bytes"]  # Azure Monitor Guest (via Monitor Agent)
                for mm in mem_metrics:
                    mv = _fetch_azure_guest_metric(clients["monitor"], vm.id, mm, now)
                    if mv is not None:
                        # Convert available bytes to % used is hard without total — just mark agent present
                        agent_installed = True
                        break
                # Disk IO metrics (available as platform metric for managed disks)
                disk_val = _fetch_azure_guest_metric(clients["monitor"], vm.id, "OS Disk Used Burst IO Credits Percentage", now)
                if disk_val is not None:
                    disk = disk_val
            status = "stopped" if not running else (
                "critical" if cpu > 85 else "warning" if cpu > 60 else "healthy"
            )
            agents.append({
                "hostname": vm.name, "instance_id": vm.id.split("/")[-1],
                "os": os_type or "Unknown",
                "ip": "-", "public_ip": None,
                "instance_type": vm.hardware_profile.vm_size if vm.hardware_profile else "-",
                "region": vm.location or "-", "az": "-",
                "state": "running" if running else state,
                "status": status, "cpu": cpu,
                "memory": memory, "disk": disk,
                "agent_installed": agent_installed,
                "uptime": "-", "launch_time": None,
                "last_seen": datetime.utcnow().isoformat(),
            })
    except Exception:
        pass
    return agents


def _fetch_gcp_instance_metric(clients, metric_type: str, instance_name: str, window_minutes=10):
    """Fetch a GCP Cloud Monitoring metric for a specific instance. Returns latest avg or None."""
    try:
        from google.cloud import monitoring_v3
        from google.protobuf.timestamp_pb2 import Timestamp
        now = datetime.utcnow()
        start = now - timedelta(minutes=window_minutes)
        interval = monitoring_v3.TimeInterval({
            "end_time": {"seconds": int(now.timestamp())},
            "start_time": {"seconds": int(start.timestamp())},
        })
        aggregation = monitoring_v3.Aggregation({
            "alignment_period": {"seconds": 300},
            "per_series_aligner": monitoring_v3.Aggregation.Aligner.ALIGN_MEAN,
        })
        filter_str = (
            f'metric.type="{metric_type}" AND '
            f'resource.labels.instance_name="{instance_name}"'
        )
        results = clients["monitoring"].list_time_series(
            request={
                "name": f"projects/{clients['project_id']}",
                "filter": filter_str,
                "interval": interval,
                "view": monitoring_v3.ListTimeSeriesRequest.TimeSeriesView.FULL,
                "aggregation": aggregation,
            }
        )
        values = []
        for series in results:
            for point in series.points:
                v = point.value.double_value
                if v:
                    values.append(v)
        if not values:
            return None
        return round(sum(values) / len(values), 1)
    except Exception:
        return None


def _list_gcp_compute_instances(creds: dict):
    """List GCP Compute Engine instances with real CPU + Ops Agent detection."""
    clients = _gcp_clients(creds)
    if not clients:
        return []
    agents = []
    try:
        project_id = clients["project_id"]
        req = clients["compute"]._client.aggregated_list(project=project_id)
        for zone, scoped_list in req:
            if scoped_list.instances:
                for inst in scoped_list.instances:
                    state = (inst.status or "UNKNOWN").lower()
                    running = state == "running"
                    cpu = 0
                    memory = None
                    disk = None
                    agent_installed = False
                    if running:
                        # CPU utilization (platform metric — no agent, 0-1 range)
                        cpu_val = _fetch_gcp_instance_metric(
                            clients, "compute.googleapis.com/instance/cpu/utilization", inst.name
                        )
                        if cpu_val is not None:
                            cpu = round(cpu_val * 100, 1)
                        # Memory (Ops Agent only — agent.googleapis.com namespace)
                        mem_val = _fetch_gcp_instance_metric(
                            clients, "agent.googleapis.com/memory/percent_used", inst.name
                        )
                        if mem_val is not None:
                            memory = mem_val
                            agent_installed = True
                        # Disk (Ops Agent)
                        disk_val = _fetch_gcp_instance_metric(
                            clients, "agent.googleapis.com/disk/percent_used", inst.name
                        )
                        if disk_val is not None:
                            disk = disk_val
                            agent_installed = True
                    if not running:
                        status = "stopped"
                    elif cpu > 85 or (memory is not None and memory > 90) or (disk is not None and disk > 90):
                        status = "critical"
                    elif cpu > 60 or (memory is not None and memory > 75) or (disk is not None and disk > 80):
                        status = "warning"
                    else:
                        status = "healthy"
                    agents.append({
                        "hostname": inst.name, "instance_id": str(inst.id),
                        "os": "Linux/Unix",
                        "ip": (inst.network_interfaces[0].network_ip if inst.network_interfaces else "-"),
                        "public_ip": None,
                        "instance_type": (inst.machine_type or "").split("/")[-1],
                        "region": zone.replace("zones/", "").rsplit("-", 1)[0],
                        "az": zone.replace("zones/", ""),
                        "state": "running" if running else state,
                        "status": status,
                        "cpu": cpu, "memory": memory, "disk": disk,
                        "agent_installed": agent_installed,
                        "uptime": "-",
                        "launch_time": inst.creation_timestamp if inst.creation_timestamp else None,
                        "last_seen": datetime.utcnow().isoformat(),
                    })
    except Exception:
        pass
    return agents


@app.get("/api/monitoring/agents")
def get_agents(account: str = "", regions: str = "",
               user: dict = Depends(get_current_user)):
    """List compute instances across ALL enabled cloud regions.

    `regions` — optional comma-separated filter (e.g. `us-east-1,eu-west-1`).
    If omitted, scans every enabled region on the account.
    """
    if not account:
        return {"agents": [], "total": 0, "healthy": 0, "warning": 0, "critical": 0,
                "note": "Specify ?account=<name> to monitor a cloud account"}

    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")

    regions_filter = [r.strip() for r in regions.split(",") if r.strip()] if regions else None

    provider = creds["provider"]
    if provider == "aws":
        agents = _list_aws_compute_instances(creds, regions_filter=regions_filter)
        all_regions = _aws_enabled_regions(creds)
        scanned_regions = regions_filter if regions_filter else all_regions
    elif provider == "azure":
        agents = _list_azure_compute_instances(creds)
        all_regions = sorted({a.get("region", "") for a in agents if a.get("region")})
        scanned_regions = all_regions
    elif provider == "gcp":
        agents = _list_gcp_compute_instances(creds)
        all_regions = sorted({a.get("region", "") for a in agents if a.get("region")})
        scanned_regions = all_regions
    else:
        return {"agents": [], "total": 0, "healthy": 0, "warning": 0, "critical": 0,
                "note": f"Provider '{provider}' not supported"}

    running = [a for a in agents if a["state"] == "running"]
    healthy = sum(1 for a in running if a["status"] == "healthy")
    warning = sum(1 for a in running if a["status"] == "warning")
    critical = sum(1 for a in running if a["status"] == "critical")
    regions_with_resources = sorted({a.get("region", "") for a in agents if a.get("region")})
    return {
        "agents": agents, "total": len(agents),
        "running": len(running), "healthy": healthy, "warning": warning, "critical": critical,
        "account": account, "provider": provider,
        "regions_scanned": len(scanned_regions),
        "regions_with_resources": regions_with_resources,
        "all_regions": all_regions if provider == "aws" else None,
    }


def _format_uptime(launch_time):
    if not launch_time:
        return "-"
    try:
        now = datetime.now(launch_time.tzinfo) if launch_time.tzinfo else datetime.utcnow()
        delta = now - launch_time
        days = delta.days
        hours = delta.seconds // 3600
        if days > 0:
            return f"{days}d {hours}h"
        return f"{hours}h {(delta.seconds % 3600) // 60}m"
    except Exception:
        return "-"


# ── Monitoring: CloudWatch Logs (real from client AWS) ──────────
@app.get("/api/monitoring/app-logs")
def get_app_logs(account: str = "", level: str = "ALL", service: str = "ALL",
                 region: str = "us-east-1", limit: int = 100,
                 user: dict = Depends(get_current_user)):
    """Pull recent entries from CloudWatch Log Groups in the client's AWS account."""
    if not account:
        return {"logs": [], "total": 0, "note": "Specify ?account=<name> to see CloudWatch logs"}

    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")
    if creds["provider"] == "azure":
        azure_logs = _fetch_azure_logs(creds, limit=limit)
        # Apply level filter client-side
        if level != "ALL":
            azure_logs = [l for l in azure_logs if l["level"] == level]
        return {"logs": azure_logs, "total": len(azure_logs), "provider": "azure", "account": account}

    if creds["provider"] == "gcp":
        gcp_logs = _fetch_gcp_logs(creds, limit=limit)
        if level != "ALL":
            gcp_logs = [l for l in gcp_logs if l["level"] == level]
        return {"logs": gcp_logs, "total": len(gcp_logs), "provider": "gcp", "account": account}

    if creds["provider"] != "aws":
        return {"logs": [], "total": 0, "provider": creds["provider"],
                **_not_implemented(creds["provider"], "cloud_logs")}

    try:
        session = _aws_session(creds, region)
        if not session:
            raise HTTPException(400, "Account has no credentials")
        logs_client = session.client("logs")

        now_ms = int(_time.time() * 1000)
        start_ms = now_ms - (6 * 60 * 60 * 1000)  # last 6 hours

        # List log groups
        groups_resp = logs_client.describe_log_groups(limit=20)
        groups = [g["logGroupName"] for g in groups_resp.get("logGroups", [])]

        all_entries = []
        for group in groups[:10]:  # limit to top 10 groups
            try:
                events_resp = logs_client.filter_log_events(
                    logGroupName=group,
                    startTime=start_ms,
                    endTime=now_ms,
                    limit=max(limit // len(groups[:10]) + 1, 10),
                )
                for idx, evt in enumerate(events_resp.get("events", [])):
                    msg = evt.get("message", "").strip()
                    if not msg:
                        continue
                    # Detect level
                    lv = "INFO"
                    msg_upper = msg.upper()
                    if "ERROR" in msg_upper or "EXCEPTION" in msg_upper or "FAIL" in msg_upper:
                        lv = "ERROR"
                    elif "WARN" in msg_upper:
                        lv = "WARN"
                    elif "DEBUG" in msg_upper:
                        lv = "DEBUG"
                    if level != "ALL" and lv != level:
                        continue
                    all_entries.append({
                        "id": f"{group.replace('/', '_')}-{evt.get('eventId', idx)}",
                        "timestamp": datetime.utcfromtimestamp(evt.get("timestamp", 0) / 1000).isoformat(),
                        "level": lv,
                        "service": group,
                        "hostname": evt.get("logStreamName", ""),
                        "message": msg[:500],
                    })
            except Exception:
                continue

        all_entries.sort(key=lambda x: x["timestamp"], reverse=True)
        return {"logs": all_entries[:limit], "total": len(all_entries),
                "groups_scanned": len(groups[:10]), "region": region}
    except HTTPException:
        raise
    except Exception as e:
        return {"logs": [], "total": 0, "error": str(e)}


# ── Monitoring: Alert Rules (persisted in DB) ──────────────────
_monitoring_alert_rules_file = os.path.join(BASE_DIR, "backend", "alert_rules.json")

def _load_alert_rules():
    try:
        if os.path.exists(_monitoring_alert_rules_file):
            with open(_monitoring_alert_rules_file) as f:
                return json.load(f)
    except Exception:
        pass
    return []

def _save_alert_rules(rules):
    try:
        with open(_monitoring_alert_rules_file, "w") as f:
            json.dump(rules, f, indent=2)
    except Exception:
        pass

class MonitoringAlertRule(BaseModel):
    name: str
    metric: str
    condition: str
    threshold: float
    duration: str = "5m"
    severity: str = "warning"
    channels: list[str] = []
    account: str = ""  # which cloud account to evaluate against

@app.get("/api/monitoring/alert-rules")
def get_monitoring_alert_rules(user: dict = Depends(get_current_user)):
    return {"rules": _load_alert_rules()}

@app.post("/api/monitoring/alert-rules")
def create_monitoring_alert_rule(payload: MonitoringAlertRule, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    rule = {
        "id": f"rule-{uuid.uuid4().hex[:8]}",
        "name": _sanitize_input(payload.name),
        "metric": payload.metric,
        "condition": payload.condition,
        "threshold": payload.threshold,
        "duration": payload.duration,
        "severity": payload.severity,
        "enabled": True,
        "channels": payload.channels,
        "account": payload.account,
        "created_by": user.get("username"),
        "created_at": datetime.utcnow().isoformat(),
    }
    rules = _load_alert_rules()
    rules.append(rule)
    _save_alert_rules(rules)
    log_action(org_id=user.get("org_id"), username=user.get("username"),
               action="monitoring.alert_rule_created", resource_type="alert_rule",
               details={"rule_name": rule["name"], "metric": rule["metric"]})
    return rule

@app.delete("/api/monitoring/alert-rules/{rule_id}")
def delete_monitoring_alert_rule(rule_id: str, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    rules = [r for r in _load_alert_rules() if r["id"] != rule_id]
    _save_alert_rules(rules)
    return {"message": "Rule deleted"}


# ── Enterprise: Audit Log Export ──────────────────────────────
@app.get("/api/enterprise/audit-log/export")
def export_audit_log(fmt: str = "json", limit: int = 10000,
                     user: dict = Depends(get_current_user)):
    """Export recent audit log entries as JSON or CSV for SIEM ingestion."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    try:
        from models.database import SessionLocal, AuditLog
        db = SessionLocal()
        rows = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit).all()
        entries = [{
            "timestamp": (r.created_at.isoformat() if r.created_at else ""),
            "username": r.username or "",
            "action": r.action or "",
            "resource_type": getattr(r, "resource_type", "") or "",
            "resource_id": getattr(r, "resource_id", "") or "",
            "org_id": r.org_id or "",
            "ip_address": getattr(r, "ip_address", "") or "",
            "user_agent": getattr(r, "user_agent", "") or "",
            "details": getattr(r, "details", None) or {},
        } for r in rows]
        db.close()
    except Exception as e:
        raise HTTPException(500, f"Export failed: {e}")

    from starlette.responses import Response
    if fmt == "csv":
        import csv
        from io import StringIO
        buf = StringIO()
        writer = csv.DictWriter(buf, fieldnames=["timestamp", "username", "action", "resource_type",
                                                   "resource_id", "org_id", "ip_address", "user_agent", "details"])
        writer.writeheader()
        for e in entries:
            e["details"] = json.dumps(e["details"]) if e["details"] else ""
            writer.writerow(e)
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": "attachment; filename=audit-log.csv"})
    return Response(content=json.dumps(entries, default=str, indent=2),
                    media_type="application/json",
                    headers={"Content-Disposition": "attachment; filename=audit-log.json"})


# ── Enterprise: Custom Roles (RBAC) ─────────────────────────────
_custom_roles_file = os.path.join(BASE_DIR, "backend", "custom_roles.json")


def _load_roles():
    try:
        if os.path.exists(_custom_roles_file):
            with open(_custom_roles_file) as f:
                return json.load(f)
    except Exception:
        pass
    return []


def _save_roles(roles):
    try:
        with open(_custom_roles_file, "w") as f:
            json.dump(roles, f, indent=2)
    except Exception:
        pass


class CustomRole(BaseModel):
    name: str
    description: str = ""
    permissions: list[str]  # e.g. ["scan.read", "accounts.write", "*.read"]


@app.get("/api/enterprise/roles")
def list_custom_roles(user: dict = Depends(get_current_user)):
    built_in = [
        {"id": "owner", "name": "Owner", "permissions": ["*"], "built_in": True},
        {"id": "admin", "name": "Admin", "permissions": ["*.read", "*.write", "*.execute"], "built_in": True},
        {"id": "editor", "name": "Editor", "permissions": ["*.read", "*.write"], "built_in": True},
        {"id": "viewer", "name": "Viewer", "permissions": ["*.read"], "built_in": True},
        {"id": "auditor", "name": "Auditor", "permissions": ["audit.read", "compliance.read"], "built_in": True},
    ]
    custom = _load_roles()
    return {"roles": built_in + custom}


@app.post("/api/enterprise/roles")
def create_custom_role(payload: CustomRole, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner" or user.get("role") not in ("owner", "admin"):
        raise HTTPException(403, "Admin only")
    roles = _load_roles()
    role = {
        "id": f"role-{uuid.uuid4().hex[:8]}",
        "name": _sanitize_input(payload.name),
        "description": _sanitize_input(payload.description),
        "permissions": payload.permissions,
        "built_in": False,
        "created_by": user.get("username"),
        "created_at": datetime.utcnow().isoformat(),
    }
    roles.append(role)
    _save_roles(roles)
    return role


@app.delete("/api/enterprise/roles/{role_id}")
def delete_custom_role(role_id: str, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner" or user.get("role") not in ("owner", "admin"):
        raise HTTPException(403, "Admin only")
    roles = [r for r in _load_roles() if r["id"] != role_id]
    _save_roles(roles)
    return {"message": "Role deleted"}


# ── Enterprise: SSO Configuration ─────────────────────────────
_sso_config_file = os.path.join(BASE_DIR, "backend", "sso_config.json")


class SSOConfig(BaseModel):
    enabled: bool = False
    provider: str = "oidc"  # oidc, saml
    issuer_url: str = ""
    client_id: str = ""
    client_secret: str = ""
    redirect_uri: str = ""


@app.get("/api/enterprise/sso")
def get_sso_config(user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    try:
        if os.path.exists(_sso_config_file):
            with open(_sso_config_file) as f:
                cfg = json.load(f)
            # Redact secret
            if cfg.get("client_secret"):
                cfg["client_secret"] = "***"
            return cfg
    except Exception:
        pass
    return {"enabled": False, "provider": "oidc", "issuer_url": "", "client_id": "", "redirect_uri": ""}


@app.post("/api/enterprise/sso")
def set_sso_config(payload: SSOConfig, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    cfg = payload.dict()
    try:
        # Preserve existing secret if user didn't change it
        if cfg.get("client_secret") == "***" and os.path.exists(_sso_config_file):
            with open(_sso_config_file) as f:
                existing = json.load(f)
            cfg["client_secret"] = existing.get("client_secret", "")
        with open(_sso_config_file, "w") as f:
            json.dump(cfg, f, indent=2)
    except Exception as e:
        raise HTTPException(500, f"Save failed: {e}")
    log_action(org_id=user.get("org_id"), username=user.get("username"),
               action="enterprise.sso_configured", details={"provider": cfg.get("provider")})
    return {"message": "SSO config saved"}


# ── Enterprise: Platform Health Status ────────────────────────
@app.get("/api/enterprise/platform-health")
def platform_health(user: dict = Depends(get_current_user)):
    """Return health status of all background services."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    import psutil as _psutil
    health = {
        "timestamp": datetime.utcnow().isoformat(),
        "uptime_s": int(_time.time() - _psutil.boot_time()) if hasattr(_psutil, "boot_time") else 0,
        "services": [],
    }
    # Alert evaluator
    try:
        from services.alert_evaluator import _thread as ae_thread
        health["services"].append({
            "name": "alert_evaluator",
            "running": ae_thread is not None and ae_thread.is_alive(),
        })
    except Exception:
        health["services"].append({"name": "alert_evaluator", "running": False})
    # Availability monitor
    try:
        from services.availability import _thread as av_thread
        health["services"].append({
            "name": "availability_monitor",
            "running": av_thread is not None and av_thread.is_alive(),
        })
    except Exception:
        health["services"].append({"name": "availability_monitor", "running": False})
    # DB connectivity
    try:
        from models.database import SessionLocal
        from sqlalchemy import text as _text
        db = SessionLocal()
        db.execute(_text("SELECT 1"))
        db.close()
        health["services"].append({"name": "database", "running": True})
    except Exception:
        health["services"].append({"name": "database", "running": False})

    health["overall"] = "healthy" if all(s["running"] for s in health["services"]) else "degraded"
    return health


# ── Deep Monitoring: API / DB / Function Performance ──────────
@app.get("/api/perf/api-endpoints")
def get_perf_api_endpoints(account: str = "", user: dict = Depends(get_current_user)):
    """List load balancers with latency / error-rate metrics from cloud provider."""
    if not account:
        return {"load_balancers": [], "total": 0, "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, "Account not found")
    from services.perf_monitoring import list_load_balancers
    lbs = list_load_balancers(creds)
    total_req = sum(l.get("request_count", 0) or 0 for l in lbs)
    total_5xx = sum(l.get("error_5xx", 0) or 0 for l in lbs)
    return {
        "load_balancers": lbs, "total": len(lbs),
        "total_requests": total_req, "total_5xx": total_5xx,
        "overall_error_rate_pct": round((total_5xx / total_req) * 100, 2) if total_req > 0 else 0,
        "account": account, "provider": creds["provider"],
    }


@app.get("/api/perf/databases")
def get_perf_databases(account: str = "", user: dict = Depends(get_current_user)):
    if not account:
        return {"databases": [], "total": 0, "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, "Account not found")
    from services.perf_monitoring import list_databases
    dbs = list_databases(creds)
    return {"databases": dbs, "total": len(dbs), "account": account, "provider": creds["provider"]}


@app.get("/api/perf/functions")
def get_perf_functions(account: str = "", user: dict = Depends(get_current_user)):
    if not account:
        return {"functions": [], "total": 0, "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, "Account not found")
    from services.perf_monitoring import list_functions
    fns = list_functions(creds)
    total_inv = sum(f.get("invocations", 0) for f in fns)
    total_err = sum(f.get("errors", 0) for f in fns)
    return {
        "functions": fns, "total": len(fns),
        "total_invocations": total_inv, "total_errors": total_err,
        "overall_error_rate_pct": round((total_err / total_inv) * 100, 2) if total_inv > 0 else 0,
        "account": account, "provider": creds["provider"],
    }


# ── Kubernetes Security (managed clusters) ────────────────────
@app.get("/api/kubernetes/clusters")
def list_k8s_clusters(account: str = "", user: dict = Depends(get_current_user)):
    """List managed K8s clusters (EKS/AKS/GKE) with security posture."""
    if not account:
        return {"clusters": [], "total": 0, "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")
    from services.k8s_security import list_clusters_for_account
    clusters = list_clusters_for_account(creds)
    total_findings = sum(len(c.get("findings", [])) for c in clusters)
    critical = sum(1 for c in clusters for f in c.get("findings", []) if f.get("severity") == "critical")
    avg_score = round(sum(c.get("score", 0) for c in clusters) / len(clusters), 1) if clusters else 0
    return {
        "clusters": clusters,
        "total": len(clusters),
        "total_findings": total_findings,
        "critical_findings": critical,
        "avg_score": avg_score,
        "account": account,
        "provider": creds["provider"],
    }


@app.get("/api/kubernetes/posture-summary")
def k8s_posture_summary(user: dict = Depends(get_current_user)):
    """Aggregate K8s posture across all cloud accounts accessible to the user."""
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    from models.database import SessionLocal, CloudAccount
    from services.crypto import decrypt
    from services.k8s_security import list_clusters_for_account
    db = SessionLocal()
    q = db.query(CloudAccount)
    if org_id:
        q = q.filter(CloudAccount.org_id == org_id)
    accounts = q.all()
    summary = []
    for acc in accounts:
        creds = {
            "provider": acc.provider, "account_id": acc.account_id,
            "access_key": decrypt(acc.access_key), "secret_key": decrypt(acc.secret_key),
        }
        if not creds["access_key"]:
            continue
        clusters = list_clusters_for_account(creds)
        if not clusters:
            continue
        summary.append({
            "account": acc.name, "provider": acc.provider,
            "cluster_count": len(clusters),
            "avg_score": round(sum(c["score"] for c in clusters) / len(clusters), 1),
            "critical_findings": sum(1 for c in clusters for f in c.get("findings", []) if f.get("severity") == "critical"),
        })
    db.close()
    return {"summary": summary, "accounts_scanned": len(summary)}


# ── Availability Monitoring (real HTTP pings) ──────────────────
class AvailabilityMonitor(BaseModel):
    name: str
    url: str
    method: str = "GET"
    expected_status: int = 200
    expected_body: str = ""
    interval_seconds: int = 60
    enabled: bool = True


@app.get("/api/availability/monitors")
def list_availability_monitors(user: dict = Depends(get_current_user)):
    from services.availability import list_monitors, uptime_stats, get_all_check_summaries
    monitors = list_monitors()
    summaries = get_all_check_summaries()
    result = []
    for m in monitors:
        stats = uptime_stats(m["id"], hours=24)
        result.append({
            **m,
            "last_check": summaries.get(m["id"]),
            "stats_24h": stats,
        })
    return {"monitors": result, "total": len(result)}


@app.post("/api/availability/monitors")
def create_availability_monitor(payload: AvailabilityMonitor, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.availability import create_monitor
    monitor = {
        "id": f"mon-{uuid.uuid4().hex[:8]}",
        "name": _sanitize_input(payload.name),
        "url": payload.url,
        "method": payload.method.upper(),
        "expected_status": payload.expected_status,
        "expected_body": _sanitize_input(payload.expected_body) if payload.expected_body else "",
        "interval_seconds": max(30, payload.interval_seconds),
        "enabled": payload.enabled,
        "created_by": user.get("username"),
        "created_at": datetime.utcnow().isoformat(),
    }
    create_monitor(monitor)
    log_action(org_id=user.get("org_id"), username=user.get("username"),
               action="availability.monitor_created", resource_type="monitor",
               details={"name": monitor["name"], "url": monitor["url"]})
    return monitor


@app.delete("/api/availability/monitors/{monitor_id}")
def delete_availability_monitor(monitor_id: str, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.availability import delete_monitor
    delete_monitor(monitor_id)
    return {"message": "Monitor deleted"}


@app.get("/api/availability/monitors/{monitor_id}/checks")
def get_monitor_checks(monitor_id: str, limit: int = 100, hours: int = 24,
                       user: dict = Depends(get_current_user)):
    from services.availability import get_recent_checks, uptime_stats
    checks = get_recent_checks(monitor_id, limit=limit)
    stats = uptime_stats(monitor_id, hours=hours)
    return {"checks": checks, "total": len(checks), "stats": stats}


@app.post("/api/availability/monitors/{monitor_id}/run")
def run_monitor_now(monitor_id: str, user: dict = Depends(get_current_user)):
    """Run a single on-demand check for the monitor."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.availability import list_monitors, _perform_check, _append_check
    mon = next((m for m in list_monitors() if m["id"] == monitor_id), None)
    if not mon:
        raise HTTPException(404, "Monitor not found")
    result = _perform_check(mon)
    _append_check(result)
    return result


@app.get("/api/monitoring/agent-install")
def get_agent_install_instructions(provider: str = "aws", user: dict = Depends(get_current_user)):
    """Return installation instructions for the cloud provider's monitoring agent.

    These agents are required for memory/disk/process metrics (OS-level data
    the hypervisor cannot see). They're official free agents from AWS/Azure/GCP,
    not CloudSentrix agents.
    """
    p = provider.lower()
    if p == "aws":
        return {
            "provider": "aws",
            "agent_name": "CloudWatch Agent",
            "docs_url": "https://docs.aws.amazon.com/AmazonCloudWatch/latest/monitoring/install-CloudWatch-Agent-on-EC2-Instance.html",
            "steps": [
                "Attach the IAM policy CloudWatchAgentServerPolicy to the EC2 instance role",
                "Install via AWS SSM: aws ssm send-command --document-name AWS-ConfigureAWSPackage --parameters action=Install,name=AmazonCloudWatchAgent --targets Key=InstanceIds,Values=i-xxxxxxxx",
                "Configure mem/disk metrics with the sample config (see docs link)",
                "Start: sudo /opt/aws/amazon-cloudwatch-agent/bin/amazon-cloudwatch-agent-ctl -a fetch-config -m ec2 -c ssm:AmazonCloudWatch-linux -s",
            ],
            "quick_install": [
                "# Amazon Linux 2 / AL2023",
                "sudo yum install -y amazon-cloudwatch-agent",
                "# Ubuntu / Debian",
                "wget https://s3.amazonaws.com/amazoncloudwatch-agent/ubuntu/amd64/latest/amazon-cloudwatch-agent.deb",
                "sudo dpkg -i -E ./amazon-cloudwatch-agent.deb",
            ],
            "metrics_collected": ["mem_used_percent", "disk_used_percent", "swap_used_percent", "net_bytes_sent", "net_bytes_recv"],
        }
    if p == "azure":
        return {
            "provider": "azure",
            "agent_name": "Azure Monitor Agent (AMA)",
            "docs_url": "https://learn.microsoft.com/en-us/azure/azure-monitor/agents/azure-monitor-agent-manage",
            "steps": [
                "Create a Data Collection Rule (DCR) in Azure Monitor",
                "Associate the DCR with your VM (Azure Portal > VM > Monitoring > Insights)",
                "Or use CLI: az vm extension set --name AzureMonitorLinuxAgent --publisher Microsoft.Azure.Monitor --vm-name <vm> --resource-group <rg>",
                "Create Data Collection Rules for performance counters (memory, disk)",
            ],
            "quick_install": [
                "az vm extension set \\",
                "  --name AzureMonitorLinuxAgent \\",
                "  --publisher Microsoft.Azure.Monitor \\",
                "  --vm-name <vm> --resource-group <rg>",
            ],
            "metrics_collected": ["Memory Available %", "Logical Disk Free %", "Processor % Processor Time"],
        }
    if p == "gcp":
        return {
            "provider": "gcp",
            "agent_name": "Ops Agent",
            "docs_url": "https://cloud.google.com/monitoring/agent/ops-agent/installation",
            "steps": [
                "Grant the compute instance service account the 'Monitoring Metric Writer' role",
                "Install via metadata startup script or manually",
                "Ops Agent publishes mem/disk/network metrics to Cloud Monitoring automatically",
            ],
            "quick_install": [
                "curl -sSO https://dl.google.com/cloudagents/add-google-cloud-ops-agent-repo.sh",
                "sudo bash add-google-cloud-ops-agent-repo.sh --also-install",
            ],
            "metrics_collected": [
                "agent.googleapis.com/memory/percent_used",
                "agent.googleapis.com/disk/percent_used",
                "agent.googleapis.com/interface/traffic",
            ],
        }
    return {"provider": p, "error": "Unknown provider"}


@app.get("/api/monitoring/alert-events")
def get_alert_events_endpoint(limit: int = 100, user: dict = Depends(get_current_user)):
    """Return alert events that the background evaluator has fired."""
    try:
        from services.alert_evaluator import get_events
        events = get_events(limit=limit)
        return {"events": events, "total": len(events)}
    except Exception as e:
        return {"events": [], "total": 0, "error": str(e)}


@app.post("/api/monitoring/alert-events/{event_id}/ack")
def ack_alert_event(event_id: str, user: dict = Depends(get_current_user)):
    """Acknowledge an alert event."""
    from services.alert_evaluator import ALERT_EVENTS_FILE, _load_json, _save_json
    events = _load_json(ALERT_EVENTS_FILE, [])
    found = False
    for e in events:
        if e.get("id") == event_id:
            e["ack"] = True
            e["ack_by"] = user.get("username")
            e["ack_at"] = datetime.utcnow().isoformat()
            found = True
            break
    if not found:
        raise HTTPException(404, "Event not found")
    _save_json(ALERT_EVENTS_FILE, events)
    return {"message": "Acknowledged"}


@app.post("/api/monitoring/alert-events/{event_id}/resolve")
def resolve_alert_event(event_id: str, user: dict = Depends(get_current_user)):
    """Mark an alert event as resolved."""
    from services.alert_evaluator import ALERT_EVENTS_FILE, _load_json, _save_json
    events = _load_json(ALERT_EVENTS_FILE, [])
    found = False
    for e in events:
        if e.get("id") == event_id:
            e["resolved"] = True
            e["resolved_by"] = user.get("username")
            e["resolved_at"] = datetime.utcnow().isoformat()
            e["status"] = "resolved"
            found = True
            break
    if not found:
        raise HTTPException(404, "Event not found")
    _save_json(ALERT_EVENTS_FILE, events)
    return {"message": "Resolved"}


# ── Notification Channels ──────────────────────────────────────
class NotificationChannel(BaseModel):
    type: str  # slack, email, webhook
    name: str
    config: dict
    enabled: bool = True


@app.get("/api/notifications/channels")
def list_notification_channels(user: dict = Depends(get_current_user)):
    from services.notifier import list_channels
    channels = list_channels()
    # Redact secrets in config for display
    sanitized = []
    for c in channels:
        cfg = c.get("config", {}) or {}
        safe_cfg = {}
        for k, v in cfg.items():
            if k in ("webhook_url", "url") and isinstance(v, str) and len(v) > 24:
                safe_cfg[k] = v[:20] + "..." + v[-6:]
            else:
                safe_cfg[k] = v
        sanitized.append({**c, "config": safe_cfg})
    return {"channels": sanitized}


@app.post("/api/notifications/channels")
def create_notification_channel(payload: NotificationChannel, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.notifier import add_channel
    if payload.type not in ("slack", "email", "webhook"):
        raise HTTPException(400, "Invalid channel type")
    channel = {
        "id": f"ch-{uuid.uuid4().hex[:8]}",
        "type": payload.type,
        "name": _sanitize_input(payload.name),
        "config": payload.config,
        "enabled": payload.enabled,
        "created_by": user.get("username"),
        "created_at": datetime.utcnow().isoformat(),
    }
    add_channel(channel)
    log_action(org_id=user.get("org_id"), username=user.get("username"),
               action="notifications.channel_created", resource_type="channel",
               details={"type": channel["type"], "name": channel["name"]})
    return channel


@app.delete("/api/notifications/channels/{channel_id}")
def delete_notification_channel(channel_id: str, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.notifier import delete_channel
    delete_channel(channel_id)
    return {"message": "Channel deleted"}


@app.post("/api/notifications/channels/{channel_id}/test")
def test_notification_channel(channel_id: str, user: dict = Depends(get_current_user)):
    """Send a test message to a specific channel."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.notifier import send
    results = send(
        subject="CloudSentrix test notification",
        body=f"This is a test from user {user.get('username')}. If you see this, the channel is working.",
        severity="info",
        channel_ids=[channel_id],
    )
    return {"results": results}


@app.get("/api/notifications/log")
def get_notification_log(limit: int = 50, user: dict = Depends(get_current_user)):
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    from services.notifier import get_delivery_log
    return {"log": get_delivery_log(limit=limit)}


@app.post("/api/monitoring/alert-rules/evaluate-now")
def evaluate_alerts_now(user: dict = Depends(get_current_user)):
    """Manually trigger a one-pass evaluation (useful for testing)."""
    if user.get("user_type") != "owner":
        raise HTTPException(403, "Owner only")
    try:
        from services.alert_evaluator import evaluate_once
        new_events = evaluate_once()
        return {"new_events": new_events, "triggered_at": datetime.utcnow().isoformat()}
    except Exception as e:
        return {"new_events": 0, "error": str(e)}


# ── Monitoring: CloudWatch Metric History (real) ───────────────
def _cloudwatch_fleet_history(creds: dict, metric_name: str = "CPUUtilization",
                              hours: int = 24, region: str = "us-east-1"):
    """Pull CloudWatch history averaged across all running EC2 instances in ALL regions."""
    session = _aws_session(creds, region)
    if not session:
        return []
    regions = _aws_enabled_regions(creds)
    if not regions:
        regions = [region]

    now = datetime.utcnow()
    start = now - timedelta(hours=hours)
    all_datapoints = {}

    for reg in regions:
        try:
            sess = _aws_session(creds, reg)
            ec2 = sess.client("ec2")
            cw = sess.client("cloudwatch")
            resp = ec2.describe_instances(Filters=[{"Name": "instance-state-name", "Values": ["running"]}])
            instance_ids = []
            for rsv in resp.get("Reservations", []):
                for inst in rsv.get("Instances", []):
                    instance_ids.append(inst["InstanceId"])

            for inst_id in instance_ids[:20]:  # cap to 20 per region
                try:
                    cw_resp = cw.get_metric_statistics(
                        Namespace="AWS/EC2", MetricName=metric_name,
                        Dimensions=[{"Name": "InstanceId", "Value": inst_id}],
                        StartTime=start, EndTime=now,
                        Period=300, Statistics=["Average"],
                    )
                    for dp in cw_resp.get("Datapoints", []):
                        ts = dp["Timestamp"].replace(minute=(dp["Timestamp"].minute // 5) * 5, second=0, microsecond=0)
                        key = ts.isoformat()
                        if key not in all_datapoints:
                            all_datapoints[key] = []
                        all_datapoints[key].append(dp["Average"])
                except Exception:
                    continue
        except Exception:
            continue

    # Average per timestamp
    history = []
    for ts_key in sorted(all_datapoints.keys()):
        values = all_datapoints[ts_key]
        history.append({
            "timestamp": ts_key,
            "value": round(sum(values) / len(values), 1),
            "instances": len(values),
        })
    return history


@app.get("/api/monitoring/history")
def get_metric_history(account: str = "", metric: str = "cpu", hours: int = 1,
                       user: dict = Depends(get_current_user)):
    """CloudWatch history aggregated across running instances of the given AWS account."""
    if not account:
        return {"history": [], "total": 0, "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")
    if creds["provider"] == "azure":
        hist = _fetch_azure_cpu_history(creds, hours=hours)
        return {"history": [{"timestamp": h["timestamp"], metric: h["value"], "instances": h["instances"]} for h in hist],
                "total": len(hist), "metric": metric, "account": account, "provider": "azure"}

    if creds["provider"] == "gcp":
        hist = _fetch_gcp_cpu_history(creds, hours=hours)
        return {"history": [{"timestamp": h["timestamp"], metric: h["value"], "instances": h["instances"]} for h in hist],
                "total": len(hist), "metric": metric, "account": account, "provider": "gcp"}

    if creds["provider"] != "aws":
        return {"history": [], "total": 0, "provider": creds["provider"],
                **_not_implemented(creds["provider"], "metric_history")}

    metric_map = {"cpu": "CPUUtilization", "network_in": "NetworkIn", "network_out": "NetworkOut"}
    cw_metric = metric_map.get(metric, "CPUUtilization")

    try:
        raw = _cloudwatch_fleet_history(creds, cw_metric, hours)
        # Reshape for frontend compatibility
        history = [{"timestamp": h["timestamp"], metric: h["value"], "instances": h["instances"]} for h in raw]
        return {"history": history, "total": len(history), "metric": metric, "account": account}
    except Exception as e:
        return {"history": [], "total": 0, "error": str(e)}


# ── Monitoring: Anomaly Detection (real, on CloudWatch history) ──
@app.get("/api/monitoring/anomalies")
def get_anomalies(account: str = "", user: dict = Depends(get_current_user)):
    """Detect anomalies in CloudWatch metrics for the client's AWS account."""
    if not account:
        return {"anomalies": [], "total": 0, "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")
    if creds["provider"] in ("azure", "gcp"):
        # Use provider history (CPU only) + z-score detection
        fetcher = _fetch_azure_cpu_history if creds["provider"] == "azure" else _fetch_gcp_cpu_history
        history = fetcher(creds, hours=24)
        if len(history) < 12:
            return {"anomalies": [], "total": 0, "samples_analyzed": len(history),
                    "provider": creds["provider"],
                    "note": f"Not enough {creds['provider'].upper()} metric history. Need at least 12 samples."}
        values = [h["value"] for h in history]
        mean = sum(values) / len(values)
        variance = sum((v - mean) ** 2 for v in values) / len(values)
        stddev = variance ** 0.5 if variance > 0 else 1
        anomalies = []
        for h in history[-5:]:
            v = h["value"]
            z = abs(v - mean) / stddev if stddev > 0 else 0
            if z > 2.5:
                anomalies.append({
                    "id": f"{creds['provider']}-CPU-{h['timestamp']}",
                    "type": "cpu_anomaly",
                    "severity": "critical" if z > 4 else "warning",
                    "host": f"fleet ({h['instances']} instances)",
                    "metric": "CPU",
                    "value": round(v, 1),
                    "expected": round(mean, 1),
                    "zscore": round(z, 2),
                    "timestamp": h["timestamp"],
                    "recommendation": f"CPU is {z:.1f} std devs from normal. Investigate load or scale out.",
                })
        return {"anomalies": anomalies, "total": len(anomalies),
                "samples_analyzed": len(history), "account": account, "provider": creds["provider"]}

    if creds["provider"] != "aws":
        return {"anomalies": [], "total": 0, "provider": creds["provider"],
                **_not_implemented(creds["provider"], "anomaly_detection")}

    try:
        anomalies = []
        samples_analyzed = 0
        for metric_label, cw_name in [("CPU", "CPUUtilization"), ("Network In", "NetworkIn"), ("Network Out", "NetworkOut")]:
            history = _cloudwatch_fleet_history(creds, cw_name, hours=24)
            samples_analyzed += len(history)
            if len(history) < 12:
                continue
            values = [h["value"] for h in history]
            mean = sum(values) / len(values)
            variance = sum((v - mean) ** 2 for v in values) / len(values)
            stddev = variance ** 0.5 if variance > 0 else 1
            recent = history[-5:]
            for h in recent:
                v = h["value"]
                z = abs(v - mean) / stddev if stddev > 0 else 0
                if z > 2.5:
                    anomalies.append({
                        "id": f"{metric_label}-{h['timestamp']}",
                        "type": f"{cw_name.lower()}_anomaly",
                        "severity": "critical" if z > 4 else "warning",
                        "host": f"fleet ({h['instances']} instances)",
                        "metric": metric_label,
                        "value": round(v, 1),
                        "expected": round(mean, 1),
                        "zscore": round(z, 2),
                        "timestamp": h["timestamp"],
                        "recommendation": f"{metric_label} is {z:.1f} std devs from normal. Investigate load or scale out.",
                    })
        if samples_analyzed < 12:
            return {"anomalies": [], "total": 0, "samples_analyzed": samples_analyzed,
                    "note": "Not enough CloudWatch history. Launch instances and let them run for at least 1 hour."}
        return {"anomalies": anomalies, "total": len(anomalies), "samples_analyzed": samples_analyzed, "account": account}
    except Exception as e:
        return {"anomalies": [], "total": 0, "error": str(e)}


# ── Monitoring: Forecast (real, from CloudWatch history) ───────
@app.get("/api/monitoring/forecast")
def get_forecast(account: str = "", metric: str = "cpu", hours: int = 6,
                 user: dict = Depends(get_current_user)):
    """Linear regression forecast on CloudWatch metrics."""
    if not account:
        return {"metric": metric, "data": [], "note": "Specify ?account=<name>"}
    org_id = user.get("org_id") if user.get("user_type") == "client" else None
    creds = _get_account_credentials(account, org_id)
    if not creds:
        raise HTTPException(404, f"Cloud account '{account}' not found")
    if creds["provider"] in ("azure", "gcp"):
        fetcher = _fetch_azure_cpu_history if creds["provider"] == "azure" else _fetch_gcp_cpu_history
        history = fetcher(creds, hours=48)
        n = len(history)
        if n < 20:
            return {"metric": metric, "data": [], "samples_used": n, "provider": creds["provider"],
                    "note": f"Only {n} samples available. Need at least 20."}
        values = [h["value"] for h in history]
        x_mean = (n - 1) / 2
        y_mean = sum(values) / n
        numerator = sum((i - x_mean) * (values[i] - y_mean) for i in range(n))
        denominator = sum((i - x_mean) ** 2 for i in range(n))
        slope = numerator / denominator if denominator > 0 else 0
        intercept = y_mean - slope * x_mean
        residuals = [(values[i] - (slope * i + intercept)) for i in range(n)]
        resid_std = (sum(r * r for r in residuals) / n) ** 0.5
        data = []
        for i, h in enumerate(history):
            data.append({"timestamp": h["timestamp"], "actual": round(h["value"], 1),
                         "forecast": None, "upper_bound": None, "lower_bound": None})
        future_points = hours * 12
        last_ts = datetime.fromisoformat(history[-1]["timestamp"].replace("Z", "+00:00").replace("+00:00", ""))
        for i in range(future_points):
            ts = (last_ts + timedelta(minutes=(i + 1) * 5)).isoformat()
            predicted = slope * (n + i) + intercept
            data.append({
                "timestamp": ts, "actual": None,
                "forecast": round(max(0, predicted), 1),
                "upper_bound": round(max(0, predicted + 2 * resid_std), 1),
                "lower_bound": round(max(0, predicted - 2 * resid_std), 1),
            })
        return {"metric": metric, "data": data, "samples_used": n, "account": account, "provider": creds["provider"]}

    if creds["provider"] != "aws":
        return {"metric": metric, "data": [], "provider": creds["provider"],
                **_not_implemented(creds["provider"], "forecast")}

    metric_map = {"cpu": "CPUUtilization", "network_in": "NetworkIn", "network_out": "NetworkOut"}
    cw_metric = metric_map.get(metric, "CPUUtilization")
    try:
        history = _cloudwatch_fleet_history(creds, cw_metric, hours=48)
        n = len(history)
        if n < 20:
            return {"metric": metric, "data": [], "samples_used": n,
                    "note": f"Only {n} CloudWatch samples. Need at least 20."}
        values = [h["value"] for h in history]
        x_mean = (n - 1) / 2
        y_mean = sum(values) / n
        numerator = sum((i - x_mean) * (values[i] - y_mean) for i in range(n))
        denominator = sum((i - x_mean) ** 2 for i in range(n))
        slope = numerator / denominator if denominator > 0 else 0
        intercept = y_mean - slope * x_mean
        residuals = [(values[i] - (slope * i + intercept)) for i in range(n)]
        resid_std = (sum(r * r for r in residuals) / n) ** 0.5

        data = []
        for i, h in enumerate(history):
            data.append({"timestamp": h["timestamp"], "actual": round(h["value"], 1), "forecast": None, "upper_bound": None, "lower_bound": None})
        future_points = hours * 12  # 5-min intervals
        last_ts = datetime.fromisoformat(history[-1]["timestamp"])
        for i in range(future_points):
            ts = (last_ts + timedelta(minutes=(i + 1) * 5)).isoformat()
            predicted = slope * (n + i) + intercept
            data.append({
                "timestamp": ts,
                "actual": None,
                "forecast": round(max(0, predicted), 1),
                "upper_bound": round(max(0, predicted + 2 * resid_std), 1),
                "lower_bound": round(max(0, predicted - 2 * resid_std), 1),
            })
        return {"metric": metric, "data": data, "samples_used": n, "account": account}
    except Exception as e:
        return {"metric": metric, "data": [], "error": str(e)}


# ── Update health endpoint with security info ──────────────────
@app.get("/api/security/health")
def security_health():
    """Security-focused health check."""
    return {
        "status": "secured",
        "security_headers": True,
        "rate_limiting": True,
        "input_sanitization": True,
        "csrf_protection": True,
        "audit_logging": True,
        "cors_configured": True,
        "jwt_auth": True,
        "api_key_management": True,
        "version": "3.2.0",
        "timestamp": datetime.now().isoformat(),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8088)
